"""build_control_logic.py -- per-controller control logic / Sequence of Operations.

Builds one structured control-logic model for every DDC controller (panel ->
controller -> equipment -> I/O points), composing a Sequence of Operations from
equipment-class templates that reference the controller's actual point tags. The
strategy is consistent with the implemented engine in backend/control.py
(R-1 economic dispatch, wet-bulb CW reset, CHW trim-and-respond) and sim.py.

Renders to:
  docs/control_logic.md                 full narrative, panel -> controller
  exports/red5-dhcp_control-logic.html  interactive (search, print per controller)
  exports/red5-dhcp_control-logic.xlsx  Index + one sheet per controller (SOO + points)
  exports/red5-dhcp_control-logic.csv   flat: Panel, Controller, Equipment, Strategy, SOO
  canvas dhcp-control-logic (in-IDE)     panel -> controller -> SOO
"""
from __future__ import annotations

import csv
import datetime
import html
import json
import os

import generate_io_list as g

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(HERE, "exports")
DOCS_DIR = os.path.join(HERE, "docs")
CANVAS_DIR = "/Users/jinkim/.cursor/projects/Users-jinkim-CURSOR/canvases"
TODAY = datetime.date.today().isoformat()

PANEL_META = {p[0]: (p[1], p[2]) for p in g.PANELS}


# ---------------------------------------------------------------------------
# Equipment classification
# ---------------------------------------------------------------------------
def classify(dev: str) -> str:
    d = dev.upper()
    if d.startswith("R-1"):
        return "chiller"
    if d.startswith("DHC-CHW"):
        return "dhc_chw"
    if d.startswith("DHC-STEAM"):
        return "dhc_steam"
    if d.startswith("HWT"):
        return "hotwell"
    if d.startswith(("HP-3", "HP-5")):
        return "hotwell_pump"
    if d.startswith(("HP-1", "HP-2")):
        return "hw_pump"
    if d.startswith("HP-4"):
        return "sec_pump"
    if d.startswith("CDP"):
        return "cond_pump"
    if d.startswith("CP-"):
        return "chw_pump"
    if d.startswith("CT-"):
        return "tower"
    if d.startswith("CHGV"):
        return "changeover"
    if d.startswith("EX-1"):
        return "freecool"
    if d.startswith("EMHX"):
        return "emhx"
    if d.startswith(("EX-2", "EX-3", "EX4", "HEX", "EX-")):
        return "hx"
    if d.startswith("EXT"):
        return "expansion"
    if d.startswith("AC"):
        return "ahu"
    if d.startswith(("OAU", "EVU", "SEV", "SA", "EA", "EF", "SF", "FAN")):
        return "vent"
    if d.startswith("DHW-"):
        return "dhw"
    if d.startswith(("MP-", "P-", "DP-", "BL-", "RCVT", "ELVT", "REUSET", "FIRET")):
        return "sanitary"
    if d.startswith("FILT-"):
        return "filtration"
    if d.startswith(("SMF-", "SMK-")):
        return "smoke"
    if d.startswith("ST-"):
        return "oa_station"
    if d.startswith("FCU"):
        return "fcu"
    if d.startswith(("PAC", "PCU", "PMAC")):
        return "packaged"
    if d.startswith("LTG"):
        return "lighting"
    if d.startswith("MTR"):
        return "meter"
    return "generic"


CLASS_LABEL = {
    "chiller": "Water-cooled screw chiller (R-1, local backup)",
    "dhc_chw": "DHC chilled-water intake (primary source)",
    "dhc_steam": "DHC steam intake / PRV station",
    "hotwell": "Hot-well / condensate tank",
    "hotwell_pump": "Hot-well & condensate-return pumps",
    "hw_pump": "Hot-water distribution pumps",
    "sec_pump": "Secondary (buffer-loop) distribution pumps",
    "cond_pump": "Condenser-water pumps",
    "chw_pump": "Chilled-water distribution pumps",
    "tower": "Cooling tower (INV fan cells)",
    "changeover": "Source-changeover valve",
    "freecool": "Winter free-cooling heat exchanger",
    "emhx": "Emergency cooling heat exchanger",
    "hx": "Plate heat exchanger (heating/tempered loop)",
    "expansion": "Expansion / pressurisation set",
    "ahu": "Air-handling unit",
    "vent": "Ventilation / exhaust fan",
    "oa_station": "Outdoor-air station",
    "fcu": "Fan-coil units",
    "packaged": "Packaged DX unit",
    "lighting": "Lighting group",
    "meter": "Energy meter",
    "dhw": "Domestic-hot-water plant (steam-heated storage)",
    "sanitary": "Plumbing / sanitary pumps & tanks",
    "filtration": "Recreational-water filtration plant",
    "smoke": "Smoke-exhaust (排煙) fans",
    "generic": "Field device",
}


# ---------------------------------------------------------------------------
# SOO templates -> list[str] bullets. `devs` = member device tags of this class.
# `sfx(dev)` returns the set of point suffixes present for that device.
# ---------------------------------------------------------------------------
def _n(devs):
    return len(devs)


def _lead_lag(devs, what):
    if len(devs) <= 1:
        return []
    return [f"Lead/lag/standby across {len(devs)} {what} ({', '.join(devs)}): stage on "
            f"demand, rotate lead on equal run-hours, prove each start before staging the next."]


def soo_chiller(devs, sfx):
    d = devs[0]
    return [
        (f"Purpose", [
            f"{d} is the local water-cooled twin-screw chiller (Ebara RHS DW202M2, 370 kW, "
            f"COP 4.5) that backs up / peak-shaves the DHC supply on the 36/37F loop."]),
        ("Economic dispatch (DHC vs R-1)", [
            "DHC is the default source. Stage the chiller only when its marginal cost beats DHC "
            "by >5% or DHC is unavailable: run when `DHC_cost > R-1_cost x 1.05` and loop PLR > 0.15.",
            f"Enable via `{d}.SS`; trim capacity with `{d}.DMD`. During the DHC coincident-peak "
            "window the demand adder pushes DHC above R-1 -> peak-shave mode.",
            "Bumpless changeover through CHGV-1/2 (branch vs through-main); no compressor start "
            "without both evaporator and condenser flow proven."]),
        ("Capacity, staging & resets", [
            "Twin 45 kW screws: one up to ~55% loop PLR, both above; balanced load-share; "
            "anti-short-cycle min-on/off; rotate lead daily.",
            f"CHW reset (trim-and-respond) 7->10 C to most-open coil; hold via `{d}.CHWST` "
            f"(return `{d}.CHWRT`). Condenser-water reset = wet-bulb + 3.5 C, floored at 19 C "
            f"(R-407C min-lift), tracked on `{d}.CWRT` via CT-3/CDP-3."]),
        ("Safeties, backup & monitoring", [
            f"Flow-proof + freeze + hi-discharge/low-suction interlocks; trip on `{d}.TRIP`; "
            f"honour `{d}.LR` (local/remote).",
            "Weekly readiness exercise (Mon 10:00) when otherwise idle; live COP / kW-per-RT and "
            "evap/cond approach for FDD (low-DT, condenser fouling)."]),
    ]


def soo_dhc_chw(devs, sfx):
    d = devs[0]
    return [
        ("Purpose", [
            "Primary cooling source: modulate the DHC intake so the building CHW header / thermal "
            "demand is met at minimum district energy + demand cost."]),
        ("Control", [
            f"Modulate intake control valve `{d}.CV` (feedback `{d}.CVFB`) to hold the CHW supply "
            f"setpoint on `{d}.CS-T` (return `{d}.CR-T`); cap opening during the district peak window.",
            f"Meter thermal energy `{d}.GJ` and flow `{d}.FLW` for demand limiting; coordinate "
            "changeover with R-1 (CHGV-1/2)."]),
        ("Safeties & monitoring", [
            f"High/low supply-temp and header-pressure alarms (`{d}.CS-T`, `{d}.CS-P`); "
            "low-flow / valve-fault supervision; open the valve on loss of BMS (fail-safe cooling)."]),
    ]


def soo_dhc_steam(devs, sfx):
    d = devs[0]
    return [
        ("Purpose", ["DHC steam intake: reduce 0.8 MPa district steam to the 0.2 MPa house header "
                     "and meter mass flow + condensate return."]),
        ("Control", [
            f"PRV holds the 0.2 MPa header setpoint; meter steam mass `{d}.MASS` and condensate "
            f"hot-return `{d}.HR-M` back to the hot-well.",
            "Enable in heating season only; isolate and drain in cooling season (Apr–May off)."]),
        ("Safeties", ["Over-pressure relief, high-header-pressure alarm, condensate-level and "
                      "trap supervision."]),
    ]


def soo_pumps(devs, sfx, what, header="header differential pressure", extra=None):
    b = [(f"Purpose", [f"{CLASS_LABEL_HINT.get(what, what)}: maintain loop flow for the served load."])]
    seq = [f"Start/stop on system demand via `{devs[0]}.SS`" + (", one pump at a time with flow/Δp proof." if _n(devs) > 1 else ".")]
    seq += _lead_lag(devs, "pumps")
    b.append(("Sequencing", seq))
    mod = []
    if any("SPD" in sfx(d) for d in devs):
        mod.append(f"VFD (`*.SPD`, feedback `*.SPDFB`) modulates to hold the {header}; "
                   "reset the Δp setpoint down to the most-open served valve (variable-flow).")
    else:
        mod.append(f"Constant-speed; maintain {header} via staging and the min-flow bypass.")
    if extra:
        mod += extra
    b.append(("Modulation & reset", mod))
    b.append(("Safeties & monitoring", [
        "Prove flow before loading; trip on `*.TRIP`; low-Δp / no-flow and pump-fault alarms; "
        "protect duty OR standby availability at all times."]))
    return b


CLASS_LABEL_HINT = {
    "chw": "Chilled-water distribution pumps", "hw": "Hot-water distribution pumps",
    "cond": "Condenser-water pumps", "sec": "Secondary buffer-loop pumps",
    "hotwell": "Hot-well / condensate-return pumps",
}


def soo_tower(devs, sfx):
    cells = [d for d in devs if d.count("-") >= 2]
    return [
        ("Purpose", ["Reject condenser heat and hold the condenser-water setpoint at minimum fan energy."]),
        ("Staging & reset", [
            "Condenser-water setpoint = wet-bulb + approach, floored for chiller min-lift.",
            f"INV fan cells `*.SPD` (feedback `*.SPDFB`) modulate together to hold setpoint; "
            f"stage cells ({', '.join(cells) if cells else 'per demand'}) as load rises; start/stop `*.SS`.",
            "Sequence with condenser pumps + chiller; run only when the chiller/free-cooling calls."]),
        ("Safeties & monitoring", [
            "Basin E/H heater on low temp (freeze); makeup/level & conductivity (blowdown, "
            "Legionella control); fan INV fault `*.TRIP`; run status `*.RUN`."]),
    ]


def soo_hx(devs, sfx, kind="tempered loop"):
    return [
        ("Purpose", [f"Transfer heat to the {kind}; hold the secondary supply temperature."]),
        ("Control", [
            f"Modulate primary control valve `*.PV` (feedback `*.PVFB`) to hold the secondary "
            f"outlet setpoint `*.S-OUT` from primary inlet `*.P-IN`."]
            + _lead_lag(devs, "exchangers")),
        ("Safeties", ["Freeze / high-temp limit on the secondary; valve-fault supervision; "
                      "isolate on served-loop shutdown."]),
    ]


def soo_freecool(devs, sfx):
    return [
        ("Purpose", ["Winter/shoulder free-cooling: use the cooling tower to make chilled water "
                     "via the HX when wet-bulb is low enough, displacing the chiller."]),
        ("Control", ["Enable when tower-approachable CHW temp < CHW setpoint; modulate the primary "
                     "valve to hold the loop setpoint; hand over to/from mechanical cooling with hysteresis."]),
        ("Safeties", ["Freeze protection; isolate when disabled."]),
    ]


def soo_ahu(devs, sfx):
    multi = _n(devs) > 1
    who = ", ".join(devs)
    b = [("Purpose", [f"Air-handling unit{'s' if multi else ''} {who}: maintain supply-air "
                      "temperature and space conditions on the occupancy schedule."])]
    b.append(("Start / fan control", [
        f"Start/stop `*.SS` on the occupancy time-program; prove `*.RUN`, alarm `*.TRIP`.",
        ("VFD supply fan `*.SPD` (feedback `*.SPDFB`) to duct static / demand."
         if any("SPDFB" in sfx(d) for d in devs)
         else "Constant-volume supply fan (fixed speed)."),
    ]))
    coil = [f"Cooling coil `*.CCV` modulates to hold supply-air temp `*.SAT` (return `*.RAT`)."]
    if any("HCV" in sfx(d) for d in devs):
        coil.append("Heating coil `*.HCV` sequenced with a deadband (no simultaneous heat/cool).")
    coil.append("Shoulder-season (Apr–May): drive to OA economizer / free-cooling — 100% OA, "
                "cooling coil closed — per the cutover plan.")
    b.append(("Temperature control", coil))
    b.append(("Safeties & monitoring", [
        "Freeze-stat, fan-status proof, filter-DP and smoke/fire interlock (shut fan, spring-return "
        "dampers); high supply-temp alarm `*.SAT`."]))
    return b


def soo_vent(devs, sfx):
    return [("Purpose", ["Ventilation / exhaust: run to the occupancy schedule and IAQ demand."]),
            ("Control", ["Start/stop `*.SS` on schedule; interlock with associated AHU / fire mode; "
                         "prove `*.RUN`, alarm `*.TRIP`."])]


def soo_oa_station(devs, sfx):
    return [("Purpose", ["Outdoor-air station: temper and deliver ventilation air to the loop."]),
            ("Control", ["Modulate to the OA supply setpoint; enable with the served distribution; "
                         "economizer-favourable in shoulder season."])]


def soo_fcu(devs, sfx):
    return [("Purpose", ["Fan-coil zone group: maintain space temperature."]),
            ("Control", ["Zone thermostat cycles fan / modulates the coil valve; batch on/off and "
                         "night-setback by the schedule; status/alarm to BMS."])]


def soo_packaged(devs, sfx):
    return [
        ("Purpose", [f"Packaged DX unit{'s' if _n(devs) > 1 else ''} ({', '.join(devs)}) — "
                     "several serve IT/electrical rooms and run continuously."]),
        ("Control & monitoring", [
            "BMS enables/schedules via `*.SS` and monitors run `*.RUN` + fault `*.TRIP`; "
            "capacity is on the unit's integral DX thermostat.",
            "Electrical/IT-room units are 24/7 — treat as critical (stage spare cooling before any swap)."]),
    ]


def soo_lighting(devs, sfx):
    b = [("Purpose", ["Common-area & facade lighting groups on time-program / astronomical clock."]),
         ("Control", [
             "On/Off `*.CMD` by schedule (and photocell/astro for facade); status `*.ST` proof.",
         ])]
    if any(d.startswith("LTG-AVIATION") for d in devs):
        b.append(("Life-safety", [
            "Aviation obstruction light: regulatory — independent circuit, dusk-to-dawn/photocell, "
            "never commanded dark; failure alarms to central."]))
    return b


def soo_meter(devs, sfx):
    return [("Purpose", ["Energy metering (no control output)."]),
            ("Function", ["Integrate pulse/analog energy for trend, demand limiting and peak "
                          "monitoring; feed the dispatch/peak-shave logic."])]


def soo_hotwell(devs, sfx):
    return [("Purpose", ["Hot-well / condensate tank: buffer condensate return to the district."]),
            ("Control", ["Level control with makeup; high/low-level alarms."])]


def soo_expansion(devs, sfx):
    return [("Purpose", ["Loop pressurisation / expansion set."]),
            ("Control", ["Maintain fill pressure via makeup valve/pump; low-pressure & level alarms."])]


def soo_changeover(devs, sfx):
    return [("Purpose", ["Source-changeover valve set (R-1 branch vs DHC through-main)."]),
            ("Control", ["Sequenced bumpless transfer with flow proof before compressor enable; "
                         "position feedback and end-switch supervision."])]


def soo_dhw(devs, sfx):
    return [
        ("Purpose", ["Domestic-hot-water plant: hold stored/delivered DHW temperature per vertical "
                     "zone at minimum steam use, with Legionella protection."]),
        ("Control", [
            "Modulate the steam charging valve `*.STG-CV` to hold the storage setpoint `*.STG-T` "
            "(≥60°C store / ≥55°C delivery); recirc pump `*.PMP-SS` on schedule with duty/standby.",
            "Meter storage steam `*.STM-M`; night/low-demand setback with a periodic thermal-"
            "disinfection cycle."]),
        ("Safeties & monitoring", [
            "High-temp / scald and low-delivery-temp alarms (`*.SUP-T`); pump fault `*.PMP-TRIP`; "
            "low-storage-temp Legionella advisory."]),
    ]


def soo_sanitary(devs, sfx):
    return [
        ("Purpose", ["Plumbing / sanitary: maintain potable (上水) & reclaimed (中水) supply and "
                     "clear drainage/sewage sumps."]),
        ("Control", [
            "Booster / transfer pumps lead/lag on tank level & pressure (`*.SS`, prove `*.RUN`); "
            "sump & sewage pumps run on float level with BMS run/high-level supervision.",
            "Aeration blowers cycle on DO/timer for the wastewater-treatment tanks."]
            + _lead_lag([d for d in devs if d.startswith(("MP", "P-"))], "booster pumps")),
        ("Safeties & monitoring", [
            "Tank high/low-level & dry-run alarms; sump high-water (flooding) alarm `*.HLV`; "
            "fire-reserve tank low-level is statutory; pump fault `*.TRIP`."]),
    ]


def soo_filtration(devs, sfx):
    return [
        ("Purpose", ["Recreational-water (bath / sauna / pool / waterfall / fountain / rainwater) "
                     "circulation & filtration."]),
        ("Control", ["Circulation pump `*.SS` on the daily schedule; periodic backwash; "
                     "prove `*.RUN`, alarm `*.TRIP`; maintain turnover/turbidity per venue."]),
    ]


def soo_smoke(devs, sfx):
    return [
        ("Purpose", ["Smoke exhaust (排煙): clear smoke from the fire zone on alarm."]),
        ("Life-safety control", [
            "Life-safety start is HARD-WIRED from the disaster-prevention (fire) panel — the BMS does "
            "not gate it. On the fire batch `SMK-FIRE.ALM`, the affected `SMF-*` start and supply "
            "AHUs/dampers interlock shut.",
            "BMS provides monitoring (`*.RUN`, `*.TRIP`) and a supervised smoke-mode start `*.SS`; "
            "weekly/periodic run test with runtime logging."]),
    ]


def soo_generic(devs, sfx):
    return [("Purpose", [f"Supervise {', '.join(devs)}: start/stop, status and alarm as available."])]


CLASS_SOO = {
    "chiller": soo_chiller, "dhc_chw": soo_dhc_chw, "dhc_steam": soo_dhc_steam,
    "tower": soo_tower, "freecool": soo_freecool, "emhx": lambda d, s: soo_hx(d, s, "emergency DHC loop"),
    "hx": soo_hx, "ahu": soo_ahu, "vent": soo_vent, "oa_station": soo_oa_station,
    "fcu": soo_fcu, "packaged": soo_packaged, "lighting": soo_lighting, "meter": soo_meter,
    "hotwell": soo_hotwell, "expansion": soo_expansion, "changeover": soo_changeover,
    "dhw": soo_dhw, "sanitary": soo_sanitary, "filtration": soo_filtration, "smoke": soo_smoke,
    "generic": soo_generic,
    "cond_pump": lambda d, s: soo_pumps(d, s, "cond"),
    "chw_pump": lambda d, s: soo_pumps(d, s, "chw"),
    "hw_pump": lambda d, s: soo_pumps(d, s, "hw"),
    "sec_pump": lambda d, s: soo_pumps(d, s, "sec"),
    "hotwell_pump": lambda d, s: soo_pumps(d, s, "hotwell", header="hot-well level"),
}

STRATEGY = {
    "chiller": "Economic dispatch + CHW/CW reset + twin-screw staging",
    "dhc_chw": "Intake valve modulates to CHW setpoint / thermal demand",
    "dhc_steam": "PRV to 0.2 MPa header + condensate metering",
    "tower": "INV fan cells hold wet-bulb-reset condenser setpoint",
    "cond_pump": "Duty/standby with chiller; flow-proof",
    "chw_pump": "Lead/lag VFD to header Δp (variable flow)",
    "hw_pump": "Heating-season lead/lag to header Δp",
    "sec_pump": "Buffer-loop lead/lag to Δp",
    "hotwell_pump": "Level-controlled condensate/kitchen pumps",
    "hx": "Primary valve modulates to secondary outlet setpoint",
    "freecool": "Tower free-cooling when wet-bulb low",
    "emhx": "Emergency DHC HX standby",
    "ahu": "Schedule + SAT coil control + economizer",
    "vent": "Scheduled ventilation / exhaust",
    "oa_station": "OA tempering to setpoint",
    "fcu": "Zone thermostat + schedule batch",
    "packaged": "Enable/monitor; integral DX (some 24/7)",
    "lighting": "Time/astro On-Off + status",
    "meter": "Energy integration (monitor only)",
    "hotwell": "Level control + makeup",
    "expansion": "Loop pressurisation",
    "changeover": "Bumpless source transfer + flow proof",
    "dhw": "Steam charging to storage setpoint + Legionella cycle",
    "sanitary": "Lead/lag booster + float-level sump/sewage",
    "filtration": "Scheduled circulation + backwash",
    "smoke": "Hard-wired fire start; BMS monitors + smoke-mode",
    "generic": "Start/stop + status/alarm",
}


# ---------------------------------------------------------------------------
# Build the model
# ---------------------------------------------------------------------------
def build_model():
    cpts = {}
    for r in g.rows:
        cpts.setdefault(r["Controller"], []).append(r)
    panel_order = [p[0] for p in g.PANELS]
    ordered = sorted(g.controllers,
                     key=lambda c: (panel_order.index(c[1]) if c[1] in panel_order else 99, c[0]))

    panels = {}
    for cid, panel, devices in ordered:
        pts = cpts.get(cid, [])
        if not pts:
            continue
        sfx_by_dev = {}
        for r in pts:
            tag = r["Point Tag"]
            dev = r["Device ID"]
            suf = tag.split(".", 1)[1] if "." in tag else tag
            sfx_by_dev.setdefault(dev, set()).add(suf.upper())

        # group member devices by class in first-seen order
        classes = {}
        for r in pts:
            classes.setdefault(classify(r["Device ID"]), [])
            if r["Device ID"] not in classes[classify(r["Device ID"])]:
                classes[classify(r["Device ID"])].append(r["Device ID"])

        def sfx(dev):
            return sfx_by_dev.get(dev, set())

        sections = []
        strat = []
        for cls, cdevs in classes.items():
            fn = CLASS_SOO.get(cls, soo_generic)
            sections.append({"class": cls, "label": CLASS_LABEL.get(cls, cls),
                             "devices": cdevs, "soo": [{"h": h, "b": b} for h, b in fn(cdevs, sfx)]})
            strat.append(STRATEGY.get(cls, ""))

        io = {"AI": 0, "AO": 0, "BI": 0, "BO": 0}
        for r in pts:
            io[r["I/O Type"]] = io.get(r["I/O Type"], 0) + 1

        rec = {
            "id": cid, "panel": panel, "devices": devices,
            "equipment": ", ".join(sorted({CLASS_LABEL.get(classify(r["Device ID"]), "") for r in pts})),
            "strategy": "; ".join([s for s in dict.fromkeys(strat) if s]),
            "npoints": len(pts), "io": io, "sections": sections,
            "points": [{"tag": r["Point Tag"], "io": r["I/O Type"], "desc": r["Point Description"],
                        "sig": r["Signal / Range"], "units": r["Units"], "sp": r["SP/Sched"],
                        "alarm": r["Alarm"]} for r in pts],
        }
        panels.setdefault(panel, {"id": panel, "desc": PANEL_META.get(panel, ("", ""))[0],
                                  "loc": PANEL_META.get(panel, ("", ""))[1], "controllers": []})
        panels[panel]["controllers"].append(rec)

    plist = [panels[p] for p in panel_order if p in panels]
    return {
        "title": "Red5-DHCP — control logic / Sequence of Operations",
        "generated": TODAY,
        "totals": {"panels": len(plist), "controllers": sum(len(p["controllers"]) for p in plist),
                   "points": len(g.rows)},
        "panels": plist,
    }


# ---------------------------------------------------------------------------
# Renderers
# ---------------------------------------------------------------------------
def render_md(m, path):
    L = [f"# {m['title']}", "",
         f"*Generated {m['generated']} · {m['totals']['controllers']} controllers · "
         f"{m['totals']['panels']} panels · {m['totals']['points']} I/O points.*", "",
         "Control strategy is consistent with the implemented engine in "
         "[`backend/control.py`](../backend/control.py) and `backend/sim.py`.", ""]
    for p in m["panels"]:
        L.append(f"## {p['id']} — {p['desc']}")
        L.append(f"*{p['loc']} · {len(p['controllers'])} controllers*")
        L.append("")
        for c in p["controllers"]:
            L.append(f"### {c['id']}")
            L.append(f"**Equipment:** {c['devices']}  ")
            L.append(f"**I/O:** {c['npoints']} pts "
                     f"(AI {c['io']['AI']} · AO {c['io']['AO']} · BI {c['io']['BI']} · BO {c['io']['BO']})")
            L.append("")
            for sec in c["sections"]:
                if len(c["sections"]) > 1:
                    L.append(f"**[{sec['label']}]** — {', '.join(sec['devices'])}")
                for blk in sec["soo"]:
                    L.append(f"- **{blk['h']}:**")
                    for line in blk["b"]:
                        L.append(f"    - {line}")
                L.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L))


def render_csv(m, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Panel", "Controller", "Equipment", "Control strategy", "Points",
                    "Sequence of operations"])
        for p in m["panels"]:
            for c in p["controllers"]:
                soo = " | ".join(f"{blk['h']}: " + " ".join(blk["b"])
                                 for sec in c["sections"] for blk in sec["soo"])
                w.writerow([p["id"], c["id"], c["devices"], c["strategy"], c["npoints"], soo])


def render_xlsx(m, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9"); B = Border(left=thin, right=thin, top=thin, bottom=thin)
    top = Alignment(wrap_text=True, vertical="top")

    wb = Workbook(); idx = wb.active; idx.title = "Index"
    idx.append([m["title"]]); idx.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F3864")
    idx.append([f"Generated {m['generated']} · {m['totals']['controllers']} controllers · "
                f"{m['totals']['points']} points"])
    idx.append([])
    idx.append(["Panel", "Controller", "Equipment", "Control strategy", "Points", "Sheet"])
    hr = idx.max_row
    for c in range(1, 7):
        idx.cell(row=hr, column=c).fill = HDR; idx.cell(row=hr, column=c).font = HF
    for p in m["panels"]:
        for c in p["controllers"]:
            rr = idx.max_row + 1
            idx.append([p["id"], c["id"], c["devices"], c["strategy"], c["npoints"], c["id"][:31]])
            link = idx.cell(row=rr, column=6)
            link.hyperlink = f"#'{c['id'][:31]}'!A1"; link.font = Font(color="1F5FBF", underline="single")
    for i, wdt in enumerate([13, 13, 40, 44, 7, 12], 1):
        idx.column_dimensions[get_column_letter(i)].width = wdt
    for rr in range(hr, idx.max_row + 1):
        for c in range(1, 7):
            idx.cell(row=rr, column=c).border = B; idx.cell(row=rr, column=c).alignment = top
    idx.freeze_panes = "A5"

    for p in m["panels"]:
        for c in p["controllers"]:
            ws = wb.create_sheet(c["id"][:31])
            ws.append([f"Control logic — {c['id']}"]); ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F3864")
            ws.append([f"Panel {p['id']} ({p['desc']}) · {p['loc']}"])
            ws.append([f"Equipment: {c['devices']}"])
            ws.append([f"Strategy: {c['strategy']}"])
            ws.append([])
            ws.append(["Sequence of Operations"]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            for sec in c["sections"]:
                if len(c["sections"]) > 1:
                    ws.append([f"[{sec['label']}] — {', '.join(sec['devices'])}"])
                    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True, color="1F3864")
                for blk in sec["soo"]:
                    ws.append([blk["h"]]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
                    for line in blk["b"]:
                        ws.append(["  • " + line])
            ws.append([])
            ws.append(["Point schedule"]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            hrow = ws.max_row + 1
            ws.append(["Point Tag", "I/O", "Description", "Signal / Range", "Units", "SP", "Alarm"])
            for cc in range(1, 8):
                ws.cell(row=hrow, column=cc).fill = HDR; ws.cell(row=hrow, column=cc).font = HF
            for pt in c["points"]:
                ws.append([pt["tag"], pt["io"], pt["desc"], pt["sig"], pt["units"], pt["sp"], pt["alarm"]])
            ws.column_dimensions["A"].width = 90
            for col, wdt in zip("ABCDEFG", [18, 6, 40, 16, 8, 6, 7]):
                ws.column_dimensions[col].width = wdt
            for rr in range(1, ws.max_row + 1):
                ws.cell(row=rr, column=1).alignment = top
            for rr in range(hrow, ws.max_row + 1):
                for cc in range(1, 8):
                    ws.cell(row=rr, column=cc).border = B; ws.cell(row=rr, column=cc).alignment = top
    wb.save(path)


def render_html(m, path):
    def esc(s):
        return html.escape(str(s if s is not None else ""))

    blocks = []
    for p in m["panels"]:
        cblocks = []
        for c in p["controllers"]:
            secs = []
            for sec in c["sections"]:
                head = (f'<div class="cls">{esc(sec["label"])} — {esc(", ".join(sec["devices"]))}</div>'
                        if len(c["sections"]) > 1 else "")
                items = "".join(
                    f'<div class="blk"><b>{esc(blk["h"])}</b><ul>'
                    + "".join(f"<li>{esc(line)}</li>" for line in blk["b"]) + "</ul></div>"
                    for blk in sec["soo"])
                secs.append(head + items)
            rows = "".join(
                f'<tr><td>{esc(pt["tag"])}</td><td class=c>{esc(pt["io"])}</td><td>{esc(pt["desc"])}</td>'
                f'<td>{esc(pt["sig"])}</td><td>{esc(pt["units"])}</td>'
                f'<td class=c>{"●" if pt["sp"]=="Y" else ""}</td><td class=c>{"▲" if pt["alarm"]=="Y" else ""}</td></tr>'
                for pt in c["points"])
            blob = esc((c["id"] + " " + p["id"] + " " + c["devices"] + " " + c["strategy"]).lower())
            cblocks.append(
                f'<details class="ctrl" data-s="{blob}"><summary>{esc(c["id"])}'
                f'<span class="r">{esc(c["strategy"])} · {c["npoints"]} pts</span></summary>'
                f'<div class="eq">Equipment: {esc(c["devices"])}</div>'
                f'<div class="soo">{"".join(secs)}</div>'
                '<details class="pts"><summary>Point schedule</summary>'
                '<table><thead><tr><th>Point Tag</th><th class=c>I/O</th><th>Description</th>'
                '<th>Signal / Range</th><th>Units</th><th class=c>SP</th><th class=c>Alm</th></tr></thead>'
                f'<tbody>{rows}</tbody></table></details></details>')
        blocks.append(
            f'<section class="panel"><h2>{esc(p["id"])} — {esc(p["desc"])}'
            f'<span class="r">{esc(p["loc"])} · {len(p["controllers"])} controllers</span></h2>'
            + "".join(cblocks) + "</section>")

    T = m["totals"]
    doc = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(m['title'])}</title>
<style>
  :root {{ --ink:#1b2430; --dim:#5b6675; --line:#d5dbe4; --band:#1F3864; }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; font:13px/1.5 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif; color:var(--ink); background:#f4f6f9; padding:24px; }}
  .wrap {{ max-width:1120px; margin:0 auto; }}
  h1 {{ font-size:21px; margin:0 0 2px; }} .sub {{ color:var(--dim); margin:0 0 14px; }}
  .bar {{ position:sticky; top:0; background:#f4f6f9; padding:8px 0 12px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; z-index:5; }}
  input[type=search] {{ border:1px solid var(--line); border-radius:7px; padding:7px 10px; min-width:300px; font-size:13px; }}
  button {{ border:1px solid var(--line); background:#fff; border-radius:7px; padding:6px 11px; cursor:pointer; font-size:12px; }}
  section.panel {{ margin:0 0 18px; }}
  h2 {{ font-size:15px; color:var(--band); border-bottom:2px solid var(--band); padding-bottom:4px; margin:18px 0 8px; }}
  h2 .r, summary .r {{ float:right; color:var(--dim); font-weight:400; font-size:12px; }}
  details.ctrl {{ background:#fff; border:1px solid var(--line); border-radius:9px; margin:0 0 8px; }}
  details.ctrl > summary {{ padding:10px 14px; font-weight:600; cursor:pointer; list-style:none; }}
  summary::-webkit-details-marker {{ display:none; }}
  .eq {{ color:var(--dim); font-size:12px; padding:0 14px 6px; }}
  .soo {{ padding:0 14px 8px; }} .cls {{ font-weight:600; color:var(--band); margin:8px 0 2px; }}
  .blk {{ margin:2px 0 6px; }} .blk b {{ font-size:12.5px; }} .blk ul {{ margin:3px 0 0; padding-left:20px; }} .blk li {{ margin:2px 0; }}
  details.pts {{ margin:0 14px 12px; }} details.pts > summary {{ font-size:12px; color:var(--dim); cursor:pointer; }}
  table {{ border-collapse:collapse; width:100%; font-size:12px; margin-top:6px; }}
  th, td {{ border:1px solid var(--line); padding:4px 7px; text-align:left; vertical-align:top; }}
  thead th {{ background:#eef1f6; }} td.c, th.c {{ text-align:center; }}
  .hidden {{ display:none !important; }}
  .foot {{ color:var(--dim); font-size:12px; margin-top:18px; }}
  @media print {{ body {{ background:#fff; padding:0; }} .bar {{ display:none; }} details {{ }} details.ctrl {{ break-inside:avoid; }} }}
</style></head>
<body><div class="wrap">
<h1>{esc(m['title'])}</h1>
<p class="sub">{T['controllers']} controllers · {T['panels']} panels · {T['points']} I/O points · generated {m['generated']}. Strategy matches backend/control.py + sim.py.</p>
<div class="bar">
  <input type="search" id="q" placeholder="find controller / equipment / strategy\u2026">
  <button id="expand">Expand all</button><button id="collapse">Collapse all</button>
  <button onclick="window.print()">Print / Save PDF</button>
  <span class="sub" id="count"></span>
</div>
{"".join(blocks)}
<p class="foot">● = has setpoint/schedule · ▲ = alarmed point. Source: generate_io_list.py.</p>
</div>
<script>
const items=[...document.querySelectorAll('details.ctrl')];
const cnt=document.getElementById('count'); cnt.textContent=items.length+' controllers';
document.getElementById('q').addEventListener('input',e=>{{
  const s=e.target.value.trim().toLowerCase(); let n=0;
  for(const d of items){{const hit=!s||d.dataset.s.includes(s);d.classList.toggle('hidden',!hit);if(hit)n++;d.open=!!s&&hit;}}
  cnt.textContent=n+' / '+items.length+' controllers';
}});
document.getElementById('expand').onclick=()=>items.forEach(d=>d.open=true);
document.getElementById('collapse').onclick=()=>items.forEach(d=>d.open=false);
</script>
</body></html>
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)


CANVAS_TEMPLATE = r'''import {
  useCanvasState,
  Stack, Row, Grid, Spacer,
  H1, H2, Text, Pill, Stat, Table, CollapsibleSection, TextInput,
} from "cursor/canvas";

type Blk = { h: string; b: string[] };
type Sec = { class: string; label: string; devices: string[]; soo: Blk[] };
type Pt = { tag: string; io: string; desc: string; sig: string; units: string; sp: string; alarm: string };
type Ctrl = { id: string; panel: string; devices: string; equipment: string; strategy: string;
  npoints: number; io: { AI: number; AO: number; BI: number; BO: number }; sections: Sec[]; points: Pt[] };
type Panel = { id: string; desc: string; loc: string; controllers: Ctrl[] };
type Model = { title: string; generated: string;
  totals: { panels: number; controllers: number; points: number }; panels: Panel[] };

const DATA: Model = __DATA__;

function match(c: Ctrl, q: string) {
  const s = q.toLowerCase();
  return (c.id + " " + c.panel + " " + c.devices + " " + c.strategy).toLowerCase().includes(s);
}

export default function DHCPControlLogic() {
  const [q, setQ] = useCanvasState("q", "");
  const T = DATA.totals;
  return (
    <Stack gap={16} style={{ padding: 20 }}>
      <div>
        <H1>{DATA.title}</H1>
        <Text tone="secondary">
          Sequence of Operations for every DDC controller — panel, equipment and I/O visibility.
          Matches backend/control.py + sim.py \u00b7 {DATA.generated}
        </Text>
      </div>
      <Grid columns={3} gap={12}>
        <Stat value={T.panels} label="Panels" />
        <Stat value={T.controllers} label="Controllers (with SOO)" />
        <Stat value={T.points} label="I/O points" tone="info" />
      </Grid>
      <Row align="center" gap={10}>
        <H2 style={{ margin: 0 }}>Panel \u2192 Controller \u2192 control logic</H2>
        <Spacer />
        <TextInput type="search" value={q} onChange={setQ}
          placeholder="find controller / equipment / strategy\u2026" style={{ minWidth: 300 }} />
      </Row>
      <Stack gap={2}>
        {DATA.panels.map((p) => {
          const cs = p.controllers.filter((c) => !q || match(c, q));
          if (q && cs.length === 0) return null;
          return (
            <div key={p.id + "|" + q}>
            <CollapsibleSection title={p.id + " \u2014 " + p.desc} count={p.controllers.length}
              defaultOpen={!!q}
              trailing={<Text size="small" tone="tertiary">{p.loc}</Text>}>
              <Stack gap={2}>
                {cs.map((c) => (
                  <div key={c.id + "|" + q}>
                  <CollapsibleSection title={c.id} defaultOpen={!!q}
                    trailing={<Text size="small" tone="tertiary">{c.strategy} \u00b7 {c.npoints} pts</Text>}>
                    <Text size="small" tone="tertiary" style={{ marginBottom: 6 }}>Equipment: {c.devices}</Text>
                    {c.sections.map((sec, si) => (
                      <div key={si} style={{ marginBottom: 8 }}>
                        {c.sections.length > 1 ? (
                          <Text size="small" weight="semibold" style={{ display: "block", marginBottom: 2 }}>
                            {sec.label} \u2014 {sec.devices.join(", ")}
                          </Text>
                        ) : null}
                        {sec.soo.map((blk, bi) => (
                          <div key={bi} style={{ marginBottom: 4 }}>
                            <Text size="small" weight="semibold">{blk.h}</Text>
                            {blk.b.map((line, li) => (
                              <span key={li}>
                              <Text size="small" tone="secondary" style={{ display: "block", marginLeft: 10 }}>
                                \u2022 {line}
                              </Text>
                              </span>
                            ))}
                          </div>
                        ))}
                      </div>
                    ))}
                    <CollapsibleSection title="Point schedule" count={c.points.length}>
                      <Table
                        headers={["Point Tag", "I/O", "Description", "Signal / Range", "Units"]}
                        columnAlign={["left", "center", "left", "left", "left"]}
                        striped stickyHeader
                        rows={c.points.map((pt) => [pt.tag, pt.io, pt.desc, pt.sig, pt.units])}
                      />
                    </CollapsibleSection>
                  </CollapsibleSection>
                  </div>
                ))}
              </Stack>
            </CollapsibleSection>
            </div>
          );
        })}
      </Stack>
    </Stack>
  );
}
'''


def render_canvas(m, path):
    out = CANVAS_TEMPLATE.replace("__DATA__", json.dumps(m, ensure_ascii=False))
    for e, ch in (("\\u2014", "\u2014"), ("\\u00b7", "\u00b7"), ("\\u2192", "\u2192"),
                  ("\\u2026", "\u2026"), ("\\u2022", "\u2022")):
        out = out.replace(e, ch)
    with open(path, "w", encoding="utf-8") as f:
        f.write(out)


if __name__ == "__main__":
    os.makedirs(EXPORT_DIR, exist_ok=True)
    os.makedirs(DOCS_DIR, exist_ok=True)
    m = build_model()
    render_md(m, os.path.join(DOCS_DIR, "control_logic.md"))
    stem = os.path.join(EXPORT_DIR, "red5-dhcp_control-logic")
    render_csv(m, stem + ".csv")
    render_xlsx(m, stem + ".xlsx")
    render_html(m, stem + ".html")
    try:
        os.makedirs(CANVAS_DIR, exist_ok=True)
        render_canvas(m, os.path.join(CANVAS_DIR, "dhcp-control-logic.canvas.tsx"))
        print("canvas written")
    except PermissionError:
        print("canvas skipped (sandbox); run with elevated perms to write canvas")
    print(f"control logic: {m['totals']['controllers']} controllers, {m['totals']['points']} points")
