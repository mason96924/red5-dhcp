"""build_cutover.py -- downloadable BMS controller-replacement cutover plan.

Re-based on the AUTHORITATIVE physical panel schedule (panels_schedule.json,
from 판넬별 포인트 정리 26.07.30, now with per-panel floor + equipment served):
the plan covers all 67 field panels / 4,467 I/O points, grouped by physical
floor/area, with per-phase panel & point counts
that sum to the full I/O. Mirrors the dhcp-controller-cutover-plan canvas into
portable, print-friendly files under exports/: HTML (print-to-PDF), XLSX
(Phases / Playbook / Risks / Summary) and CSV (phase table).

Phase→area mapping is an engineering-judgement grouping (idle rooftop/heat-source
plant first, occupied floors on economizer overnight, basement plant/hydronics
last with DHC-intake coordination). It is documented in the output.
"""
from __future__ import annotations

import csv
import datetime
import html
import json
import os

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(HERE, "exports")
SCHEDULE = os.path.join(HERE, "panels_schedule.json")
TODAY = datetime.date.today().isoformat()

# --- load physical schedule ------------------------------------------------
_sched = json.load(open(SCHEDULE, encoding="utf-8"))
_PANELS = _sched["panels"]
SRC = _sched.get("source", "")

N_PANELS = len(_PANELS)
N_POINTS = _sched["usedTotal"]
N_CAP = _sched["capTotal"]
N_MODULES = sum(_sched["modTotals"])

FLOOR_AREAS = ["Floor 31", "Floor 21", "Floor 10", "Floor 4", "Floor 3", "Floor 2", "Floor 1"]
BASEMENT_AREAS = ["Basement B1", "Basement B2", "Basement B3"]


def _is_central(p):
    return "시스템" in p["name"]


def _pick_stats(pred):
    names, pts, npan = [], 0, 0
    for p in _PANELS:
        if pred(p):
            names.append(p["name"])
            pts += p["usedTot"]
            npan += 1
    return names, npan, pts


TITLE = "Red5-DHCP — BMS controller replacement: minimal-interruption cutover"
SUBTITLE = (f"Migrate the hotel's {N_PANELS} field panels ({N_MODULES} Delta controllers + modules, "
            f"{N_POINTS:,} I/O points) from Azbil savic-net FX2 to new DDC in the April–May shoulder "
            "season, using OA economizer free-cooling and local manual operation to keep the hotel "
            "conditioned throughout.")

STATS = [("Physical panels to migrate", f"{N_PANELS}"),
         ("Delta controllers + modules", f"{N_MODULES}"),
         ("I/O points (commissioning checklist)", f"{N_POINTS:,}"),
         ("Cutover window (Mar prep)", "Apr–May")]

WHY = ("Tokyo shoulder-season OA (~10–22 °C, low enthalpy) lets the airside economizer meet the load "
       "with no mechanical cooling, so the towers and R-1 chiller are already idle and swap cold. "
       "Heating is over, so hot-water hydronics are low-risk. DHC stays live as a warm-spell safety "
       "net and its intake is migrated last.")
ENABLER = ("Each motor starter has a 手元/遠方 (local/remote) selector + DC24V ext start-stop + status "
           "feedback (confirmed on E-01/13-4). During a swap the equipment runs in LOCAL manual — fan "
           "hand-on, OA damper fixed ~100%, pump fixed-speed — independent of the BMS.")

# phase -> (n, window, label, predicate, state, method, risk)
PHASE_DEFS = [
    ("0", "March (prep)", "System / central head-end (B2F)",
     _is_central,
     "n/a — runs in parallel",
     f"Stand up new head-end alongside FX2; prefab panels; bench-test programs; point-verify vs the "
     f"{N_POINTS:,}-pt I/O list", "Prep"),
    ("1", "early Apr", "Penthouse (PH1/PH2) — rooftop plant, R-1 chiller, OAUs",
     lambda p: p["area"] == "Penthouse (PH)",
     "Idle — no mechanical cooling; rooftop towers / R-1 / OAU off",
     "Power down & swap cold; recommission before any warm spell", "Low"),
    ("2", "April–May", "Guest & public floors — 31F, 21F, 10F, 4F–1F (incl. EVU-6)",
     lambda p: p["area"] in FLOOR_AREAS,
     "Economizer free-cooling / DHC carries the occupied floors",
     "Guest-floor RS/RCP risers to local; swap overnight, one riser/floor at a time", "Moderate"),
    ("3", "Apr–May (last)", "Basements B1–B3 — plant, hydronics, DHC intake, packaged units",
     lambda p: p["area"] in BASEMENT_AREAS and not _is_central(p),
     "Heating off; only trickle CHW; DHC live as safety net",
     "Pumps / HX / hydronics one at a time on local fixed-speed; duty & standby never both down; "
     "DHC-intake panels done last with district coordination (low-occupancy weekend)", "Critical"),
]


def build_phases():
    """Expand PHASE_DEFS with real panel names + panel/point counts from the schedule."""
    rows = []
    for n, window, label, pred, state, method, risk in PHASE_DEFS:
        names, npan, pts = _pick_stats(pred)
        panels_txt = f"{label} — {', '.join(names)}"
        rows.append((n, window, panels_txt, npan, pts, state, method, risk))
    return rows


PHASES = build_phases()
_TOT_PAN = sum(r[3] for r in PHASES)
_TOT_PTS = sum(r[4] for r in PHASES)

PLAYBOOK = [
    ("AHU / OAU fans", "COS→手元, fan hand-ON, OA damper fixed ~100% (economizer), DHC coil valve left manually crackable",
     "Warm spell → be able to open the cooling valve within minutes"),
    ("CHW / HW pumps", "One pump hand-ON at fixed speed; DP/bypass set manually",
     "Keep duty OR standby always available; protect domestic-hot-water continuity"),
    ("Cooling towers / condenser / R-1", "Leave isolated & de-energised",
     "None — off-season; recommission before enabling mechanical cooling"),
    ("DHC intake", "Intake isolation + energy valves and pumps to local manual, fixed position",
     "Primary source — coordinate with district operator; do these basement panels last"),
    ("Packaged units (IT/elec rooms)", "DX on local thermostat, continuous run",
     "24/7 critical rooms — stage spare/portable cooling; swap one unit at a time"),
    ("Lighting", "Local switching / manual override per circuit",
     "Aviation obstruction light is regulatory life-safety — independent circuit, always lit"),
]

RISKS = [
    ("Cutover discipline", "warn",
     "Only one system in manual at any time. Guest-facing zones swapped overnight, back-of-house first. "
     "Old controller stays landed on a marshalling strip so you can revert to FX2 within minutes. "
     "Freeze all scope changes during the window; spares on site."),
    ("Never interrupt", "crit",
     "Life-safety interlocks — smoke control, stair pressurization, exhaust — remain live and coordinated "
     "with the fire-alarm system. Aviation obstruction light and IT/electrical-room cooling stay energised."),
    ("Verify every point", "info",
     f"The {N_POINTS:,}-point physical I/O schedule is the master checklist. Per panel: point-to-point "
     "verify each DO/DI/BTOT/AI/AO, functional-test, then trend 24–48 h before returning to auto."),
    ("Weather safety net", "ok",
     "DHC stays operational until the final phase, so a warm spell is covered. The sequence guarantees the "
     "building always has either economizer free-cooling or the DHC source available."),
]

RISK_HEX = {"Prep": "E7ECF5", "Low": "DDF0E4", "Moderate": "FBEEDA", "Critical": "F7DEDE"}
RISK_HTML = {"Prep": "#eef2fb", "Low": "#e7f6ec", "Moderate": "#fdf2df", "Critical": "#fbe3e3"}
CALLOUT_HTML = {"warn": "#fdf2df", "crit": "#fbe3e3", "info": "#eef2fb", "ok": "#e7f6ec"}

PHASE_COLS = ["Phase", "Window", "Areas / panels", "Panels", "Points",
              "Shoulder-season state", "Swap method", "Risk"]


# --- CSV -------------------------------------------------------------------
def write_csv(path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(PHASE_COLS)
        w.writerows(PHASES)
        w.writerow(["TOTAL", "", "", _TOT_PAN, _TOT_PTS, "", "", ""])
        w.writerow([])
        w.writerow(["Manual-operation playbook"])
        w.writerow(["System", "Local-mode setup", "Watch-out"])
        w.writerows(PLAYBOOK)


# --- XLSX ------------------------------------------------------------------
def write_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HDR = PatternFill("solid", fgColor="1F3864")
    HF = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    B = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, n):
        for c in range(1, n + 1):
            ws.cell(row=1, column=c).fill = HDR
            ws.cell(row=1, column=c).font = HF

    def widths(ws, ws_w):
        for i, w in enumerate(ws_w, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def frame(ws, n):
        for rr in range(2, ws.max_row + 1):
            for c in range(1, n + 1):
                ws.cell(row=rr, column=c).border = B
                ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([TITLE]); ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F3864")
    ws.append([SUBTITLE]); ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="555555")
    ws.append([f"Generated {TODAY}"])
    ws.append([])
    for label, val in STATS:
        ws.append([label, val])
    ws.append([])
    ws.append(["Why April–May", WHY])
    ws.append(["Enabler — local/remote", ENABLER])
    widths(ws, [30, 90])
    for rr in range(1, ws.max_row + 1):
        ws.cell(row=rr, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Phases")
    ws.append(PHASE_COLS)
    hdr(ws, len(PHASE_COLS))
    for row in PHASES:
        ws.append(list(row))
        fill = RISK_HEX.get(row[7])
        if fill:
            for c in range(1, len(PHASE_COLS) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill("solid", fgColor=fill)
    ws.append(["TOTAL", "", "", _TOT_PAN, _TOT_PTS, "", "", ""])
    for c in range(1, len(PHASE_COLS) + 1):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    widths(ws, [7, 13, 52, 8, 8, 28, 44, 11])
    ws.freeze_panes = "A2"
    frame(ws, len(PHASE_COLS))

    ws = wb.create_sheet("Playbook")
    ws.append(["System", "Local-mode setup while BMS is disconnected", "Watch-out"])
    hdr(ws, 3)
    for row in PLAYBOOK:
        ws.append(list(row))
    widths(ws, [28, 60, 46])
    ws.freeze_panes = "A2"
    frame(ws, 3)

    ws = wb.create_sheet("Risks")
    ws.append(["Control", "Detail"])
    hdr(ws, 2)
    for title, _t, text in RISKS:
        ws.append([title, text])
    widths(ws, [24, 100])
    ws.freeze_panes = "A2"
    frame(ws, 2)

    wb.save(path)


# --- HTML (light, print-friendly) -----------------------------------------
def _rows_html():
    out = []
    for n, window, panels, npan, pts, state, method, risk in PHASES:
        out.append(
            f'<tr style="background:{RISK_HTML.get(risk, "#fff")}">'
            f"<td class=c>{html.escape(n)}</td><td>{html.escape(window)}</td>"
            f"<td>{html.escape(panels)}</td><td class=c>{npan}</td><td class=c>{pts:,}</td>"
            f"<td>{html.escape(state)}</td>"
            f"<td>{html.escape(method)}</td><td class=c><b>{html.escape(risk)}</b></td></tr>")
    out.append(
        f'<tr style="font-weight:700"><td></td><td>TOTAL</td><td></td>'
        f'<td class=c>{_TOT_PAN}</td><td class=c>{_TOT_PTS:,}</td><td></td><td></td><td></td></tr>')
    return "".join(out)


def _play_html():
    return "".join(
        f"<tr><td><b>{html.escape(s)}</b></td><td>{html.escape(setup)}</td><td>{html.escape(w)}</td></tr>"
        for s, setup, w in PLAYBOOK)


def _risk_html():
    return "".join(
        f'<div class="call" style="background:{CALLOUT_HTML[t]}"><b>{html.escape(title)}</b>'
        f"<p>{html.escape(text)}</p></div>" for title, t, text in RISKS)


def write_html(path):
    stats = "".join(f'<div class="stat"><b>{v}</b><span>{html.escape(l)}</span></div>' for l, v in STATS)
    doc = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(TITLE)}</title>
<style>
  :root {{ --ink:#1b2430; --dim:#5b6675; --line:#d8dee7; --band:#1F3864; }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; font:14px/1.5 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;
          color:var(--ink); background:#f4f6f9; padding:28px; }}
  .page {{ max-width:1180px; margin:0 auto; background:#fff; border:1px solid var(--line);
           border-radius:10px; padding:28px 32px; }}
  h1 {{ font-size:22px; margin:0 0 4px; }} .sub {{ color:var(--dim); margin:0 0 18px; max-width:90ch; }}
  h2 {{ font-size:16px; margin:26px 0 8px; border-bottom:2px solid var(--band); padding-bottom:4px; color:var(--band); }}
  .stats {{ display:flex; gap:26px; flex-wrap:wrap; margin:0 0 16px; }}
  .stat b {{ font-size:22px; display:block; }} .stat span {{ color:var(--dim); font-size:12px; }}
  .cols {{ display:flex; gap:14px; flex-wrap:wrap; }}
  .call {{ flex:1; min-width:280px; border:1px solid var(--line); border-radius:8px; padding:12px 14px; }}
  .call b {{ display:block; margin-bottom:4px; }} .call p {{ margin:0; color:#33404f; font-size:13px; }}
  table {{ border-collapse:collapse; width:100%; margin:6px 0 6px; font-size:13px; }}
  th, td {{ border:1px solid var(--line); padding:7px 9px; text-align:left; vertical-align:top; }}
  thead th {{ background:var(--band); color:#fff; }} td.c, th.c {{ text-align:center; white-space:nowrap; }}
  .note {{ color:var(--dim); font-size:12px; margin:4px 0 0; }}
  .grid2 {{ display:grid; grid-template-columns:1fr 1fr; gap:12px; }}
  .foot {{ color:var(--dim); font-size:12px; margin-top:22px; }}
  @media print {{ body {{ background:#fff; padding:0; }} .page {{ border:0; }} }}
</style></head>
<body><div class="page">
<h1>{html.escape(TITLE)}</h1>
<p class="sub">{html.escape(SUBTITLE)}</p>
<div class="stats">{stats}</div>
<div class="cols">
  <div class="call" style="background:#e7f6ec"><b>Why April–May</b><p>{html.escape(WHY)}</p></div>
  <div class="call" style="background:#eef2fb"><b>The enabler: local/remote at every starter</b><p>{html.escape(ENABLER)}</p></div>
</div>
<h2>Phased sequence — coldest/idle first, primary source last</h2>
<p class="note">Row shading = comfort/continuity risk during that phase. One system in manual at a time; guest floors overnight. Panel / point counts are from the {SRC} physical schedule (grouped by the panel's own floor) and sum to {_TOT_PAN} panels / {_TOT_PTS:,} points.</p>
<table><thead><tr><th class=c>#</th><th>Window</th><th>Areas / panels</th><th class=c>Panels</th><th class=c>Points</th><th>Shoulder-season state</th><th>Swap method</th><th class=c>Risk</th></tr></thead>
<tbody>{_rows_html()}</tbody></table>
<p class="note">Phase→floor grouping is an engineering-judgement mapping (rooftop/heat-source plant idle first, occupied floors on economizer overnight, basement plant/hydronics + DHC-intake last). The schedule now names the equipment each RCP/CP panel serves (AC-* AHUs, EVU-* OAUs, SF/EF fans, PCU packaged units); guest-floor RS risers hold most of the newly-surfaced points.</p>
<h2>Manual-operation playbook (during each panel's swap)</h2>
<table><thead><tr><th>System</th><th>Local-mode setup while BMS is disconnected</th><th>Watch-out</th></tr></thead>
<tbody>{_play_html()}</tbody></table>
<h2>Risk controls & rollback</h2>
<div class="grid2">{_risk_html()}</div>
<p class="foot">Generated {TODAY} · Red5-DHCP · savic-net FX2 → new DDC migration · source: panels_schedule.json (판넬별 포인트 정리 {SRC}). Print to PDF for distribution.</p>
</div></body></html>
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)


if __name__ == "__main__":
    os.makedirs(EXPORT_DIR, exist_ok=True)
    stem = os.path.join(EXPORT_DIR, "red5-dhcp_controller-cutover-plan")
    write_csv(stem + ".csv")
    write_xlsx(stem + ".xlsx")
    write_html(stem + ".html")
    assert _TOT_PAN == N_PANELS and _TOT_PTS == N_POINTS, (_TOT_PAN, N_PANELS, _TOT_PTS, N_POINTS)
    print(f"cutover plan exported: {len(PHASES)} phases, {_TOT_PAN} panels / {_TOT_PTS} points "
          f"-> {stem}.{{csv,xlsx,html}}")
