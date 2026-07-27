"""build_switchboard.py -- power-panel / motor-control-panel (MCC) schedule.

Models the electrical distribution that the ESCO energy-saving retrofit drawings
actually define, then exports CSV / XLSX / HTML into ``exports/``.

Source drawings (2015-04-30 竣工図, (株)ヤマグチ):
  - E-01  動力制御盤改修図(撤去・改修)     panel + feeder schedule
  - E-02  幹線動力設備平面図              main feeder routing / cable sizes
  - 13-4  動力盤改造図                    P-PH1-2 single-line + CP-8 starter detail
  - 共-02 電気設備 特記仕様書             general electrical spec

Scope note: these retrofit drawings cover only the circuits the ESCO project
touched on the PH1F machine-room power panels (RC-1 chiller feeders + the new
CP-8 DHC chilled-water pump).  The whole-building substation and per-floor
lighting/receptacle panelboards are not part of this drawing set.
"""
from __future__ import annotations

import csv
import datetime
import html
import json
import os

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(HERE, "exports")
TODAY = datetime.date.today().isoformat()

TITLE = "Red5-DHCP \u2014 Power-panel / MCC schedule (ESCO retrofit)"
SUBTITLE = ("PH1F machine-room 動力制御盤 feeders for the RC-1 (36/37F) chiller "
            "and the new CP-8 DHC chilled-water pump. Source: E-01 / E-02 / 13-4 動力盤改造図.")

# --- schedule model -------------------------------------------------------
# Each panel: id, name, location, supply, note, circuits[]
# Each circuit: circuit, load, kw, device, cable, control, work
PANELS = [
    {
        "id": "P-PH1-1",
        "name": "動力制御盤 P-PH1-1 (Power/motor-control panel)",
        "location": "PH1F machine room",
        "supply": "3φ3W 400V",
        "note": "Existing panel; CP-8 chilled-water-pump circuit newly added under the ESCO works.",
        "circuits": [
            {"circuit": "CP-8", "load": "CP-8 冷水ポンプ (DHC-side CHW pump)", "kw": 3.7,
             "device": "ELCB 3P 30AF/20AT (MWS-CV) + MS (MSO-N21) + THR 2.2–5A",
             "cable": "CV3.5sq-4C (\u2205 22)",
             "control": "WH 20/5A 1kWh-pulse \u2192 BMS; local/remote COS; ext start-stop DC24V; run/fault status out",
             "work": "New"},
        ],
    },
    {
        "id": "P-PH1-2",
        "name": "動力制御盤 P-PH1-2 (Power/motor-control panel)",
        "location": "PH1F machine room",
        "supply": "3φ3W 400V",
        "note": "Existing panel modified; 600AF main + revenue/sub kWh meter; feeds the new RC-1 local switch panel.",
        "circuits": [
            {"circuit": "MAIN", "load": "Panel incomer / main breaker", "kw": None,
             "device": "MCCB 3P 600AF/600AT", "cable": "CVT100sq-E14 (existing conduit reused)",
             "control": "WH 200/5A 1kWh-pulse \u2192 BMS", "work": "Modify"},
            {"circuit": "F-RC1", "load": "Feeder \u2192 手元開閉器盤 (RC-1用) local panel", "kw": None,
             "device": "MCCB 3P 250AF/200AT", "cable": "CVT100sq-E14",
             "control": "\u2014", "work": "Modify"},
        ],
    },
    {
        "id": "LSP-RC1",
        "name": "手元開閉器盤 (RC-1用) — RC-1 local switch/disconnect panel",
        "location": "PH1F machine room (adjacent RC-1 / 36-37F chiller)",
        "supply": "3φ3W 400V from P-PH1-2",
        "note": "New local panel for the RC-1 (Ebara twin-screw) chiller: one feeder per compressor + two spare feeders.",
        "circuits": [
            {"circuit": "RC-1(1)", "load": "36/37F chiller RC-1 — compressor #1", "kw": 43.5,
             "device": "MCCB 3P 250AF/150AT", "cable": "CVT22sq-E5.5",
             "control": "Chiller-integral control; BMS run/alarm + kW", "work": "New"},
            {"circuit": "RC-1(2)", "load": "36/37F chiller RC-1 — compressor #2", "kw": 43.5,
             "device": "MCCB 3P 250AF/150AT", "cable": "CVT22sq-E5.5",
             "control": "Chiller-integral control; BMS run/alarm + kW", "work": "New"},
            {"circuit": "SP-1", "load": "予備電源 (spare feeder 1)", "kw": None,
             "device": "MCCB 3P 50AF/30AT", "cable": "CV5.5sq-3C", "control": "\u2014", "work": "New"},
            {"circuit": "SP-2", "load": "予備電源 (spare feeder 2)", "kw": None,
             "device": "MCCB 3P 50AF/30AT", "cable": "CV5.5sq-3C", "control": "\u2014", "work": "New"},
        ],
    },
]

SCHED_COLS = ["Panel", "Location", "Supply", "Circuit", "Load / served", "kW",
              "Protective device", "Feeder cable", "Metering / control", "Work"]


def dataset():
    total_kw = 0.0
    ncirc = 0
    for p in PANELS:
        for c in p["circuits"]:
            ncirc += 1
            if c["kw"]:
                total_kw += c["kw"]
    return {
        "title": TITLE, "subtitle": SUBTITLE, "generated": TODAY,
        "panels": PANELS,
        "totals": {"panels": len(PANELS), "circuits": ncirc,
                   "connected_kw": round(total_kw, 1)},
    }


# --- CSV -------------------------------------------------------------------
def write_csv(ds, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(SCHED_COLS)
        for p in ds["panels"]:
            for c in p["circuits"]:
                w.writerow([p["id"], p["location"], p["supply"], c["circuit"],
                            c["load"], c["kw"] if c["kw"] is not None else "",
                            c["device"], c["cable"], c["control"], c["work"]])


# --- XLSX ------------------------------------------------------------------
def write_xlsx(ds, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    PANEL_FILL = PatternFill("solid", fgColor="D6DCE4")
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([ds["title"]]); ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F3864")
    ws.append([ds["subtitle"]]); ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="555555")
    ws.append([f"Generated {ds['generated']}  ·  source: E-01 / E-02 / 13-4 動力盤改造図"])
    ws.append([])
    T = ds["totals"]
    ws.append(["Panels", T["panels"]])
    ws.append(["Circuits", T["circuits"]])
    ws.append(["Connected load (kW)", T["connected_kw"]])
    ws.append([])
    ws.append(["Panel", "Name", "Location", "Supply", "Circuits", "Note"])
    hrow = ws.max_row
    for c in range(1, 7):
        ws.cell(row=hrow, column=c).fill = HDR_FILL; ws.cell(row=hrow, column=c).font = HDR_FONT
    for p in ds["panels"]:
        ws.append([p["id"], p["name"], p["location"], p["supply"], len(p["circuits"]), p["note"]])
    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = [14, 44, 30, 22, 9, 60][i - 1]
    for rr in range(hrow, ws.max_row + 1):
        for c in range(1, 7):
            ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    # Schedule (grouped, outline)
    ws = wb.create_sheet("Schedule")
    ws.append(SCHED_COLS)
    for c in range(1, len(SCHED_COLS) + 1):
        ws.cell(row=1, column=c).fill = HDR_FILL; ws.cell(row=1, column=c).font = HDR_FONT
    ws.sheet_properties.outlinePr.summaryBelow = False
    for p in ds["panels"]:
        ws.append([f"{p['id']} — {p['name']}", p["location"], p["supply"], "", "", "", "", "", "", ""])
        for c in range(1, len(SCHED_COLS) + 1):
            ws.cell(row=ws.max_row, column=c).fill = PANEL_FILL
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
        for c in p["circuits"]:
            ws.append([p["id"], p["location"], p["supply"], c["circuit"], c["load"],
                       c["kw"] if c["kw"] is not None else "", c["device"], c["cable"],
                       c["control"], c["work"]])
            ws.row_dimensions[ws.max_row].outline_level = 1
    widths = [14, 26, 16, 10, 34, 6, 34, 22, 46, 8]
    for i in range(1, len(SCHED_COLS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.freeze_panes = "A2"
    for rr in range(2, ws.max_row + 1):
        for c in range(1, len(SCHED_COLS) + 1):
            ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


# --- HTML ------------------------------------------------------------------
HTML_TEMPLATE = """<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
  :root { --bg:#0f1115; --panel:#171a21; --ink:#e6e8ec; --dim:#9aa2b1; --line:#2a2f3a; --accent:#5b9dd9; }
  * { box-sizing:border-box; }
  body { margin:0; font:14px/1.45 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif; background:var(--bg); color:var(--ink); padding:24px; }
  h1 { font-size:22px; margin:0 0 2px; } .sub { color:var(--dim); margin:0 0 16px; max-width:70ch; }
  .stats { display:flex; gap:22px; flex-wrap:wrap; margin:0 0 18px; }
  .stat b { font-size:22px; display:block; } .stat span { color:var(--dim); font-size:12px; }
  input[type=search] { background:var(--panel); border:1px solid var(--line); color:var(--ink); padding:7px 10px; border-radius:7px; min-width:280px; font-size:13px; margin:0 0 14px; }
  details.panel { border:1px solid var(--line); border-radius:9px; margin:0 0 10px; background:var(--panel); }
  details.panel > summary { padding:11px 14px; font-weight:600; cursor:pointer; list-style:none; }
  summary::-webkit-details-marker { display:none; }
  summary .r { float:right; color:var(--dim); font-weight:400; font-size:12px; }
  .meta { color:var(--dim); font-size:12px; margin:2px 0 8px 16px; }
  table { border-collapse:collapse; width:calc(100% - 16px); margin:0 0 12px 16px; font-size:12.5px; }
  th, td { border:1px solid var(--line); padding:6px 9px; text-align:left; vertical-align:top; }
  thead th { background:#20242e; } td.kw { text-align:right; white-space:nowrap; }
  .new { color:#57b894; } .mod { color:#e0a458; }
  .hidden { display:none !important; }
  .foot { color:var(--dim); font-size:12px; margin-top:20px; max-width:80ch; }
</style></head>
<body>
<h1>__TITLE__</h1>
<p class="sub">__SUBTITLE__</p>
<div class="stats" id="stats"></div>
<input type="search" id="q" placeholder="filter circuit / load / device / cable\u2026">
<div id="tree"></div>
<p class="foot">Generated __GENERATED__ &middot; source drawings: E-01 動力制御盤改修図, E-02 幹線動力設備平面図, 13-4 動力盤改造図 (2015-04-30, (株)ヤマグチ). Scope: ESCO retrofit power circuits only.</p>
<script>
const DATA = __DATA__;
function esc(s){return (''+(s??'')).replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));}
const T=DATA.totals;
document.getElementById('stats').innerHTML=[['Panels',T.panels],['Circuits',T.circuits],['Connected kW',T.connected_kw]]
 .map(([l,v])=>`<div class="stat"><b>${v}</b><span>${l}</span></div>`).join('');
let h='';
for(const p of DATA.panels){
  h+=`<details class="panel" open><summary>${esc(p.id)} \u2014 ${esc(p.name)}<span class="r">${p.circuits.length} circuits</span></summary>`
   +`<div class="meta">${esc(p.location)} \u00b7 ${esc(p.supply)} \u2014 ${esc(p.note)}</div>`
   +'<table><thead><tr><th>Circuit</th><th>Load / served</th><th>kW</th><th>Protective device</th><th>Feeder cable</th><th>Metering / control</th><th>Work</th></tr></thead><tbody>';
  for(const c of p.circuits){
    const blob=(c.circuit+' '+c.load+' '+c.device+' '+c.cable+' '+c.control).toLowerCase();
    const wc=c.work==='New'?'new':(c.work==='Modify'?'mod':'');
    h+=`<tr data-s="${esc(blob)}"><td>${esc(c.circuit)}</td><td>${esc(c.load)}</td>`
     +`<td class="kw">${c.kw??''}</td><td>${esc(c.device)}</td><td>${esc(c.cable)}</td>`
     +`<td>${esc(c.control)}</td><td class="${wc}">${esc(c.work)}</td></tr>`;
  }
  h+='</tbody></table></details>';
}
document.getElementById('tree').innerHTML=h;
document.getElementById('q').addEventListener('input',e=>{
  const s=e.target.value.trim().toLowerCase();
  for(const p of document.querySelectorAll('details.panel')){
    let any=false;
    for(const tr of p.querySelectorAll('tbody tr')){const hit=!s||tr.dataset.s.includes(s);tr.classList.toggle('hidden',!hit);any=any||hit;}
    p.classList.toggle('hidden',!any); p.open=true;
  }
});
</script>
</body></html>
"""


def write_html(ds, path):
    out = (HTML_TEMPLATE
           .replace("__TITLE__", html.escape(ds["title"]))
           .replace("__SUBTITLE__", html.escape(ds["subtitle"]))
           .replace("__GENERATED__", ds["generated"])
           .replace("__DATA__", json.dumps(ds, ensure_ascii=False)))
    for e, ch in (("\\u2014", "\u2014"), ("\\u00b7", "\u00b7"), ("\\u2192", "\u2192"), ("\\u2026", "\u2026")):
        out = out.replace(e, ch)
    with open(path, "w", encoding="utf-8") as f:
        f.write(out)


if __name__ == "__main__":
    os.makedirs(EXPORT_DIR, exist_ok=True)
    ds = dataset()
    stem = os.path.join(EXPORT_DIR, "red5-dhcp_switchboard")
    write_csv(ds, stem + ".csv")
    write_xlsx(ds, stem + ".xlsx")
    write_html(ds, stem + ".html")
    print(f"switchboard: panels={ds['totals']['panels']} circuits={ds['totals']['circuits']} "
          f"kW={ds['totals']['connected_kw']}")
