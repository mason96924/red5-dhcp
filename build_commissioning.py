"""build_commissioning.py -- per-panel point-to-point commissioning sheets.

Driven by the AUTHORITATIVE physical panel schedule (panels_schedule.json,
transcribed from 판넬별 포인트 정리(델타 컨트롤러 및 모듈 포함)_26.07.29.xlsx), so the
checklist carries a tick-box for every physically wired point in every panel
(4,467 points across 67 panels), not just the functionally-characterized subset.

For each panel we enumerate its used points by type -- DO / DI / BTOT / AI / AO --
into line items (e.g. RS-PH2-1-DI01 ... DI42) with tick-box columns
(wired / terminated / P2P verified / functional test) plus observed-value and
notes fields, in three portable forms under exports/:

  red5-dhcp_commissioning.xlsx   Index + one worksheet per panel
  red5-dhcp_commissioning.html   interactive: clickable checkboxes, search,
                                 one print page per panel (print-to-PDF)
  red5-dhcp_commissioning.csv    combined flat sheet (Panel columns)

Carried by the field crew during each panel/controller swap (see the cutover plan).
Note: the schedule gives per-panel TYPE COUNTS, not vendor tags -- the enumerated
tags are positional (panel + type + index). Where a point overlaps the functional
model, its real tag/description lives in red5-dhcp_full / _control-logic.
"""
from __future__ import annotations

import csv
import datetime
import html
import json
import os

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(HERE, "exports")
SCHEDULE = os.path.join(HERE, "panels_schedule.json")
TODAY = datetime.date.today().isoformat()

CHECK_COLS = ["Wired", "Terminated", "P2P verified", "Func. test"]
BASE_COLS = ["#", "Point Tag", "Type", "Description", "Signal / Range", "Units"]
TAIL_COLS = ["Observed value / state", "Notes"]
BOX = "\u2610"  # ballot box

AREA_ORDER = [
    "Penthouse (PH)", "Floor 31", "Floor 21", "Floor 10", "Floor 4", "Floor 3",
    "Floor 2", "Floor 1", "Basement B1", "Basement B2", "Basement B3",
    "Heat source (R-1)", "System / central", "Plant / misc",
]

# type -> (long name, signal/range, units)
TYPE_META = {
    "DO":   ("Digital output", "Relay - 0 / 1", "on-off"),
    "DI":   ("Digital input", "Dry contact - 0 / 1", "status"),
    "BTOT": ("Pulse / totalizer", "Pulse accumulate", "kWh - m3"),
    "AI":   ("Analog input", "4-20 mA / 0-10 V / RTD", "eng. unit"),
    "AO":   ("Analog output", "0-10 V / 4-20 mA", "%"),
}
TYPE_ORDER = ["DO", "DI", "BTOT", "AI", "AO"]


def load_panels():
    d = json.load(open(SCHEDULE, encoding="utf-8"))
    panels = d["panels"]
    order = {a: i for i, a in enumerate(AREA_ORDER)}
    panels.sort(key=lambda p: (order.get(p["area"], 99), int(p["no"])))
    return panels, d


def panel_points(p):
    """Enumerate positional point rows for one panel: (tag, type, desc, sig, units)."""
    u = p["used"]
    rows = []
    for t in TYPE_ORDER:
        n = u.get(t, 0)
        if not n:
            continue
        longname, sig, units = TYPE_META[t]
        for k in range(1, n + 1):
            tag = f"{p['name']}-{t}{k:02d}"
            rows.append((tag, t, f"{longname} #{k}", sig, units))
    return rows


def panels_with_points(panels):
    out = []
    for p in panels:
        pts = panel_points(p)
        if pts:
            out.append((p, pts))
    return out


# ---------------------------------------------------------------------------
# CSV (combined)
# ---------------------------------------------------------------------------
def write_csv(path, data):
    total = 0
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Panel", "Area", "Panel used/cap"] + BASE_COLS[1:] + CHECK_COLS + TAIL_COLS)
        for p, pts in data:
            cap = f"{p['usedTot']}/{p['capTot']}"
            for i, (tag, t, desc, sig, units) in enumerate(pts, 1):
                w.writerow([p["name"], p["area"], cap, tag, t, desc, sig, units,
                            BOX, BOX, BOX, BOX, "", ""])
                total += 1
    return total


# ---------------------------------------------------------------------------
# XLSX (index + sheet per panel)
# ---------------------------------------------------------------------------
def write_xlsx(path, data, meta):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HDR = PatternFill("solid", fgColor="1F3864")
    HF = Font(bold=True, color="FFFFFF", size=10)
    CHK = PatternFill("solid", fgColor="FFF6D9")
    thin = Side(style="thin", color="D9D9D9")
    B = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center")
    top = Alignment(wrap_text=True, vertical="top")

    cols = BASE_COLS + CHECK_COLS + TAIL_COLS
    widths = [4, 20, 7, 22, 22, 11] + [10, 11, 12, 10] + [22, 24]
    total_pts = sum(len(pts) for _, pts in data)

    wb = Workbook()
    idx = wb.active
    idx.title = "Index"
    idx.append(["Red5-DHCP - per-panel commissioning checklists (physical schedule)"])
    idx.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F3864")
    idx.append([f"Generated {TODAY}  -  {len(data)} panels  -  {total_pts} points  "
                f"(schedule: {meta['usedTotal']} used / {meta['capTotal']} capacity)"])
    idx.cell(row=2, column=1).font = Font(italic=True, size=10, color="555555")
    idx.append([])
    idx.append(["#", "Panel", "Area", "Used", "Cap", "Sheet"])
    hrow = idx.max_row
    for c in range(1, 7):
        idx.cell(row=hrow, column=c).fill = HDR
        idx.cell(row=hrow, column=c).font = HF

    def sheet_name(p):
        base = p["name"].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "")
        base = base.replace("[", "(").replace("]", ")").replace(":", "-")
        return (f"{p['no']}-{base}")[:31]

    used_names = set()
    names = {}
    for p, pts in data:
        nm = sheet_name(p)
        while nm in used_names:
            nm = (nm[:29] + "~")[:31]
        used_names.add(nm)
        names[p["no"]] = nm

    for p, pts in data:
        nm = names[p["no"]]
        rr = idx.max_row + 1
        idx.append([p["no"], p["name"], p["area"], p["usedTot"], p["capTot"], nm])
        link = idx.cell(row=rr, column=6)
        link.hyperlink = f"#'{nm}'!A1"
        link.font = Font(color="1F5FBF", underline="single")
    for i, wdt in enumerate([5, 20, 18, 7, 7, 24], 1):
        idx.column_dimensions[get_column_letter(i)].width = wdt
    for rr in range(hrow, idx.max_row + 1):
        for c in range(1, 7):
            idx.cell(row=rr, column=c).border = B
            idx.cell(row=rr, column=c).alignment = top
    idx.freeze_panes = "A5"

    for p, pts in data:
        ws = wb.create_sheet(names[p["no"]])
        ws.append([f"Commissioning checklist - {p['name']}"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F3864")
        ws.append([f"Panel #{p['no']} - {p['area']}", "",
                   f"Used {p['usedTot']} / capacity {p['capTot']}"])
        if p.get("note"):
            ws.append([f"Note: {p['note']}"])
        ws.append(["Technician: ______________________", "", "Date: __________",
                   "", "Signature: ______________________"])
        ws.append([])
        hr = ws.max_row + 1
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(row=hr, column=c).fill = HDR
            ws.cell(row=hr, column=c).font = HF
            ws.cell(row=hr, column=c).alignment = ctr
        chk_start = len(BASE_COLS) + 1
        chk_end = len(BASE_COLS) + len(CHECK_COLS)
        for i, (tag, t, desc, sig, units) in enumerate(pts, 1):
            ws.append([i, tag, t, desc, sig, units, BOX, BOX, BOX, BOX, "", ""])
            rr = ws.max_row
            for c in range(chk_start, chk_end + 1):
                cell = ws.cell(row=rr, column=c)
                cell.fill = CHK
                cell.alignment = ctr
                cell.font = Font(size=12)
        for i, wdt in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = wdt
        for rr in range(hr, ws.max_row + 1):
            for c in range(1, len(cols) + 1):
                ws.cell(row=rr, column=c).border = B
                if not (chk_start <= c <= chk_end):
                    ws.cell(row=rr, column=c).alignment = top
        ws.freeze_panes = f"A{hr + 1}"

    wb.save(path)
    return total_pts


# ---------------------------------------------------------------------------
# HTML (interactive; one print page per panel)
# ---------------------------------------------------------------------------
def write_html(path, data, meta):
    def esc(s):
        return html.escape(str(s if s is not None else ""))

    total_pts = sum(len(pts) for _, pts in data)
    blocks = []
    for p, pts in data:
        search_blob = esc((p["name"] + " " + p["area"] + " " + str(p["no"])).lower())
        rows = []
        for i, (tag, t, desc, sig, units) in enumerate(pts, 1):
            rows.append(
                "<tr><td class=c>{i}</td><td>{tag}</td><td class=c>{t}</td><td>{d}</td>"
                "<td>{sig}</td><td>{u}</td>"
                "<td class=c><input type=checkbox></td><td class=c><input type=checkbox></td>"
                "<td class=c><input type=checkbox></td><td class=c><input type=checkbox></td>"
                "<td><input class=t type=text></td><td><input class=t type=text></td></tr>".format(
                    i=i, tag=esc(tag), t=esc(t), d=esc(desc), sig=esc(sig), u=esc(units)))
        note = f'<div>Note: {esc(p["note"])}</div>' if p.get("note") else ""
        blocks.append(
            f'<details class="ctrl" data-s="{search_blob}"><summary>{esc(p["name"])}'
            f'<span class="r">#{esc(p["no"])} - {esc(p["area"])} - {len(pts)} points</span></summary>'
            f'<div class="hd"><div>Panel #{esc(p["no"])} - {esc(p["area"])} - '
            f'used {p["usedTot"]} / capacity {p["capTot"]}</div>{note}'
            f'<div class="sign">Technician <input class=t type=text> &nbsp; Date <input class=t type=text>'
            f' &nbsp; Signature <input class=t type=text></div></div>'
            '<table><thead><tr><th class=c>#</th><th>Point Tag</th><th class=c>Type</th><th>Description</th>'
            '<th>Signal / Range</th><th>Units</th>'
            '<th class=c>Wired</th><th class=c>Term.</th><th class=c>P2P</th><th class=c>Func.</th>'
            '<th>Observed</th><th>Notes</th></tr></thead>'
            f'<tbody>{"".join(rows)}</tbody></table></details>')

    doc = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Red5-DHCP - commissioning checklists</title>
<style>
  :root {{ --ink:#1b2430; --dim:#5b6675; --line:#d5dbe4; --band:#1F3864; }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; font:13px/1.45 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;
          color:var(--ink); background:#f4f6f9; padding:24px; }}
  .wrap {{ max-width:1200px; margin:0 auto; }}
  h1 {{ font-size:21px; margin:0 0 2px; }} .sub {{ color:var(--dim); margin:0 0 14px; }}
  .bar {{ position:sticky; top:0; background:#f4f6f9; padding:8px 0 12px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; z-index:5; }}
  input[type=search] {{ border:1px solid var(--line); border-radius:7px; padding:7px 10px; min-width:280px; font-size:13px; }}
  button {{ border:1px solid var(--line); background:#fff; border-radius:7px; padding:6px 11px; cursor:pointer; font-size:12px; }}
  details.ctrl {{ background:#fff; border:1px solid var(--line); border-radius:9px; margin:0 0 10px; }}
  summary {{ padding:11px 14px; font-weight:600; cursor:pointer; list-style:none; }}
  summary::-webkit-details-marker {{ display:none; }}
  summary .r {{ float:right; color:var(--dim); font-weight:400; font-size:12px; }}
  .hd {{ color:var(--dim); font-size:12px; padding:0 14px 8px; }}
  .hd .sign {{ margin-top:6px; color:var(--ink); }}
  table {{ border-collapse:collapse; width:100%; font-size:12px; }}
  th, td {{ border:1px solid var(--line); padding:5px 7px; text-align:left; vertical-align:top; }}
  thead th {{ background:var(--band); color:#fff; position:sticky; top:52px; }}
  td.c, th.c {{ text-align:center; white-space:nowrap; }}
  input.t {{ width:100%; border:0; border-bottom:1px solid var(--line); font-size:12px; padding:2px 0; }}
  input[type=checkbox] {{ width:16px; height:16px; }}
  .hidden {{ display:none !important; }}
  .foot {{ color:var(--dim); font-size:12px; margin-top:18px; }}
  @media print {{
    body {{ background:#fff; padding:0; }} .bar {{ display:none; }}
    details.ctrl {{ break-inside:avoid; page-break-after:always; border:0; }}
    details.ctrl[open] summary .r {{ color:#000; }}
    details:not([open]) {{ display:none; }}
    thead th {{ position:static; }}
  }}
</style></head>
<body><div class="wrap">
<h1>Red5-DHCP - per-panel commissioning checklists</h1>
<p class="sub">{len(data)} panels - {total_pts} points (physical schedule: {meta['usedTotal']} used / {meta['capTotal']} capacity). Point-to-point verify each I/O, then trend 24-48 h before returning to auto. Fill on a tablet or print to PDF (one page per panel). Tags are positional (panel + type + index) from the 26.07.29 schedule.</p>
<div class="bar">
  <input type="search" id="q" placeholder="find panel / area / no\u2026">
  <button id="expand">Expand all</button>
  <button id="collapse">Collapse all</button>
  <button onclick="window.print()">Print / Save PDF</button>
  <span class="sub" id="count"></span>
</div>
{"".join(blocks)}
<p class="foot">Generated {TODAY} - source: panels_schedule.json (판넬별 포인트 정리 26.07.29) - Red5-DHCP savic-net FX2 -> new DDC migration.</p>
</div>
<script>
const items=[...document.querySelectorAll('details.ctrl')];
document.getElementById('count').textContent=items.length+' panels';
document.getElementById('q').addEventListener('input',e=>{{
  const s=e.target.value.trim().toLowerCase();
  let n=0;
  for(const d of items){{const hit=!s||d.dataset.s.includes(s);d.classList.toggle('hidden',!hit);if(hit)n++;d.open=!!s&&hit;}}
  document.getElementById('count').textContent=n+' / '+items.length+' panels';
}});
document.getElementById('expand').onclick=()=>items.forEach(d=>d.open=true);
document.getElementById('collapse').onclick=()=>items.forEach(d=>d.open=false);
</script>
</body></html>
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)
    return total_pts


if __name__ == "__main__":
    os.makedirs(EXPORT_DIR, exist_ok=True)
    panels, meta = load_panels()
    data = panels_with_points(panels)
    stem = os.path.join(EXPORT_DIR, "red5-dhcp_commissioning")
    n_csv = write_csv(stem + ".csv", data)
    n_xlsx = write_xlsx(stem + ".xlsx", data, meta)
    n_html = write_html(stem + ".html", data, meta)
    assert n_csv == n_xlsx == n_html, (n_csv, n_xlsx, n_html)
    print(f"commissioning sheets: {len(data)} panels, {n_csv} points "
          f"(schedule {meta['usedTotal']} used / {meta['capTotal']} cap) -> {stem}.{{csv,xlsx,html}}")
