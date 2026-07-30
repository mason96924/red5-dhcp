"""build_panels_io.py -- downloadable physical panel schedule (Panels · Controllers · I/O).

Reads panels_schedule.json (transcribed from 판넬별 포인트 정리 델타 컨트롤러 및 모듈 포함,
rev in the file's `source`) and emits, under exports/:

  Red5-DHCP-Panels-IO.csv    flat 67-panel schedule (+ TOTAL), UTF-8-BOM
  Red5-DHCP-Panels-IO.xlsx   `Panel schedule` + `Summary` sheets
  Red5-DHCP-Panels-IO.html   self-contained page (stats, reconciliation, by-area tables)

Each panel row carries floor, equipment served, Delta module counts
(ROOM/CON-T1L/EXPAND-04/4F4xP/8xP/8PxP), used points by type (DO/DI/BTOT/AI/AO),
used total and installed capacity. Mirrors the dhcp-panels-io-pivot canvas.
"""
from __future__ import annotations

import csv
import datetime
import html
import json
import os

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(HERE, "exports")
SCHEDULE = os.path.join(HERE, "panels_schedule.json")
GEN = datetime.date.today().isoformat()

_d = json.load(open(SCHEDULE, encoding="utf-8"))
PANELS = _d["panels"]
T = _d["typeTotals"]
MOD = _d["modTotals"]
USED = _d["usedTotal"]
CAP = _d["capTotal"]
SRC = _d.get("source", "")
OVERLAP = 1160
NEWTBD = USED - OVERLAP
SPARE = CAP - USED
MODSUM = sum(MOD)

ORDER = ["Penthouse (PH)", "Floor 31", "Floor 21", "Floor 10", "Floor 4", "Floor 3",
         "Floor 2", "Floor 1", "Basement B1", "Basement B2", "Basement B3"]
MOD_LABELS = ["Red5-PLUS-ROOM", "CON-T1L", "Red5-EXPAND-04", "Red5-MODULE-4F4xP", "Red5-MODULE-8xP", "Red5-MODULE-8PxP"]
COLS = ["No", "Panel", "Floor", "Equipment served", "ROOM", "CON-T1L", "EXPAND-04", "4F4xP", "8xP", "8PxP",
        "DO", "DI", "BTOT", "AI", "AO", "Used", "Cap", "Note"]
STAMP = f"{SRC} panel schedule (판넬별 포인트 정리 · 델타 컨트롤러 및 모듈 포함)"


def row_of(p):
    u = p["used"]
    return [p["no"], p["name"], p["floor"], p["equip"], *p["mods"],
            u["DO"], u["DI"], u["BTOT"], u["AI"], u["AO"], p["usedTot"], p["capTot"], p["note"]]


by_area = {a: [p for p in PANELS if p["area"] == a] for a in ORDER}
ordered = [p for a in ORDER for p in by_area[a]]


def write_csv(path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([f"Red5-DHCP — Panels · Controllers · I/O   ({STAMP})"])
        w.writerow([f"{len(PANELS)} panels · {MODSUM} Delta controllers+modules · {USED} used pts · {CAP} capacity · "
                    f"overlap {OVERLAP} / new-TBD {NEWTBD} / spare {SPARE}"])
        w.writerow([])
        w.writerow(COLS)
        for p in ordered:
            w.writerow(row_of(p))
        w.writerow(["TOTAL", "", "", "", *MOD, T["DO"], T["DI"], T["BTOT"], T["AI"], T["AO"], USED, CAP, ""])


def write_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HDR = PatternFill("solid", fgColor="1F3864")
    HF = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    B = Border(left=thin, right=thin, top=thin, bottom=thin)
    top = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    ws.title = "Panel schedule"
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        ws.cell(row=1, column=c).fill = HDR
        ws.cell(row=1, column=c).font = HF
    for p in ordered:
        ws.append(row_of(p))
    ws.append(["TOTAL", "", "", "", *MOD, T["DO"], T["DI"], T["BTOT"], T["AI"], T["AO"], USED, CAP, ""])
    for c in range(1, len(COLS) + 1):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    widths = [5, 14, 7, 26, 7, 8, 10, 7, 6, 6, 6, 6, 7, 6, 6, 7, 6, 34]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    for rr in range(1, ws.max_row + 1):
        for c in range(1, len(COLS) + 1):
            ws.cell(row=rr, column=c).border = B
            ws.cell(row=rr, column=c).alignment = top
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    rows = [
        ["Red5-DHCP — reconciliation summary", ""],
        ["Source", STAMP],
        ["Generated", GEN],
        ["", ""],
        ["Metric", "Value"],
        ["Physical panels", len(PANELS)],
        ["Delta controllers + modules", MODSUM],
        ["Used I/O points", USED],
        ["Installed capacity", CAP],
        ["Overlap (characterized, purpose known)", OVERLAP],
        ["New — purpose TBD", NEWTBD],
        ["Spare to capacity", SPARE],
        ["Panels naming equipment served", sum(1 for p in PANELS if p["equip"])],
        ["", ""],
        ["Delta module type", "Count"],
    ] + [[MOD_LABELS[i], MOD[i]] for i in range(6)] + [
        ["", ""],
        ["Point type", "Used count"],
        ["DO (binary output)", T["DO"]],
        ["DI (binary input)", T["DI"]],
        ["BTOT (pulse/totalizer)", T["BTOT"]],
        ["AI (analog input)", T["AI"]],
        ["AO (analog output)", T["AO"]],
    ]
    for r in rows:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 60
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F3864")

    wb.save(path)


def _bar(segs, total):
    parts = []
    for lab, val, col in segs:
        pct = 100.0 * val / total if total else 0
        parts.append(f'<div class="seg" style="width:{pct:.3f}%;background:{col}" title="{lab}: {val}"></div>')
    return '<div class="bar">' + "".join(parts) + "</div>"


def _esc(x):
    return html.escape(str(x))


def write_html(path):
    area_sections = []
    for a in ORDER:
        rows = by_area[a]
        if not rows:
            continue
        aUsed = sum(p["usedTot"] for p in rows)
        aCap = sum(p["capTot"] for p in rows)
        trs = []
        for p in rows:
            u = p["used"]
            trs.append("<tr>" + "".join(f"<td>{_esc(v)}</td>" for v in
                [p["no"], p["name"], p["floor"], p["equip"], *p["mods"], u["DO"], u["DI"], u["BTOT"],
                 u["AI"], u["AO"], p["usedTot"], p["capTot"], p["note"]]) + "</tr>")
        head = "".join(f"<th>{_esc(h)}</th>" for h in
            ["No", "Panel", "Floor", "Equip. served", "ROOM", "CON", "EXP", "4F4xP", "8xP", "8PxP",
             "DO", "DI", "BTOT", "AI", "AO", "Used", "Cap", "Note"])
        area_sections.append(f'''
    <details open>
      <summary><b>{_esc(a)}</b> <span class="muted">· {len(rows)} panels · {aUsed} used / {aCap} cap</span></summary>
      <table class="grid"><thead><tr>{head}</tr></thead><tbody>{''.join(trs)}</tbody></table>
    </details>''')

    hw_rows = "".join(f"<tr><td>{_esc(MOD_LABELS[i])}</td><td class='r'>{MOD[i]}</td></tr>" for i in range(6))
    recon_bar = _bar([("Overlap (characterized)", OVERLAP, "#3fa266"),
                      ("New — purpose TBD", NEWTBD, "#ea7a3c"),
                      ("Spare capacity", SPARE, "#d7dde5")], CAP)
    type_bar = _bar([("DO", T["DO"], "#7c5cd6"), ("DI", T["DI"], "#2563eb"),
                     ("BTOT", T["BTOT"], "#e0af22"), ("AI", T["AI"], "#3fa266"),
                     ("AO", T["AO"], "#ea7a3c")], USED)
    n_equip = sum(1 for p in PANELS if p["equip"])

    doc = f'''<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Red5-DHCP — Panels · Controllers · I/O</title>
<style>
  :root {{ --ink:#1b2430; --muted:#64748b; --line:#e6eaf0; --bg:#f7f9fc; }}
  * {{ box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif; color:var(--ink);
         margin:0; background:var(--bg); }}
  .wrap {{ max-width:1240px; margin:0 auto; padding:28px 22px 60px; }}
  h1 {{ font-size:26px; margin:0 0 4px; }}
  h2 {{ font-size:19px; margin:26px 0 8px; }}
  h3 {{ font-size:15px; margin:0 0 8px; }}
  .sub {{ color:var(--muted); font-size:13.5px; margin-bottom:18px; }}
  .stats {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin:14px 0; }}
  .stat {{ background:#fff; border:1px solid var(--line); border-radius:12px; padding:14px 16px; }}
  .stat .v {{ font-size:26px; font-weight:700; }}
  .stat .l {{ color:var(--muted); font-size:12.5px; margin-top:2px; }}
  .stat.info .v {{ color:#2563eb; }}
  .callout {{ border-radius:12px; padding:14px 16px; margin:14px 0; font-size:13.5px; line-height:1.5; }}
  .warn {{ background:#fff7ed; border:1px solid #fed7aa; }}
  .infoc {{ background:#eff6ff; border:1px solid #bfdbfe; }}
  .callout b.t {{ display:block; font-size:14px; margin-bottom:5px; }}
  .bar {{ display:flex; height:22px; border-radius:7px; overflow:hidden; border:1px solid var(--line); }}
  .seg {{ height:100%; }}
  .barlbl {{ display:flex; justify-content:space-between; font-size:12.5px; color:var(--muted); margin:14px 0 5px; }}
  .two {{ display:grid; grid-template-columns:1fr 1fr; gap:18px; align-items:start; }}
  table {{ border-collapse:collapse; width:100%; background:#fff; }}
  table.grid {{ font-size:12px; margin-top:6px; }}
  th,td {{ border:1px solid var(--line); padding:5px 7px; text-align:right; }}
  th {{ background:#f0f3f8; position:sticky; top:0; }}
  td:nth-child(2), th:nth-child(2), td:nth-child(4), th:nth-child(4), td:last-child, th:last-child {{ text-align:left; }}
  td:nth-child(3), th:nth-child(3) {{ text-align:center; }}
  table.kv td {{ text-align:left; }} table.kv td.r {{ text-align:right; }}
  details {{ background:#fff; border:1px solid var(--line); border-radius:12px; padding:10px 14px; margin:8px 0; }}
  summary {{ cursor:pointer; font-size:14px; }}
  .muted {{ color:var(--muted); font-weight:400; font-size:12.5px; }}
  .foot {{ color:var(--muted); font-size:12px; margin-top:30px; }}
</style></head>
<body><div class="wrap">
  <h1>Red5-DHCP — Panels · Controllers · I/O</h1>
  <div class="sub">Authoritative physical panel schedule (Delta controllers &amp; modules) — with each panel's
    floor and equipment served — reconciled against the prior functional point list.
    Source: {_esc(STAMP)} · generated {GEN}.</div>

  <div class="stats">
    <div class="stat"><div class="v">{len(PANELS)}</div><div class="l">Physical panels</div></div>
    <div class="stat"><div class="v">{MODSUM}</div><div class="l">Delta controllers + modules</div></div>
    <div class="stat info"><div class="v">{USED}</div><div class="l">Used I/O points</div></div>
    <div class="stat"><div class="v">{CAP}</div><div class="l">Installed capacity</div></div>
  </div>

  <div class="callout warn"><b class="t">Reconciliation — overlap vs newly identified</b>
    The physical schedule lists {USED} used I/O points across {len(PANELS)} panels ({MODSUM} Delta
    controllers + modules; {CAP} points of installed capacity). Of these, {OVERLAP} match points already
    characterized in the prior functional list (overlap — purpose known); the remaining {NEWTBD} are newly
    surfaced. This {SRC} revision also names the equipment served for {n_equip} panels, tying RCP/CP panels
    directly to the functional model.</div>

  <div class="barlbl"><span>Used-point reconciliation — {USED} used / {CAP} capacity</span>
    <span>Overlap {OVERLAP} · New/TBD {NEWTBD} · Spare {SPARE}</span></div>
  {recon_bar}

  <div class="two" style="margin-top:22px">
    <div>
      <h3>Delta hardware rollup</h3>
      <table class="kv"><thead><tr><th style="text-align:left">Controller / module</th><th>Count</th></tr></thead>
      <tbody>{hw_rows}</tbody></table>
    </div>
    <div>
      <h3>Used point-type mix</h3>
      <div class="barlbl"><span>{USED} used points</span>
        <span>DO {T["DO"]} · DI {T["DI"]} · BTOT {T["BTOT"]} · AI {T["AI"]} · AO {T["AO"]}</span></div>
      {type_bar}
      <div class="muted" style="margin-top:8px">Type key: DO ≈ binary output (BO), DI ≈ binary input (BI),
        BTOT = pulse/totalizer meter points (kWh · flow), AI/AO analog.</div>
    </div>
  </div>

  <h2>Physical panel schedule — by area / floor</h2>
  <div class="muted" style="margin-bottom:6px">Per panel: floor, equipment served, Delta module count
    (ROOM / CON-T1L / EXPAND-04 / 4F4xP / 8xP / 8PxP), used points by type, used total, installed capacity.</div>
  {''.join(area_sections)}

  <div class="foot">Red5-DHCP · {_esc(STAMP)} · totals reconcile to the schedule TOTAL row
    ({len(PANELS)} panels · {MODSUM} modules · {USED} used · {CAP} capacity).</div>
</div></body></html>'''
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)


if __name__ == "__main__":
    os.makedirs(EXPORT_DIR, exist_ok=True)
    stem = os.path.join(EXPORT_DIR, "Red5-DHCP-Panels-IO")
    write_csv(stem + ".csv")
    write_xlsx(stem + ".xlsx")
    write_html(stem + ".html")
    print(f"panels-io exported: {len(PANELS)} panels / {USED} used / {CAP} cap (source {SRC}) "
          f"-> {stem}.{{csv,xlsx,html}}")
