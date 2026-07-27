"""build_exports.py -- downloadable pivots (CSV / XLSX / HTML) from the BMS model.

Produces, per dataset, three shippable files under ``exports/``:
  - <name>.csv   flat point list
  - <name>.xlsx  Summary + collapsible Panel>Controller>Point outline pivot + Flat
  - <name>.html  self-contained interactive pivot (search + expand/collapse)

Datasets:
  - full          every point in the model
  - lighting      LCP-LTG / System=Lighting (照明一覧 groups)
  - electrical    electrical cross-cut: energy meters, device kW power, and
                  electrical-protection (trip/overload/phase/demand) points
The switchboard/MCC schedule (from the E-drawings) is built separately in
build_switchboard.py and exported with the same writers here.
"""
from __future__ import annotations

import csv
import datetime
import html
import json
import os

import generate_io_list as g
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(HERE, "exports")
TODAY = datetime.date.today().isoformat()

IO_KEYS = ("AI", "AO", "BI", "BO")
FLAT_COLS = ["Panel", "Panel description", "Controller", "Device ID",
             "Point Tag", "I/O", "System", "Units", "Description", "Notes / Basis"]


# ---------------------------------------------------------------------------
# Dataset assembly (panels -> controllers -> points, mirrors the canvas model)
# ---------------------------------------------------------------------------
def _io_zero():
    return {k: 0 for k in IO_KEYS}


def build_dataset(title, subtitle, predicate):
    ctrl_panel = {c[0]: c[1] for c in g.controllers}
    ctrl_devices = {c[0]: c[2] for c in g.controllers}
    ctrl_order = [c[0] for c in g.controllers]

    panel_desc = {p[0]: p[1] for p in g.PANELS}
    panel_loc = {p[0]: p[2] for p in g.PANELS}

    sel = [r for r in g.rows if predicate(r)]

    ctrl_points = {}
    for r in sel:
        ctrl_points.setdefault(r["Controller"], []).append(r)

    panels = []
    totals = _io_zero()
    dev_set = set()
    panel_ids = [p[0] for p in g.PANELS]
    for pid in panel_ids:
        cids = [c for c in ctrl_order if ctrl_panel.get(c) == pid and ctrl_points.get(c)]
        if not cids:
            continue
        pio = _io_zero()
        npts = 0
        controllers = []
        for cid in cids:
            pts = ctrl_points.get(cid, [])
            cio = _io_zero()
            for r in pts:
                cio[r["I/O Type"]] = cio.get(r["I/O Type"], 0) + 1
                dev_set.add(r["Device ID"])
            for k in IO_KEYS:
                pio[k] += cio[k]
            npts += len(pts)
            controllers.append({
                "id": cid, "devices": ctrl_devices.get(cid, ""),
                "nPoints": len(pts), "io": cio,
                "points": [{
                    "tag": r["Point Tag"], "io": r["I/O Type"], "sys": r["System"],
                    "dev": r["Device ID"], "units": r["Units"] or "",
                    "desc": r["Point Description"], "notes": r["Notes / Basis"] or "",
                } for r in pts],
            })
        for k in IO_KEYS:
            totals[k] += pio[k]
        panels.append({
            "id": pid, "description": panel_desc.get(pid, ""),
            "location": panel_loc.get(pid, ""), "nctrl": len(cids),
            "nPoints": npts, "io": pio, "controllers": controllers,
        })

    return {
        "title": title, "subtitle": subtitle, "generated": TODAY,
        "totals": {"panels": len(panels),
                   "controllers": sum(p["nctrl"] for p in panels),
                   "devices": len(dev_set), "points": len(sel), **totals},
        "panels": panels,
    }


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def write_csv(ds, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(FLAT_COLS)
        for p in ds["panels"]:
            for c in p["controllers"]:
                for pt in c["points"]:
                    w.writerow([p["id"], p["description"], c["id"], pt["dev"],
                                pt["tag"], pt["io"], pt["sys"], pt["units"],
                                pt["desc"], pt["notes"]])


# ---------------------------------------------------------------------------
# XLSX  (Summary + collapsible outline Pivot + Flat)
# ---------------------------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
PANEL_FILL = PatternFill("solid", fgColor="D6DCE4")
CTRL_FILL = PatternFill("solid", fgColor="EDEFF4")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_xlsx(ds, path):
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws.append([ds["title"]])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F3864")
    ws.append([ds["subtitle"]])
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="555555")
    ws.append([f"Generated {ds['generated']}  ·  source: generate_io_list.py"])
    ws.append([])
    T = ds["totals"]
    ws.append(["Panels", T["panels"]])
    ws.append(["Controllers", T["controllers"]])
    ws.append(["Devices", T["devices"]])
    ws.append(["I/O points", T["points"]])
    ws.append([])
    hdr = ["Panel", "Description", "Location", "Ctrls", "AI", "AO", "BI", "BO", "Points"]
    ws.append(hdr)
    hrow = ws.max_row
    for c in range(1, len(hdr) + 1):
        ws.cell(row=hrow, column=c).fill = HDR_FILL
        ws.cell(row=hrow, column=c).font = HDR_FONT
    for p in ds["panels"]:
        ws.append([p["id"], p["description"], p["location"], p["nctrl"],
                   p["io"]["AI"], p["io"]["AO"], p["io"]["BI"], p["io"]["BO"], p["nPoints"]])
    ws.append(["GRAND TOTAL", "", "", T["controllers"],
               T["AI"], T["AO"], T["BI"], T["BO"], T["points"]])
    for c in range(1, len(hdr) + 1):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=c).fill = PANEL_FILL
    _widths(ws, [16, 40, 26, 8, 6, 6, 6, 6, 9])
    for rr in range(hrow, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Pivot (collapsible outline) ----
    ws = wb.create_sheet("Pivot")
    pcols = ["Panel / Controller / Point", "Description / devices", "Device",
             "I/O", "System", "Units"]
    ws.append(pcols)
    _style_header(ws, len(pcols))
    ws.sheet_properties.outlinePr.summaryBelow = False
    for p in ds["panels"]:
        ws.append([p["id"], p["description"], "", "", "", ""])
        for c in range(1, len(pcols) + 1):
            ws.cell(row=ws.max_row, column=c).fill = PANEL_FILL
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
        for ctrl in p["controllers"]:
            ws.append(["    " + ctrl["id"], ctrl["devices"], "", "", "", ""])
            crow = ws.max_row
            ws.row_dimensions[crow].outline_level = 1
            for c in range(1, len(pcols) + 1):
                ws.cell(row=crow, column=c).fill = CTRL_FILL
                ws.cell(row=crow, column=c).font = Font(bold=True, italic=True, size=9)
            for pt in ctrl["points"]:
                ws.append(["        " + pt["tag"], pt["desc"], pt["dev"],
                           pt["io"], pt["sys"], pt["units"]])
                ws.row_dimensions[ws.max_row].outline_level = 2
    _widths(ws, [40, 52, 14, 6, 15, 8])
    ws.freeze_panes = "A2"
    for rr in range(2, ws.max_row + 1):
        for c in range(1, len(pcols) + 1):
            ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Flat ----
    ws = wb.create_sheet("Flat")
    ws.append(FLAT_COLS)
    _style_header(ws, len(FLAT_COLS))
    for p in ds["panels"]:
        for c in p["controllers"]:
            for pt in c["points"]:
                ws.append([p["id"], p["description"], c["id"], pt["dev"],
                           pt["tag"], pt["io"], pt["sys"], pt["units"],
                           pt["desc"], pt["notes"]])
    _widths(ws, [14, 34, 14, 12, 16, 6, 14, 8, 40, 44])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FLAT_COLS))}{ws.max_row}"
    for rr in range(2, ws.max_row + 1):
        for c in range(1, len(FLAT_COLS) + 1):
            ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


# ---------------------------------------------------------------------------
# HTML (self-contained interactive pivot)
# ---------------------------------------------------------------------------
HTML_TEMPLATE = """<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
  :root { --bg:#0f1115; --panel:#171a21; --row:#1c2029; --ink:#e6e8ec; --dim:#9aa2b1;
          --line:#2a2f3a; --accent:#5b9dd9; --ai:#5b9dd9; --ao:#57b894; --bi:#e0a458; --bo:#a982d9; }
  * { box-sizing:border-box; }
  body { margin:0; font:14px/1.45 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;
         background:var(--bg); color:var(--ink); padding:24px; }
  h1 { font-size:22px; margin:0 0 2px; }
  .sub { color:var(--dim); margin:0 0 16px; }
  .stats { display:flex; gap:22px; flex-wrap:wrap; margin:0 0 14px; }
  .stat b { font-size:22px; display:block; }
  .stat span { color:var(--dim); font-size:12px; }
  .mix { height:10px; border-radius:5px; overflow:hidden; display:flex; margin:6px 0 18px; background:var(--line); }
  .mix i { display:block; height:100%; }
  .legend { color:var(--dim); font-size:12px; margin:-12px 0 18px; }
  .legend em { font-style:normal; color:var(--ink); }
  .dot { display:inline-block; width:9px; height:9px; border-radius:2px; margin:0 4px 0 12px; vertical-align:middle; }
  .toolbar { display:flex; gap:10px; align-items:center; margin:0 0 12px; flex-wrap:wrap; }
  input[type=search] { background:var(--panel); border:1px solid var(--line); color:var(--ink);
         padding:7px 10px; border-radius:7px; min-width:280px; font-size:13px; }
  button.f { background:var(--panel); border:1px solid var(--line); color:var(--dim);
         padding:5px 10px; border-radius:20px; cursor:pointer; font-size:12px; }
  button.f.on { color:var(--ink); border-color:var(--accent); }
  table.sum { border-collapse:collapse; width:100%; margin:0 0 22px; font-size:13px; }
  table.sum th, table.sum td { border:1px solid var(--line); padding:6px 9px; text-align:right; }
  table.sum th:nth-child(-n+3), table.sum td:nth-child(-n+3) { text-align:left; }
  table.sum thead th { background:#20242e; }
  table.sum tr.tot td { font-weight:700; background:#20242e; }
  details.panel { border:1px solid var(--line); border-radius:9px; margin:0 0 8px; background:var(--panel); }
  details.panel > summary { padding:11px 14px; font-weight:600; cursor:pointer; list-style:none; }
  details.ctrl { margin:2px 0 2px 18px; }
  details.ctrl > summary { padding:7px 12px; cursor:pointer; list-style:none; color:var(--ink); }
  summary::-webkit-details-marker { display:none; }
  summary .r { float:right; color:var(--dim); font-weight:400; font-size:12px; }
  .serves { color:var(--dim); font-size:12px; margin:2px 0 6px 30px; }
  table.pts { border-collapse:collapse; width:calc(100% - 30px); margin:0 0 10px 30px; font-size:12.5px; }
  table.pts th, table.pts td { border:1px solid var(--line); padding:5px 8px; text-align:left; vertical-align:top; }
  table.pts thead th { background:#20242e; position:sticky; top:0; }
  table.pts td.io { text-align:center; white-space:nowrap; }
  .io-AI{color:var(--ai)} .io-AO{color:var(--ao)} .io-BI{color:var(--bi)} .io-BO{color:var(--bo)}
  .hidden { display:none !important; }
  .foot { color:var(--dim); font-size:12px; margin-top:20px; }
</style></head>
<body>
<h1>__TITLE__</h1>
<p class="sub">__SUBTITLE__</p>
<div class="stats" id="stats"></div>
<div class="mix" id="mix"></div>
<div class="legend">
  <span class="dot" style="background:var(--ai)"></span><em>AI</em> analog in
  <span class="dot" style="background:var(--ao)"></span><em>AO</em> analog out
  <span class="dot" style="background:var(--bi)"></span><em>BI</em> binary in
  <span class="dot" style="background:var(--bo)"></span><em>BO</em> binary out
</div>
<h3>Panel summary</h3>
<table class="sum" id="sum"></table>
<div class="toolbar">
  <strong>Panel &rarr; Controller &rarr; I/O</strong>
  <input type="search" id="q" placeholder="filter tag / device / description / system\u2026">
  <button class="f" id="expand">Expand all</button>
  <button class="f" id="collapse">Collapse all</button>
</div>
<div id="tree"></div>
<p class="foot">Generated __GENERATED__ &middot; source: generate_io_list.py &middot; Red5-DHCP BMS model</p>
<script>
const DATA = __DATA__;
const IOC = {AI:'io-AI',AO:'io-AO',BI:'io-BI',BO:'io-BO'};
function ioText(io){return `AI ${io.AI} \u00b7 AO ${io.AO} \u00b7 BI ${io.BI} \u00b7 BO ${io.BO}`;}
function esc(s){return (s||'').replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));}

const T = DATA.totals;
document.getElementById('stats').innerHTML =
  [['Panels',T.panels],['Controllers',T.controllers],['Devices',T.devices],['I/O points',T.points]]
  .map(([l,v])=>`<div class="stat"><b>${v}</b><span>${l}</span></div>`).join('');
const mixTot = T.AI+T.AO+T.BI+T.BO || 1;
document.getElementById('mix').innerHTML =
  [['AI',T.AI],['AO',T.AO],['BI',T.BI],['BO',T.BO]]
  .map(([k,v])=>`<i style="width:${100*v/mixTot}%;background:var(--${k.toLowerCase()})" title="${k} ${v}"></i>`).join('');

let sh = '<thead><tr><th>Panel</th><th>Description</th><th>Location</th><th>Ctrls</th>'
       + '<th>AI</th><th>AO</th><th>BI</th><th>BO</th><th>Points</th></tr></thead><tbody>';
for(const p of DATA.panels){
  sh += `<tr><td>${esc(p.id)}</td><td>${esc(p.description)}</td><td>${esc(p.location)}</td>`
      + `<td>${p.nctrl}</td><td>${p.io.AI}</td><td>${p.io.AO}</td><td>${p.io.BI}</td>`
      + `<td>${p.io.BO}</td><td>${p.nPoints}</td></tr>`;
}
sh += `<tr class="tot"><td>GRAND TOTAL</td><td></td><td></td><td>${T.controllers}</td>`
    + `<td>${T.AI}</td><td>${T.AO}</td><td>${T.BI}</td><td>${T.BO}</td><td>${T.points}</td></tr></tbody>`;
document.getElementById('sum').innerHTML = sh;

const tree = document.getElementById('tree');
let th = '';
for(const p of DATA.panels){
  th += `<details class="panel" data-panel="${esc(p.id)}"><summary>${esc(p.id)}`
      + `<span class="r">${ioText(p.io)} \u00b7 ${p.nPoints} pts \u00b7 ${p.nctrl} ctrl</span></summary>`
      + `<div class="serves">${esc(p.description)} \u2014 ${esc(p.location)}</div>`;
  for(const c of p.controllers){
    th += `<details class="ctrl"><summary>${esc(c.id)}<span class="r">${ioText(c.io)} \u00b7 ${c.nPoints} pts</span></summary>`
        + `<div class="serves">Serves: ${esc(c.devices)}</div>`
        + '<table class="pts"><thead><tr><th>Point tag</th><th>Description</th><th>Device</th><th>I/O</th><th>Units</th><th>System</th></tr></thead><tbody>';
    for(const pt of c.points){
      const blob = (pt.tag+' '+pt.dev+' '+pt.desc+' '+pt.sys).toLowerCase();
      th += `<tr data-s="${esc(blob)}"><td>${esc(pt.tag)}</td><td>${esc(pt.desc)}</td>`
          + `<td>${esc(pt.dev)}</td><td class="io ${IOC[pt.io]||''}">${pt.io}</td>`
          + `<td>${esc(pt.units)}</td><td>${esc(pt.sys)}</td></tr>`;
    }
    th += '</tbody></table></details>';
  }
  th += '</details>';
}
tree.innerHTML = th;

const q = document.getElementById('q');
q.addEventListener('input', ()=>{
  const s = q.value.trim().toLowerCase();
  for(const panel of tree.querySelectorAll('details.panel')){
    let pAny = false;
    for(const ctrl of panel.querySelectorAll('details.ctrl')){
      let cAny = false;
      for(const tr of ctrl.querySelectorAll('tbody tr')){
        const hit = !s || tr.dataset.s.includes(s);
        tr.classList.toggle('hidden', !hit);
        cAny = cAny || hit;
      }
      ctrl.classList.toggle('hidden', !cAny);
      ctrl.open = !!s && cAny;
      pAny = pAny || cAny;
    }
    panel.classList.toggle('hidden', !pAny);
    panel.open = !!s && pAny;
  }
});
document.getElementById('expand').onclick = ()=>tree.querySelectorAll('details').forEach(d=>d.open=true);
document.getElementById('collapse').onclick = ()=>tree.querySelectorAll('details').forEach(d=>d.open=false);
</script>
</body></html>
"""


def write_html(ds, path):
    out = (HTML_TEMPLATE
           .replace("__TITLE__", html.escape(ds["title"]))
           .replace("__SUBTITLE__", html.escape(ds["subtitle"]))
           .replace("__GENERATED__", ds["generated"])
           .replace("__DATA__", json.dumps(ds, ensure_ascii=False)))
    for esc, ch in (("\\u2014", "\u2014"), ("\\u00b7", "\u00b7"),
                    ("\\u2192", "\u2192"), ("\\u2026", "\u2026")):
        out = out.replace(esc, ch)
    with open(path, "w", encoding="utf-8") as f:
        f.write(out)


# ---------------------------------------------------------------------------
# Predicates
# ---------------------------------------------------------------------------
def is_lighting(r):
    return (r["System"] == "Lighting" or r["Panel"] == "LCP-LTG"
            or r["Point Tag"].startswith("LTG-"))


_ELEC_KW = ("power", "electrical", "trip", "overload", "phase", "demand",
            "kwh", "kilowatt", "current", "voltage", "受電", "電力")


def is_electrical(r):
    if r["System"] == "Metering":
        return True
    if r["Point Tag"].endswith(".KW"):
        return True
    if (r["Units"] or "").lower() in ("kw", "kwh"):
        return True
    blob = (r["Point Description"] + " " + (r["Notes / Basis"] or "")).lower()
    return any(k in blob for k in _ELEC_KW)


DATASETS = [
    ("full", "Red5-DHCP \u2014 Full BMS I/O",
     "Every DDC panel \u2192 controller \u2192 point in the savic-net FX2 model.",
     lambda r: True),
    ("lighting", "Red5-DHCP \u2014 Lighting (\u7167\u660e\u4e00\u89a7)",
     "Common-area & facade lighting groups supervised on savic-net (LCP-LTG).",
     is_lighting),
    ("electrical", "Red5-DHCP \u2014 Electrical / Power monitoring",
     "BMS electrical cross-cut: energy meters, device kW power, and electrical-protection points.",
     is_electrical),
]


def export(ds, stem):
    os.makedirs(EXPORT_DIR, exist_ok=True)
    write_csv(ds, os.path.join(EXPORT_DIR, stem + ".csv"))
    write_xlsx(ds, os.path.join(EXPORT_DIR, stem + ".xlsx"))
    write_html(ds, os.path.join(EXPORT_DIR, stem + ".html"))
    print(f"  {stem:<26} panels={ds['totals']['panels']:>2} "
          f"ctrl={ds['totals']['controllers']:>3} points={ds['totals']['points']:>4}")


if __name__ == "__main__":
    print("Building exports ->", EXPORT_DIR)
    for stem, title, sub, pred in DATASETS:
        export(build_dataset(title, sub, pred), "red5-dhcp_" + stem)
    print("done.")
