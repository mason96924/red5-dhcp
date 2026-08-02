"""generate_io_list.py -- Red5-DHCP BMS I/O list + panel / controller schedule.

Builds a comprehensive multi-sheet Excel workbook for the ANA InterContinental
Tokyo District-Heating-&-Cooling (DHC) connected BMS.  The device/point model
was REBUILT from a direct reading of the Azbil savic-net FX savic-net FX2
graphics (完成図 pages 9-15, graph IDs 1000-1100) plus the air-side summary
graphs (空調関連 / 客室等) and the M-* CAD equipment schedules.

Heat source is DISTRICT HEATING & COOLING (DHC) -- the local plant is backup:
  DHC受入 (graph 1100)   DHC chilled water CS/CR (GJ/h + m3/h + P/T metering) and
                         DHC steam SS 0.8MPa -> PRV 8k->2k -> 0.2MPa, condensate
                         return HR (metered) back to the district plant.
  ホットウェル (1004)     HWT-1 condensate/hot-well tank (level Hi/Lo, temp);
                         HP-5 condensate-return pumps -> DHC; HP-3 pumps -> B1F
                         kitchen AHUs (AC-5/6/7).
  冷却塔 (1005)          CT-1 (2 INV cells) + CT-2 (2 cells + filtration) rejecting
                         heat from packaged units (PCU/PAC/PMAC) & kitchen
                         refrigeration; CDP-1 (x3) + CDP-2 (x2) condenser pumps;
                         EX-1 winter free-cooling HX; 2 emergency cooling HX to DHC.
  低層 (1001)           CP-1 (x3) primary CHW + HP-1 (x3) hot-water pumps; CP-4/5
                         kitchen pumps; EX-2 (x2) laundry HX; EXT-1; ST-1 OA station;
                         5~20F FCU groups (N/NE/SE/S/SW, 4-pipe).
  高層 (1002)           CP-2 (x3) primary CHW + HP-2 (x3) hot-water pumps; CP-3 (INV)
                         + CP-6 pumps; EX-3 (x2) HX; EXT-2; ST-2 OA station;
                         21~35F FCU groups (N/NE/SE/S/SW, 4-pipe).
  36,37F (1003/p12)     R-1 single water-cooled SCREW chiller (Ebara RHS DW202M2,
                         370 kW / COP 4.5, 2x45 kW screw comp, R-407C), DHC-backup;
                         CDP-3 (x2) + CT-3 condenser; CP-7 (x2) CHW;
                         HEX-1 + CP-8 DHC HX; EX4 + HP-4 (x2) + EXT-3 secondary;
                         2 source-changeover valves (DHC <-> R-1).
  Air side              AC-1..27 AHUs, EVU-1..15 OAUs, EF/SF/RF vent fans.
  Also on savic-net     Packaged units PCU/PAC/PMAC, kitchen refrigeration, and
                         common-area/facade lighting (照明一覧) -- monitored here.

Run:  .venv/bin/python generate_io_list.py
Out:  Red5-DHCP_BMS_IO_List.xlsx
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "exports", "Red5-DHCP_BMS_IO_List.xlsx")

# --------------------------------------------------------------------------
# Panels (distributed DDC local control panels + central)
# --------------------------------------------------------------------------
PANELS = [
    ("BMS-CENTRAL", "Azbil savic-net FX2 central monitoring (中央監視)", "B2F Central Monitoring Rm",
     "Whole-building supervisory on savic-net FX2 (S/W spec 2015/04/09): point mgmt, "
     "start-stop/setpoint operation, status & alarm/event processing, time-program + "
     "calendar + event-program control, seasonal changeover (batch), remote-setpoint "
     "schedules, runtime/start-count & deviation/hi-lo limit monitoring, trend + periodic "
     "data collection, daily/monthly reports, energy CSV export, numeric/logic operations, "
     "power-failure/restoration handling, user/access mgmt, maintenance & spare-parts mgmt"),
    ("LCP-DHC", "DHC intake & hot-well panel", "B2F Machine Room",
     "DHC chilled-water intake (CS/CR, GJ meter), DHC steam intake (SS 8k->2k, HR condensate), "
     "hot-well HWT-1 + HP-5 condensate-return / HP-3 kitchen pumps, main energy meters"),
    ("LCP-CT", "Cooling-tower & condenser-water panel", "Rooftop / PH",
     "CT-1/CT-2 cells (INV), CDP-1/CDP-2 condenser pumps, EX-1 winter free-cooling HX, "
     "emergency cooling HX to DHC; serves packaged units & kitchen refrigeration heat rejection"),
    ("LCP-HSL", "Low-rise distribution panel (低層系統)", "B2F Machine Room",
     "CP-1 primary CHW x3, HP-1 hot-water x3, CP-4/CP-5 kitchen pumps, EX-2 laundry HX, "
     "EXT-1, ST-1 OA station; low-rise AC/EVU + 5~20F FCU headers"),
    ("LCP-HSH", "High-rise distribution panel (高層系統)", "PH Machine Room",
     "CP-2 primary CHW x3, HP-2 hot-water x3, CP-3(INV)/CP-6 pumps, EX-3 HX, EXT-2, "
     "ST-2 OA station; high-rise AC/EVU + 21~35F FCU headers"),
    ("LCP-3637", "36/37F local heat-source panel (36,37F系統)", "37F PH Machine Room",
     "R-1 chiller + CDP-3 + CT-3 condenser, CP-7 CHW pumps, HEX-1 + CP-8 DHC HX, "
     "EX4 + HP-4 secondary + EXT-3, DHC<->R-1 changeover valves; 36/37F AC/EVU/FCU"),
    ("LCP-1F", "1F–2F air-side panel", "1F EPS/AHU Room",
     "AC-7, EVU-1..3 (1F ceremony/banquet/lobby)"),
    ("LCP-2F", "2F air-side panel", "2F AHU Room",
     "AC-8..13, EVU-4..6 (atrium, Prominence, offices/tenants)"),
    ("LCP-3F", "3F air-side panel", "3F AHU Room",
     "AC-14..16, EVU-7 (Unkai/Karin/DaVinci, chapel)"),
    ("LCP-4F", "4F–5F air-side panel", "4F AHU Room",
     "AC-17..21, EVU-8..10 (kitchens, low-rise guest-room OAUs)"),
    ("LCP-PKG", "Packaged-unit & refrigeration panel", "B3F–B1F EPS",
     "PCU/PAC/PMAC packaged air-conditioners (status/alarm), kitchen refrigeration/cold-store alarms"),
    ("LCP-LTG", "Common-area & facade lighting panel", "Distributed EPS",
     "Facade (aviation light, neon, balcony) + per-floor common-area lighting groups (照明一覧)"),
    ("LCP-DHW", "Domestic hot-water (給湯) plant panel", "B2F / PH plant",
     "Steam-heated DHW storage ST-1…4 banks + zone supply/recirc pumps by vertical zone "
     "(B3F-5F / 5F-15F / 16F-25F / 26F-37F) with storage steam-flow metering — per 給湯系統図 (graph 4100)"),
    ("LCP-SAN", "Plumbing / sanitary panel (上水・中水・排水)", "B3F–B2F plant",
     "Potable (上水) & reclaimed (中水) booster sets (MP/P), receiving/elevated/fire tanks, "
     "drainage sump & sewage pumps (DP) + aeration blowers (BL) — per 衛生系統図 (graphs 4200/4300)"),
    ("LCP-FILT", "Recreational-water filtration panel (ろ過装置)", "Distributed (2F–5F)",
     "Bath (男/女) / sauna / pool / waterfall (大滝・雲海) / fountain / rainwater filtration & "
     "circulation — per ろ過装置 (graph 4400)"),
    ("LCP-SMK", "Smoke-control / 排煙 panel", "Distributed / 防災",
     "Smoke-exhaust fans SMF-1…22 with fire-mode batch (火災一括); BMS status + smoke-mode "
     "start/stop — per 排煙系統図 (graph 5000)"),
]

# device -> air-side panel by floor grouping
def air_panel(dev_id: str) -> str:
    if dev_id.startswith("AC-"):
        ac = int(dev_id.split("-")[1])
        if ac in (22, 23, 24, 25, 27): return "LCP-3637"
        if ac in (17, 18, 19, 20, 21): return "LCP-4F"
        if ac in (14, 15, 16): return "LCP-3F"
        if ac in (8, 9, 10, 11, 12, 13): return "LCP-2F"
        if ac == 7: return "LCP-1F"
        return "LCP-HSL"  # AC-1..6, AC-26 basement/kitchen
    if dev_id.startswith("EVU-"):
        n = int(dev_id.split("-")[1])
        if n in (11, 12, 13, 14, 15): return "LCP-HSH"
        if n in (8, 9, 10): return "LCP-4F"
        if n == 7: return "LCP-3F"
        if n in (4, 5, 6): return "LCP-2F"
        return "LCP-1F"  # 1..3
    return "LCP-HSH"

# --------------------------------------------------------------------------
# Signal-type shorthands
# --------------------------------------------------------------------------
AI, AO, BI, BO = "AI", "AO", "BI", "BO"
SIG = {
    "temp":   ("4-20mA", "°C"),
    "rh":     ("4-20mA", "%RH"),
    "press":  ("4-20mA", "MPa"),
    "flow":   ("4-20mA", "L/min"),
    "kw":     ("4-20mA", "kW"),
    "pct":    ("0-10V",  "%"),
    "pctfb":  ("4-20mA", "%"),
    "level":  ("4-20mA", "%"),
    "status": ("DI dry contact", "-"),
    "cmd":    ("DO relay", "-"),
    "pulse":  ("Pulse/Modbus", "unit"),
    "co2":    ("4-20mA", "ppm"),
}

rows = []   # each: dict of columns

def add(tag, system, dev, dtype, area, desc, io, sigkey, units=None,
        sp="", alarm="", trend="", notes=""):
    sig, unit = SIG[sigkey]
    rows.append({
        "Point Tag": tag, "System": system, "Device ID": dev,
        "Device Type": dtype, "Area / Served": area,
        "Panel": PANEL_HINT.get(dev) or (air_panel(dev) if dev[:2] in ("AC", "EV") else "LCP-HSH"),
        "Controller": "",  # filled later
        "Point Description": desc, "I/O Type": io,
        "Signal / Range": sig, "Units": units or unit,
        "SP/Sched": sp, "Alarm": alarm, "Trend": trend, "Notes / Basis": notes,
    })

# heat-source / plant device -> panel hint
PANEL_HINT = {}

# --------------------------------------------------------------------------
# Device inventories
# --------------------------------------------------------------------------
AC_AREAS = {
    1: "B3-4F service elevator hall", 2: "Mid banquet / B1F back corridor / F&B office",
    3: "B1F guest lift lobby & anteroom", 4: "B1F banquet lobby",
    5: "Guardmanger & Bakery kitchen", 6: "C kitchen", 7: "Main kitchen",
    8: "Cascade Cafe hall", 9: "Cascade kitchen", 10: "Atrium lobby/lounge",
    11: "Prominence I", 12: "Prominence II", 13: "Prominence III",
    14: "Unkai hall", 15: "Karin/Kenzan kitchen", 16: "DaVinci kitchen",
    17: "Unkai kitchen", 18: "Steakhouse kitchen", 19: "Steakhouse hall",
    20: "Poolside snack/pantry (4F)", 21: "5F staff canteen kitchen",
    22: "PG/Mixx kitchen (36F)", 23: "Akasaka kitchen (37F)",
    24: "Akasaka hall / Libra / Aries", 25: "Sirius hall & kitchen",
    26: "Banquet-service office (B2F)", 27: "Club Lounge (35F) - 4-pipe",
}
AC_KITCHEN = {5, 6, 7, 9, 15, 16, 17, 18, 21, 22, 23}   # supply-air low-temp kitchens
AC_NO_INV = {1, 2, 3, 8, 9}                              # "制御無し" / INV故障 per schedule
AC_HAS_HEAT = set(range(1, 28)) - {2, 3, 4}              # most have steam heat; a few cool-only
AC_HUMID = {10, 14, 19, 24, 27}                          # halls/lounges with humidification
AC_CO2 = {11, 12, 13}                                    # banquet halls w/ CO2 demand-controlled vent (自動制御 A-11 sheet)

EVU_AREAS = {
    1: "1F ceremony/photo/dress/beauty/corridor", 2: "1F small banquet/corridor/store",
    3: "Banquet entrance lobby & bell room", 4: "2F office & CPU room",
    5: "2F tenants/shops", 6: "Prominence (B1F precool)",
    7: "3F Karin/Kenzan/DaVinci/chapel", 8: "Low-rise guest rooms N (6-20F)",
    9: "Low-rise guest rooms SE (6-20F)", 10: "Low-rise guest rooms SW (6-20F)",
    11: "High-rise guest rooms N (21-35F)", 12: "High-rise guest rooms SE (21-35F)",
    13: "High-rise guest rooms SW (21-35F)", 14: "36F office/Mixx",
    15: "36F PG hall/rooms/pantry, Mixx bar/hall",
}
EVU_NO_INV = {1, 2, 3, 4, 5, 6, 7, 14, 15}   # "制御無しの為変更不可" per schedule

# fans: (tag, parent AC, inv?)
FANS = [
    ("SF-6", 1, False), ("EF-2-1", 8, False), ("EF-28", 5, False),
    ("EF-32", 6, False), ("EF-33", 6, False), ("EF-41", 7, False), ("EF-42", 7, False),
    ("EF-46", 9, False), ("EF-47", 9, False), ("EF-54", 17, False), ("EF-55", 17, False),
    ("EF-56", 14, False), ("EF-57", 18, False), ("EF-58", 18, False),
    ("EF-59", 19, False), ("EF-59-2", 19, False), ("EF-60", 15, False),
    ("EF-61", 16, False), ("EF-62", 16, False), ("EF-67", 20, False),
    ("EF-69", 21, False), ("EF-70", 21, True), ("EF-71", 20, False),
    ("EF-76", 15, False), ("EF-77", 15, False), ("EF-82", 22, False),
    ("EF-83", 22, True), ("EF-84", 23, False), ("EF-85", 24, False),
    ("EF-86", 24, False), ("EF-89", 25, False), ("EF-97", 26, False),
    ("EF-98", 27, False), ("RF-7", 3, False), ("RF-8", 4, False), ("RF-9", 26, False),
]

# FCU BMS zone-GROUPS (guest rooms 4-pipe on DHC water), confirmed from the
# Azbil graphic: 5~20F FCU (N/S/SE/SW) low-rise + 20~35F FCU high-rise, each
# batch-monitored (冷水BV/温水BV一括故障). Per-room FCUs are locally controlled;
# the BMS supervises them by riser/orientation group.
# Confirmed from 完成図 p10 (低層) & p11 (高層): FIVE orientation groups per tier
# -- 5~20F FCU (N/NE/SE/S/SW) and 21~35F FCU (N/NE/SE/S/SW), each 4-pipe
# (冷水/温水) with 冷水BV/温水BV一括故障 batch-fault monitoring.
FCU_GROUPS = [
    ("FCU-L-N",  "Low-rise guest rooms 5-20F (N)",  "LCP-HSL"),
    ("FCU-L-NE", "Low-rise guest rooms 5-20F (NE)", "LCP-HSL"),
    ("FCU-L-SE", "Low-rise guest rooms 5-20F (SE)", "LCP-HSL"),
    ("FCU-L-S",  "Low-rise guest rooms 5-20F (S)",  "LCP-HSL"),
    ("FCU-L-SW", "Low-rise guest rooms 5-20F (SW)", "LCP-HSL"),
    ("FCU-H-N",  "High-rise guest rooms 21-35F (N)",  "LCP-HSH"),
    ("FCU-H-NE", "High-rise guest rooms 21-35F (NE)", "LCP-HSH"),
    ("FCU-H-SE", "High-rise guest rooms 21-35F (SE)", "LCP-HSH"),
    ("FCU-H-S",  "High-rise guest rooms 21-35F (S)",  "LCP-HSH"),
    ("FCU-H-SW", "High-rise guest rooms 21-35F (SW)", "LCP-HSH"),
    ("FCU-3637", "36/37F FCU (4-pipe, EX4 secondary loop)", "LCP-3637"),
]

# --------------------------------------------------------------------------
# Packaged AC units (PCU / PAC / PMAC) -- water-cooled, condenser off CT-1/CT-2.
# Confirmed from 完成図 p16 (PMAC/PAC一覧) + the cooling-tower load list.
# BMS monitors status/alarm and (for most) start/stop only.
# --------------------------------------------------------------------------
PKG_UNITS = [
    ("PAC-1-1", "特高電気室 (1)", True), ("PAC-1-2", "特高電気室 (2)", True),
    ("PAC-2",   "第1電気室 / 前室", True), ("PAC-3", "第1電気室 (2)", True),
    ("PAC-4",   "第2電気室", True),        ("PAC-5", "大宴会場電気室", True),
    ("PCU-1-1", "ギャラクシー II", False), ("PCU-1-2", "ギャラクシー I", False),
    ("PCU-5",   "ルミナス", False),        ("PCU-6", "ギャラクシー III", False),
    ("PCU-7",   "グローリー", False),      ("PCU-8", "オーロラ", False),
    ("PCU-10",  "プリズム", False),        ("PCU-11", "客用ELVロビー", False),
    ("PCU-12",  "宴会場ロビー (B2F)", False), ("PCU-13", "大宴会場前室", False),
    ("PMAC-1",  "メンテナンスセンター/事務室", False),
]

# Kitchen refrigeration / cold-store batch alarms (condenser-water rejected),
# per the 冷却塔制御 sheet (冷蔵庫異常一括).
REFR_ALARMS = [
    "B1F 主厨房一括", "B1F 厨房冷蔵庫異常一括", "MR-1 ル・パティオ厨房",
    "2F ゴミ処理室", "2F カスケイド厨房", "調理冷蔵庫", "雲海厨房",
]

# Common-area / facade lighting groups on savic-net (照明一覧 graphs 3000/3001).
# The full list is ~40 floors x 2-3 circuits; modelled here as representative
# supervised groups (on/off command + status).
LTG_GROUPS = [
    ("LTG-AVIATION", "航空障害灯 (aviation obstruction light)"),
    ("LTG-NEON",     "ネオンサイン (facade neon sign)"),
    ("LTG-BALCONY-L","バルコニー照明 4~20F"),
    ("LTG-BALCONY-H","バルコニー照明 21~37F"),
    ("LTG-CORRIDOR", "客用廊下・ELVホール照明 (per-floor batch)"),
    ("LTG-LOBBY",    "客用ELVロビー・ロビー照明"),
    ("LTG-SOFFIT",   "軒天井照明 (1F/2F/3F)"),
    ("LTG-TENANT",   "テナント間接・ダウンライト"),
]

# --------------------------------------------------------------------------
# Point templates
# --------------------------------------------------------------------------
def chiller_points(dev, area, panel="LCP-3637"):
    """R-1: single local water-cooled SCREW chiller -- Ebara RHS DW202M2
    (エバラスクリュー冷凍機, 荏原冷熱システム, 2014/12): 370 kW cooling / 82.2 kW
    input (COP 4.5), twin 45 kW screw compressors, dual R-407C circuits.
    Dedicated to the 36/37F loop, run in changeover/backup with DHC."""
    T = "Water-cooled screw chiller (local backup)"
    PANEL_HINT[dev] = panel
    add(f"{dev}.RUN", "Heat source", dev, T, area, "Run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", "Heat source", dev, T, area, "Common fault / trip", BI, "status", alarm="Y")
    add(f"{dev}.LR", "Heat source", dev, T, area, "Local/Remote status", BI, "status")
    add(f"{dev}.SS", "Heat source", dev, T, area, "Start/Stop command", BO, "cmd", sp="Y",
        notes="36/37F local chiller; BACKUP to DHC (チラーバックアップ); source changeover valves")
    add(f"{dev}.DMD", "Heat source", dev, T, area, "Capacity/demand limit", AO, "pct", sp="Y", notes="Demand-limit / soft-load")
    add(f"{dev}.CHWST", "Heat source", dev, T, area, "CHW supply (leaving) temp", AI, "temp", alarm="Y", trend="Y", notes="Design 7°C")
    add(f"{dev}.CHWRT", "Heat source", dev, T, area, "CHW return (entering) temp", AI, "temp", trend="Y", notes="Design 12°C")
    add(f"{dev}.CWRT", "Heat source", dev, T, area, "Condenser water leaving temp", AI, "temp", trend="Y")
    add(f"{dev}.KW", "Heat source", dev, T, area, "Electrical input power", AI, "kw", trend="Y", notes="Ebara screw RHS DW202M2: 370 kW cool / 82.2 kW in (COP 4.5); 2×45 kW screw comp, R-407C")
    add(f"{dev}.FLW", "Heat source", dev, T, area, "Evaporator flow proof", BI, "status", alarm="Y")

def cdpump_points(dev, area):
    """CDP-n: condenser-water pump (heat rejection via cooling towers)."""
    PANEL_HINT[dev] = "LCP-CT"
    add(f"{dev}.RUN", "Condenser water", dev, "Condenser-water pump", area, "Run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", "Condenser water", dev, "Condenser-water pump", area, "Fault / breaker trip", BI, "status", alarm="Y")
    add(f"{dev}.SS", "Condenser water", dev, "Condenser-water pump", area, "Start/Stop command", BO, "cmd", notes="Group start (群起動); interlock w/ CT + load")
    add(f"{dev}.FLW", "Condenser water", dev, "Condenser-water pump", area, "Flow/DP proof", BI, "status", alarm="Y")

def vfd_pump_points(dev, dtype, area, sysname, notes_dp=""):
    add(f"{dev}.RUN", sysname, dev, dtype, area, "Run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", sysname, dev, dtype, area, "VFD/pump fault", BI, "status", alarm="Y")
    add(f"{dev}.SS", sysname, dev, dtype, area, "Start/Stop command", BO, "cmd")
    add(f"{dev}.SPD", sysname, dev, dtype, area, "VFD speed command", AO, "pct", sp="Y", notes=notes_dp)
    add(f"{dev}.SPDFB", sysname, dev, dtype, area, "VFD speed feedback", AI, "pctfb", trend="Y")
    add(f"{dev}.KW", sysname, dev, dtype, area, "Power", AI, "kw", trend="Y")
    add(f"{dev}.FLW", sysname, dev, dtype, area, "Flow/DP proof", BI, "status", alarm="Y")

def tower_cell_points(dev, area, inv=True):
    """One cooling-tower fan cell (e.g. CT-1-1). INV cells have VFD speed."""
    PANEL_HINT[dev] = "LCP-CT"
    add(f"{dev}.RUN", "Condenser water", dev, "Cooling tower cell", area, "Fan run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", "Condenser water", dev, "Cooling tower cell", area, "Fan INV fault", BI, "status", alarm="Y")
    add(f"{dev}.SS", "Condenser water", dev, "Cooling tower cell", area, "Fan start/stop", BO, "cmd")
    if inv:
        add(f"{dev}.SPD", "Condenser water", dev, "Cooling tower cell", area, "Fan VFD speed command", AO, "pct", sp="Y", notes="Wet-bulb + approach reset")
        add(f"{dev}.SPDFB", "Condenser water", dev, "Cooling tower cell", area, "Fan speed feedback", AI, "pctfb", trend="Y")

def tower_common_points(dev, area, filtration=False):
    """Tower-body common points (basin level, make-up, treatment, alarms)."""
    PANEL_HINT[dev] = "LCP-CT"
    add(f"{dev}.LVL", "Condenser water", dev, "Cooling tower", area, "Basin water level", AI, "level", alarm="Y")
    add(f"{dev}.MU", "Condenser water", dev, "Cooling tower", area, "Make-up water valve", BO, "cmd")
    add(f"{dev}.BLD", "Condenser water", dev, "Cooling tower", area, "Bleed/blowdown valve", BO, "cmd")
    add(f"{dev}.DRY", "Condenser water", dev, "Cooling tower", area, "Low-water (渇水) alarm", BI, "status", alarm="Y")
    add(f"{dev}.CHEM", "Condenser water", dev, "Cooling tower", area, "Chemical dosing (薬注) batch fault", BI, "status", alarm="Y")
    add(f"{dev}.EH", "Condenser water", dev, "Cooling tower", area, "Basin heater (E/H) enable", BO, "cmd", notes="Freeze protection")
    if filtration:
        add(f"{dev}.FILT-RUN", "Condenser water", dev, "Cooling tower", area, "Filtration (ろ過) pump run", BI, "status", trend="Y")
        add(f"{dev}.FILT-SS", "Condenser water", dev, "Cooling tower", area, "Filtration pump start/stop", BO, "cmd")

def freecool_points(dev, area):
    """EX-1: winter free-cooling heat exchanger (冬期用熱交換器)."""
    PANEL_HINT[dev] = "LCP-CT"
    add(f"{dev}.EN", "Condenser water", dev, "Free-cooling HX (winter)", area, "Free-cooling enable / changeover", BO, "cmd", sp="Y", notes="Low wet-bulb winter free-cooling")
    add(f"{dev}.ST", "Condenser water", dev, "Free-cooling HX (winter)", area, "CHW-side leaving temp", AI, "temp", trend="Y")
    add(f"{dev}.CV", "Condenser water", dev, "Free-cooling HX (winter)", area, "CW-side control valve", AO, "pct", sp="Y")

def emerg_hx_points(dev, area):
    """Low-floor emergency cooling-water HX tied to the DHC chilled headers."""
    PANEL_HINT[dev] = "LCP-CT"
    add(f"{dev}.EN", "Condenser water", dev, "Emergency cooling HX (to DHC)", area, "Emergency cooling enable", BO, "cmd", sp="Y", notes="低層階非常用; from DHC冷水受入ヘッダ")
    add(f"{dev}.ST", "Condenser water", dev, "Emergency cooling HX (to DHC)", area, "Cooling-water leaving temp", AI, "temp", trend="Y")
    add(f"{dev}.ISOL", "Condenser water", dev, "Emergency cooling HX (to DHC)", area, "DHC-side isolation valve", BO, "cmd")

def dhc_chw_points(dev, area):
    """DHC chilled-water intake (冷水受入 / DHCプラント, graph 1100): CS supply +
    CR return headers with thermal-energy metering.  This is the PRIMARY cooling
    source for the whole building."""
    PANEL_HINT[dev] = "LCP-DHC"
    add(f"{dev}.CS-T", "DHC interface", dev, "DHC chilled-water intake", area, "Supply (CS) temp", AI, "temp", alarm="Y", trend="Y", notes="DHC冷水受入 ~7°C")
    add(f"{dev}.CR-T", "DHC interface", dev, "DHC chilled-water intake", area, "Return (CR) temp", AI, "temp", trend="Y", notes="DHC冷水返送")
    add(f"{dev}.CS-P", "DHC interface", dev, "DHC chilled-water intake", area, "Supply header pressure", AI, "press", alarm="Y", trend="Y")
    add(f"{dev}.CR-P", "DHC interface", dev, "DHC chilled-water intake", area, "Return header pressure", AI, "press", trend="Y")
    add(f"{dev}.FLW", "DHC interface", dev, "DHC chilled-water intake", area, "Intake flow", AI, "flow", units="m³/h", trend="Y")
    add(f"{dev}.GJ", "DHC interface", dev, "DHC chilled-water intake", area, "Thermal-energy meter", AI, "pulse", units="GJ", trend="Y", notes="GJ/h + integrated; thermal-demand control")
    add(f"{dev}.CV", "DHC interface", dev, "DHC chilled-water intake", area, "Intake control valve", AO, "pct", sp="Y", notes="Modulate to building load")
    add(f"{dev}.CVFB", "DHC interface", dev, "DHC chilled-water intake", area, "Intake valve position fb", AI, "pctfb")
    add(f"{dev}.ISOL", "DHC interface", dev, "DHC chilled-water intake", area, "Intake isolation valve open/close", BO, "cmd", sp="Y")
    add(f"{dev}.ISOL-ST", "DHC interface", dev, "DHC chilled-water intake", area, "Isolation valve status", BI, "status", alarm="Y")

def hex_points(dev, area, panel="LCP-3637", note="36/37F DHC tie-in (チラーバックアップ)"):
    """HEX-1: DHC chilled-water plate HX feeding the 36/37F loop (with CP-8 +
    DHC thermal meter), sharing duty with R-1 in changeover."""
    PANEL_HINT[dev] = panel
    add(f"{dev}.P-IN", "Heat source", dev, "DHC chilled-water plate HX", area, "Primary (DHC) inlet temp", AI, "temp", trend="Y", notes=note)
    add(f"{dev}.P-OUT", "Heat source", dev, "DHC chilled-water plate HX", area, "Primary (DHC) outlet temp", AI, "temp", trend="Y")
    add(f"{dev}.S-OUT", "Heat source", dev, "DHC chilled-water plate HX", area, "Secondary (loop) supply temp", AI, "temp", sp="Y", alarm="Y", trend="Y", notes="Loop CHW setpoint")
    add(f"{dev}.S-FLW", "Heat source", dev, "DHC chilled-water plate HX", area, "Secondary supply flow", AI, "flow", units="m³/h", trend="Y")
    add(f"{dev}.GJ", "Heat source", dev, "DHC chilled-water plate HX", area, "DHC thermal-energy meter", AI, "pulse", units="MJ", trend="Y", notes="m³/h + MJ on the graphic")
    add(f"{dev}.PV", "Heat source", dev, "DHC chilled-water plate HX", area, "Primary control valve", AO, "pct", sp="Y", notes="Modulate to hold S-OUT")
    add(f"{dev}.PVFB", "Heat source", dev, "DHC chilled-water plate HX", area, "Primary valve position fb", AI, "pctfb")

def dhc_steam_points(dev, area):
    """DHC steam intake (蒸気受入 / DHCプラント, graph 1100): SS 0.8MPa (8k) ->
    PRV -> 0.2MPa (2k) headers; condensate hot-return HR (metered) to DHC."""
    PANEL_HINT[dev] = "LCP-DHC"
    add(f"{dev}.SS-P", "DHC interface", dev, "DHC steam intake", area, "Supply (SS 8k) pressure", AI, "press", alarm="Y", trend="Y", notes="~0.8 MPa")
    add(f"{dev}.MASS", "DHC interface", dev, "DHC steam intake", area, "Steam mass-flow meter", AI, "pulse", units="t/h", trend="Y", notes="t/h + integrated kg")
    add(f"{dev}.PRV", "DHC interface", dev, "DHC steam intake", area, "PRV 8k->2k reducing valve", AO, "pct", sp="Y", notes="0.8->0.2 MPa")
    add(f"{dev}.SSH2-P", "DHC interface", dev, "DHC steam intake", area, "Reduced (SSH 2k) pressure", AI, "press", trend="Y", notes="~0.2 MPa")
    add(f"{dev}.ISOL", "DHC interface", dev, "DHC steam intake", area, "Main isolation valve open/close", BO, "cmd", sp="Y", notes="Season enable + OA lockout")
    add(f"{dev}.ISOL-ST", "DHC interface", dev, "DHC steam intake", area, "Isolation valve status", BI, "status", alarm="Y")
    add(f"{dev}.HR-M", "DHC interface", dev, "DHC steam intake", area, "Condensate (HR) return meter", AI, "pulse", units="L", trend="Y", notes="Hot-return to DHC via HWT-1")

def hotwell_points(dev, area):
    """HWT-1 hot-well / condensate tank (graph 1004): level, temp, condensate in,
    HP-5 return-to-DHC pumps and HP-3 kitchen-AHU pumps are separate devices."""
    PANEL_HINT[dev] = "LCP-DHC"
    add(f"{dev}.LVL", "Heat source", dev, "Hot-well / condensate tank", area, "Tank level", AI, "level", alarm="Y", trend="Y")
    add(f"{dev}.HI", "Heat source", dev, "Hot-well / condensate tank", area, "High-level alarm", BI, "status", alarm="Y")
    add(f"{dev}.LO", "Heat source", dev, "Hot-well / condensate tank", area, "Low-level alarm", BI, "status", alarm="Y")
    add(f"{dev}.T", "Heat source", dev, "Hot-well / condensate tank", area, "Condensate temp", AI, "temp", trend="Y")
    add(f"{dev}.MU", "Heat source", dev, "Hot-well / condensate tank", area, "Make-up water valve", BO, "cmd")

def ex_hx_points(dev, area, panel, served):
    """Distribution / process plate HX (EX-2 laundry, EX-3 kitchen, EX4 36/37F)."""
    PANEL_HINT[dev] = panel
    add(f"{dev}.P-IN", "Heat source", dev, "Plate heat exchanger", area, f"Primary inlet temp ({served})", AI, "temp", trend="Y")
    add(f"{dev}.S-OUT", "Heat source", dev, "Plate heat exchanger", area, "Secondary outlet (supply) temp", AI, "temp", sp="Y", alarm="Y", trend="Y")
    add(f"{dev}.PV", "Heat source", dev, "Plate heat exchanger", area, "Primary control valve", AO, "pct", sp="Y", notes="Modulate to hold S-OUT")
    add(f"{dev}.PVFB", "Heat source", dev, "Plate heat exchanger", area, "Primary valve position fb", AI, "pctfb")

def ext_points(dev, area, panel):
    """Expansion tank (EXT-1/2/3): level / make-up supervision."""
    PANEL_HINT[dev] = panel
    add(f"{dev}.LVL", "Heat source", dev, "Expansion tank", area, "Tank level", AI, "level", alarm="Y", trend="Y")
    add(f"{dev}.HI", "Heat source", dev, "Expansion tank", area, "High-level alarm", BI, "status", alarm="Y")
    add(f"{dev}.LO", "Heat source", dev, "Expansion tank", area, "Low-level alarm", BI, "status", alarm="Y")

def changeover_valve_points(dev, area, note):
    PANEL_HINT[dev] = "LCP-3637"
    add(f"{dev}.CMD", "Heat source", dev, "Source-changeover valve", area, "Open/close command", BO, "cmd", sp="Y", notes=note)
    add(f"{dev}.ST", "Heat source", dev, "Source-changeover valve", area, "Position status", BI, "status", alarm="Y")

def oa_station_points(dev, area, note):
    PANEL_HINT[dev] = "LCP-HSL" if dev.endswith("1") else "LCP-HSH"
    add(f"{dev}.T", "Common", dev, "Outdoor-air station", area, "Outdoor dry-bulb temp", AI, "temp", trend="Y", notes=note)
    add(f"{dev}.RH", "Common", dev, "Outdoor-air station", area, "Outdoor humidity", AI, "rh", trend="Y")
    add(f"{dev}.WB", "Common", dev, "Outdoor-air station", area, "Wet-bulb temp (計測/演算)", AI, "temp", trend="Y", notes="Drives CW setpoint + free-cooling")

def ahu_points(n):
    dev = f"AC-{n}"
    area = AC_AREAS[n]
    add(f"{dev}.RUN", "Air side", dev, "AHU (public/kitchen)", area, "Supply fan run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", "Air side", dev, "AHU (public/kitchen)", area, "Supply fan fault", BI, "status", alarm="Y")
    add(f"{dev}.SS", "Air side", dev, "AHU (public/kitchen)", area, "Supply fan start/stop", BO, "cmd", sp="Y", notes="Time/event schedule; banquet 'occ by event'")
    if n not in AC_NO_INV:
        add(f"{dev}.SPD", "Air side", dev, "AHU (public/kitchen)", area, "Supply fan VFD (INV) speed", AO, "pct", sp="Y", notes="INV output schedule 0.4-1.0")
        add(f"{dev}.SPDFB", "Air side", dev, "AHU (public/kitchen)", area, "Supply fan speed feedback", AI, "pctfb", trend="Y")
    else:
        add(f"{dev}.SPD", "Air side", dev, "AHU (public/kitchen)", area, "Supply fan speed (fixed)", AO, "pct", notes="INFERRED / no INV control per schedule")
    add(f"{dev}.SAT", "Air side", dev, "AHU (public/kitchen)", area, "Supply air temp", AI, "temp", sp="Y", alarm="Y", trend="Y",
        notes="Kitchen SAT 5-12°C" if n in AC_KITCHEN else "SAT/RAT setpoint schedule")
    add(f"{dev}.RAT", "Air side", dev, "AHU (public/kitchen)", area, "Return air temp", AI, "temp", trend="Y")
    add(f"{dev}.CCV", "Air side", dev, "AHU (public/kitchen)", area, "Cooling coil (CHW) valve", AO, "pct", sp="Y", notes="PID to SAT; zero-energy band")
    if n in AC_HAS_HEAT:
        add(f"{dev}.HCV", "Air side", dev, "AHU (public/kitchen)", area, "Heating coil (steam) valve", AO, "pct", sp="Y", notes="Season enable; deadband vs CCV")
    if n in AC_HUMID:
        add(f"{dev}.HUM", "Air side", dev, "AHU (public/kitchen)", area, "Steam humidifier valve", AO, "pct", sp="Y")
        add(f"{dev}.RH", "Air side", dev, "AHU (public/kitchen)", area, "Return/space humidity", AI, "rh", trend="Y")
    add(f"{dev}.FLT", "Air side", dev, "AHU (public/kitchen)", area, "Filter dirty (DP)", BI, "status", alarm="Y")
    add(f"{dev}.OAD", "Air side", dev, "AHU (public/kitchen)", area, "Outdoor-air damper", AO, "pct", sp="Y",
        notes="CO2 demand-controlled reset (DCV)" if n in AC_CO2 else "")
    if n in AC_CO2:
        add(f"{dev}.CO2", "Air side", dev, "AHU (public/kitchen)", area, "Space CO2 concentration", AI, "co2",
            sp="Y", alarm="Y", trend="Y", notes="Banquet DCV: resets OA damper / fan to hold ≤1000 ppm (自動制御 A-11)")
    if n in AC_HAS_HEAT:
        add(f"{dev}.FRZ", "Air side", dev, "AHU (public/kitchen)", area, "Freeze protection stat", BI, "status", alarm="Y")

def evu_points(n):
    dev = f"EVU-{n}"
    area = EVU_AREAS[n]
    add(f"{dev}.RUN", "Air side", dev, "Outdoor-air unit (OAU)", area, "Supply fan run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", "Air side", dev, "Outdoor-air unit (OAU)", area, "Supply fan fault", BI, "status", alarm="Y")
    add(f"{dev}.SS", "Air side", dev, "Outdoor-air unit (OAU)", area, "Supply fan start/stop", BO, "cmd", sp="Y", notes="Time schedule; night setback")
    if n not in EVU_NO_INV:
        add(f"{dev}.SPD", "Air side", dev, "Outdoor-air unit (OAU)", area, "Supply fan VFD (INV) speed", AO, "pct", sp="Y", notes="INV schedule; demand vent")
        add(f"{dev}.SPDFB", "Air side", dev, "Outdoor-air unit (OAU)", area, "Supply fan speed feedback", AI, "pctfb", trend="Y")
    else:
        add(f"{dev}.SPD", "Air side", dev, "Outdoor-air unit (OAU)", area, "Supply fan speed (fixed)", AO, "pct", notes="No INV control per schedule")
    add(f"{dev}.SAT", "Air side", dev, "Outdoor-air unit (OAU)", area, "Supply (off-coil) air temp", AI, "temp", sp="Y", alarm="Y", trend="Y", notes="Seasonal SAT reset")
    add(f"{dev}.OAT", "Air side", dev, "Outdoor-air unit (OAU)", area, "Outdoor-air temp (unit inlet)", AI, "temp", trend="Y")
    add(f"{dev}.OARH", "Air side", dev, "Outdoor-air unit (OAU)", area, "Outdoor-air humidity", AI, "rh", trend="Y")
    add(f"{dev}.CCV", "Air side", dev, "Outdoor-air unit (OAU)", area, "Cooling coil (CHW) valve", AO, "pct", sp="Y")
    add(f"{dev}.HCV", "Air side", dev, "Outdoor-air unit (OAU)", area, "Heating coil (steam) valve", AO, "pct", sp="Y", notes="Season enable; deadband")
    add(f"{dev}.HUM", "Air side", dev, "Outdoor-air unit (OAU)", area, "Steam humidifier valve", AO, "pct", sp="Y")
    add(f"{dev}.FLT", "Air side", dev, "Outdoor-air unit (OAU)", area, "Filter dirty (DP)", BI, "status", alarm="Y")
    add(f"{dev}.FRZ", "Air side", dev, "Outdoor-air unit (OAU)", area, "Freeze protection stat", BI, "status", alarm="Y")

def fcu_group_points(dev, area, panel):
    """BMS batch monitoring for a 4-pipe guest-room FCU group (per riser /
    orientation). Individual FCUs run on local room thermostats; the BMS
    supervises the group and its DHC 4-pipe header valves."""
    PANEL_HINT[dev] = panel
    ff = "36/37F: EX4 secondary loop (DHC/R-1 source)" if dev == "FCU-3637" else "DHC 4-pipe; season changeover"
    add(f"{dev}.EN", "Air side", dev, "FCU zone group (4-pipe)", area, "Group enable / schedule", BO, "cmd", sp="Y", notes=ff)
    add(f"{dev}.MODE", "Air side", dev, "FCU zone group (4-pipe)", area, "Cooling/Heating changeover status", BI, "status", notes="順序/季節切替")
    add(f"{dev}.CWV", "Air side", dev, "FCU zone group (4-pipe)", area, "Chilled-water group valve", AO, "pct", sp="Y")
    add(f"{dev}.HWV", "Air side", dev, "FCU zone group (4-pipe)", area, "Hot-water group valve", AO, "pct", sp="Y", notes="Deadband vs cooling")
    add(f"{dev}.CWBV-F", "Air side", dev, "FCU zone group (4-pipe)", area, "Chilled-water BV batch fault", BI, "status", alarm="Y", notes="冷水BV一括故障")
    add(f"{dev}.HWBV-F", "Air side", dev, "FCU zone group (4-pipe)", area, "Hot-water BV batch fault", BI, "status", alarm="Y", notes="温水BV一括故障")
    add(f"{dev}.RT", "Air side", dev, "FCU zone group (4-pipe)", area, "Zone reference/return temp", AI, "temp", sp="Y", trend="Y")

def fan_points(tag, parent_ac, inv):
    area = f"Serves AC-{parent_ac} ({AC_AREAS.get(parent_ac,'')})"
    PANEL_HINT[tag] = air_panel(f"AC-{parent_ac}")
    add(f"{tag}.SS", "Ventilation", tag, "Exhaust/supply fan", area, "Start/Stop command", BO, "cmd", sp="Y", notes="Interlock w/ parent AHU")
    add(f"{tag}.RUN", "Ventilation", tag, "Exhaust/supply fan", area, "Run status", BI, "status", trend="Y")
    add(f"{tag}.TRIP", "Ventilation", tag, "Exhaust/supply fan", area, "Fault", BI, "status", alarm="Y")
    if inv:
        add(f"{tag}.SPD", "Ventilation", tag, "Exhaust/supply fan", area, "VFD (INV) speed command", AO, "pct", sp="Y")
        add(f"{tag}.SPDFB", "Ventilation", tag, "Exhaust/supply fan", area, "VFD speed feedback", AI, "pctfb")

def pkg_points(dev, area, always_on):
    """Packaged air-conditioner (PCU/PAC/PMAC) -- water-cooled off CT-1/CT-2.
    BMS supervises status/alarm and start/stop (electrical-room units run 24/7)."""
    PANEL_HINT[dev] = "LCP-PKG"
    T = "Packaged AC unit (water-cooled)"
    add(f"{dev}.RUN", "Packaged units", dev, T, area, "Run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", "Packaged units", dev, T, area, "Fault / alarm", BI, "status", alarm="Y")
    add(f"{dev}.SS", "Packaged units", dev, T, area, "Start/Stop command", BO, "cmd", sp="Y",
        notes="24/7 (electrical room)" if always_on else "Time/occupancy schedule")

def refrigeration_points():
    dev = "REFR-1"
    PANEL_HINT[dev] = "LCP-PKG"
    for i, name in enumerate(REFR_ALARMS, 1):
        add(f"{dev}.ALM{i}", "Packaged units", dev, "Kitchen refrigeration", name,
            f"Cold-store/refrigeration alarm — {name}", BI, "status", alarm="Y",
            notes="冷蔵庫異常; condenser-water rejected")

def lighting_points():
    for tag, name in LTG_GROUPS:
        PANEL_HINT[tag] = "LCP-LTG"
        add(f"{tag}.CMD", "Lighting", tag, "Lighting group", name, "On/Off command", BO, "cmd", sp="Y",
            notes="Time/calendar + astronomical (sunset) schedule")
        add(f"{tag}.ST", "Lighting", tag, "Lighting group", name, "Group status", BI, "status")

def meter_points():
    meters = [
        ("MTR-ELEC-MAIN", "Main electricity (kWh)", "kWh", "LCP-DHC"),
        ("MTR-GAS-MAIN", "City gas (m³)", "m³", "LCP-DHC"),
        ("MTR-WATER-MAIN", "City water (m³)", "m³", "LCP-DHC"),
        ("MTR-DHC-CHW", "DHC chilled-water energy (受入)", "GJ", "LCP-DHC"),
        ("MTR-DHC-STEAM", "DHC steam energy (受入)", "t/h", "LCP-DHC"),
        ("MTR-DHC-COND", "DHC condensate return (HR)", "L", "LCP-DHC"),
        ("MTR-CDW-KWH", "Cooling-tower/condenser sub-meter", "kWh", "LCP-CT"),
        ("MTR-AHU-KWH", "AHU/OAU fans sub-meter", "kWh", "LCP-HSL"),
    ]
    for tag, desc, unit, panel in meters:
        PANEL_HINT[tag] = panel
        add(f"{tag}", "Metering", tag, "Energy meter", "Whole building", desc, AI, "pulse", units=unit, trend="Y",
            notes="Integrating meter; thermal-demand & peak limiting")

def common_points():
    PANEL_HINT["POOL-FP"] = "LCP-HSL"
    add("POOL-FP.SS", "Common", "POOL-FP", "Pool filtration pump", "Pool plant", "Start/Stop command", BO, "cmd", sp="Y", notes="Time schedule 8:00-19:00")
    add("POOL-FP.RUN", "Common", "POOL-FP", "Pool filtration pump", "Pool plant", "Run status", BI, "status")
    add("POOL-FP.TRIP", "Common", "POOL-FP", "Pool filtration pump", "Pool plant", "Fault", BI, "status", alarm="Y")

# --------------------------------------------------------------------------
# Domestic hot water (給湯 / graph 4100): steam-heated storage banks with a
# supply/recirc pump per vertical zone.  Drawing tags reuse ST-*/HP-* which
# collide with the OA-station / heat-source pumps, so these are given DHW-n
# device IDs and the drawing tags are cited in the notes.
# --------------------------------------------------------------------------
DHW_ZONES = [
    ("DHW-1", "B3F–5F DHW zone",  "ST-1-1…4 storage + HP-1 supply pump"),
    ("DHW-2", "5F–15F DHW zone",  "ST-2-1/2 storage + HP-2 supply pump"),
    ("DHW-3", "16F–25F DHW zone", "ST-3-1/2 storage + HP-3 supply pump"),
    ("DHW-4", "26F–37F DHW zone", "ST-4-1/2 storage + HP-4 supply pump"),
]

def dhw_points(dev, area, basis):
    PANEL_HINT[dev] = "LCP-DHW"
    T = "Domestic-hot-water plant (steam-heated storage)"
    S = "Domestic hot water"
    add(f"{dev}.STG-T", S, dev, T, area, "Storage-tank temperature", AI, "temp", sp="Y", alarm="Y", trend="Y",
        notes=f"{basis}; hold ≥60°C store / ≥55°C delivery (Legionella)")
    add(f"{dev}.STG-CV", S, dev, T, area, "Steam charging valve", AO, "pct", sp="Y", notes="Modulate to storage setpoint")
    add(f"{dev}.SUP-T", S, dev, T, area, "DHW supply (delivery) temperature", AI, "temp", alarm="Y", trend="Y")
    add(f"{dev}.STM-M", S, dev, T, area, "Storage steam-flow meter (貯湯槽蒸気流量)", AI, "pulse", units="m³", trend="Y")
    add(f"{dev}.PMP-RUN", S, dev, T, area, "Supply/recirc pump run status", BI, "status", trend="Y")
    add(f"{dev}.PMP-TRIP", S, dev, T, area, "Supply/recirc pump fault", BI, "status", alarm="Y")
    add(f"{dev}.PMP-SS", S, dev, T, area, "Supply/recirc pump start/stop", BO, "cmd", sp="Y", notes="Duty/standby; recirc schedule")

# --------------------------------------------------------------------------
# Plumbing / sanitary (衛生 / graphs 4200 上水・中水, 4300 排水).
# Potable (上水) + reclaimed (中水) booster sets, receiving/elevated/fire tanks,
# drainage sump & sewage pumps, wastewater-treatment aeration blowers.
# --------------------------------------------------------------------------
def san_pump_points(dev, area, role, note=""):
    PANEL_HINT[dev] = "LCP-SAN"
    S = "Plumbing/Sanitary"
    add(f"{dev}.RUN", S, dev, role, area, "Run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", S, dev, role, area, "Pump fault / breaker trip", BI, "status", alarm="Y")
    add(f"{dev}.SS", S, dev, role, area, "Start/Stop command", BO, "cmd", sp="Y", notes=note or "Lead/lag; tank-level float control")

def san_sump_points(dev, area, note=""):
    PANEL_HINT[dev] = "LCP-SAN"
    S = "Plumbing/Sanitary"
    T = "Drainage sump / sewage pump"
    add(f"{dev}.RUN", S, dev, T, area, "Pump run status", BI, "status", trend="Y")
    add(f"{dev}.TRIP", S, dev, T, area, "Pump fault", BI, "status", alarm="Y")
    add(f"{dev}.HLV", S, dev, T, area, "High-water-level alarm", BI, "status", alarm="Y", notes=note or "Float; flooding alarm")

def san_tank_points(dev, area, name, fire=False):
    PANEL_HINT[dev] = "LCP-SAN"
    S = "Plumbing/Sanitary"
    T = "Water tank"
    add(f"{dev}.LVL", S, dev, T, area, f"{name} water level", AI, "level", alarm="Y", trend="Y")
    if fire:
        add(f"{dev}.LO", S, dev, T, area, f"{name} low-level alarm", BI, "status", alarm="Y", notes="Fire-reserve; statutory")
    else:
        add(f"{dev}.HI", S, dev, T, area, f"{name} high-level alarm", BI, "status", alarm="Y")
        add(f"{dev}.LO", S, dev, T, area, f"{name} low-level / dry-run alarm", BI, "status", alarm="Y")

SAN_BOOST = [   # (dev, area, role, note)
    ("MP-1", "B3F 受水槽室", "Potable transfer/booster pump", "上水; 高置水槽揚水"),
    ("MP-2", "B3F 受水槽室", "Potable transfer/booster pump", "上水; 高置水槽揚水"),
    ("MP-3", "B3F 中水室",   "Reclaimed-water booster pump",  "中水加圧"),
    ("MP-4", "B3F 中水室",   "Reclaimed-water booster pump",  "中水加圧"),
    ("P-1", "B3F 加圧ポンプ室", "Domestic-water booster pump", "上水加圧装置"),
    ("P-2", "B3F 加圧ポンプ室", "Domestic-water booster pump", "上水加圧装置"),
    ("P-3", "B3F 加圧ポンプ室", "Domestic-water booster pump", "上水加圧装置"),
    ("P-4", "B3F 加圧ポンプ室", "Domestic-water booster pump", "低層上水 P-1 群"),
    ("P-5", "B3F 加圧ポンプ室", "Domestic-water booster pump", "高層上水"),
    ("P-6", "B3F 加圧ポンプ室", "Domestic-water booster pump", "高層上水"),
]
SAN_TANKS = [   # (dev, area, name, fire?)
    ("RCVT-1", "B3F 受水槽室", "Potable receiving tank (上水受水槽)", False),
    ("ELVT-1", "PH2F 高置水槽", "Potable elevated tank (高置水槽)", False),
    ("REUSET-1", "B3F 中水受水槽", "Reclaimed receiving tank (中水受水槽)", False),
    ("FIRET-1", "B3F 消火水槽", "Fire-reserve tank (消火水槽)", True),
]
SAN_SUMPS = [   # DP-1..14 sump / sewage pumps (graph 4300)
    ("DP-1", "B3F 汚水槽-1"), ("DP-2", "B3F 汚水槽-2"), ("DP-3", "B3F 排水槽-1"),
    ("DP-4", "B3F 雑排水槽"), ("DP-5", "B3F 雑排水放流槽"), ("DP-6", "B3F 雑排水移送槽"),
    ("DP-7", "B3F 雑排水移送槽-2"), ("DP-8", "B3F 排水槽-2"), ("DP-9", "B3F トレンチ排水槽-1"),
    ("DP-10", "B3F トレンチ排水槽-2"), ("DP-11", "B3F 湧水槽-1"), ("DP-12", "B3F 湧水槽-2"),
    ("DP-13", "B3F 湧水放流槽"), ("DP-14", "B3F 雑排水放流槽-2"),
]
SAN_BLOWERS = [ # aeration blowers for wastewater treatment (graph 4300)
    ("BL-1-1", "B3F ブロアー室1 (厨房排水処理)"), ("BL-1-2", "B3F ブロアー室1 (厨房排水処理)"),
    ("BL-2-1", "B3F ブロアー室2 (雑排水処理)"),   ("BL-2-2", "B3F ブロアー室2 (雑排水処理)"),
    ("BL-F-1", "B3F 厨房排水処理槽ばっ気"),        ("BL-F-2", "B3F 厨房排水処理槽ばっ気"),
]

def sanitary_points():
    for dev, area, role, note in SAN_BOOST:
        san_pump_points(dev, area, role, note)
    for dev, area, name, fire in SAN_TANKS:
        san_tank_points(dev, area, name, fire)
    for dev, area in SAN_SUMPS:
        san_sump_points(dev, area)
    for dev, area in SAN_BLOWERS:
        san_pump_points(dev, area, "Wastewater aeration blower", note="ばっ気; DO/timer control")

# --------------------------------------------------------------------------
# Recreational-water filtration (ろ過装置 / graph 4400): circulation +
# filtration plant for baths, sauna, pool, waterfalls and the rainwater tank.
# (Pool filtration is already modelled as POOL-FP under Common.)
# --------------------------------------------------------------------------
FILT_PLANTS = [
    ("FILT-BATHM", "5F 男子浴室 ろ過装置"), ("FILT-BATHW", "5F 女子浴室 ろ過装置"),
    ("FILT-SAUNA", "4F サウナ浴室 ろ過装置"), ("FILT-SAUNAC", "4F サウナ水風呂 ろ過装置"),
    ("FILT-OTAKI", "2F 大滝 ろ過装置 / 循環水"), ("FILT-UNKAI", "3F 雲海庭園の滝 ろ過"),
    ("FILT-FOUNT", "B2F ロータリーの噴水 ろ過"), ("FILT-RAIN", "B3F 雨水ろ過装置"),
]

def filt_points(dev, area):
    PANEL_HINT[dev] = "LCP-FILT"
    S = "Water filtration"
    T = "Filtration / circulation plant"
    add(f"{dev}.RUN", S, dev, T, area, "Circulation/filtration pump run", BI, "status", trend="Y")
    add(f"{dev}.TRIP", S, dev, T, area, "Pump / filter fault", BI, "status", alarm="Y")
    add(f"{dev}.SS", S, dev, T, area, "Start/Stop command", BO, "cmd", sp="Y", notes="Time schedule; backwash cycle")

def filtration_points():
    for dev, area in FILT_PLANTS:
        filt_points(dev, area)

# --------------------------------------------------------------------------
# Smoke control (排煙 / graph 5000): SMF-1..22 smoke-exhaust fans with a
# building fire-alarm batch (火災一括).  BMS monitors status and can start the
# smoke-mode fans; life-safety start is hard-wired from the fire panel.
# --------------------------------------------------------------------------
def smoke_fan_points(n):
    dev = f"SMF-{n}"
    PANEL_HINT[dev] = "LCP-SMK"
    S = "Smoke control"
    T = "Smoke-exhaust fan (排煙機)"
    area = "Smoke zone / floor per 排煙系統図"
    add(f"{dev}.RUN", S, dev, T, area, "Fan run status", BI, "status", trend="Y")
    add(f"{dev}.SS", S, dev, T, area, "Smoke-mode start/stop", BO, "cmd", notes="Life-safety start hard-wired from fire panel; BMS monitors + smoke-mode start")
    add(f"{dev}.TRIP", S, dev, T, area, "Fan fault", BI, "status", alarm="Y")

def smoke_points():
    for n in range(1, 23):
        smoke_fan_points(n)
    PANEL_HINT["SMK-FIRE"] = "LCP-SMK"
    add("SMK-FIRE.ALM", "Smoke control", "SMK-FIRE", "Fire-alarm interface", "Building",
        "Fire-alarm batch (火災一括) from disaster-prevention panel", BI, "status", alarm="Y",
        notes="Triggers smoke-mode fan start + AHU/damper shutdown interlocks")

# --------------------------------------------------------------------------
# Build all points  (order follows the savic-net 熱源系統 menu:
#   DHC受入 -> ホットウェル -> 冷却塔 -> 低層 -> 高層 -> 36,37F -> air side)
# --------------------------------------------------------------------------
# --- DHC intake (PRIMARY source) ---
dhc_chw_points("DHC-CHW", "B2F DHC chilled-water intake (CS/CR headers)")
dhc_steam_points("DHC-STEAM", "B2F DHC steam intake (SS 8k->2k / HR)")

# --- Hot well / condensate ---
hotwell_points("HWT-1", "B2F hot-well room")
for i in (1, 2):   # HP-5 condensate-return-to-DHC pumps
    PANEL_HINT[f"HP-5-{i}"] = "LCP-DHC"
    vfd_pump_points(f"HP-5-{i}", "Condensate-return pump (to DHC)", "B2F hot-well room",
                    "Heat source", notes_dp="HP-5; HWT-1 -> DHC hot-return")
for i in (1, 2):   # HP-3 pumps to B1F kitchen AHUs (AC-5/6/7)
    PANEL_HINT[f"HP-3-{i}"] = "LCP-DHC"
    vfd_pump_points(f"HP-3-{i}", "Kitchen-AHU hot-water pump", "B2F hot-well room",
                    "Heat source", notes_dp="HP-3; serves B1F 厨房 AC-5/6/7")

# --- Cooling towers / condenser water (packaged-unit + refrigeration rejection) ---
tower_cell_points("CT-1-1", "Rooftop", inv=True)
tower_cell_points("CT-1-2", "Rooftop", inv=True)
tower_common_points("CT-1", "Rooftop", filtration=False)
tower_cell_points("CT-2-1", "Rooftop", inv=True)
tower_cell_points("CT-2-2", "Rooftop", inv=True)
tower_common_points("CT-2", "Rooftop", filtration=True)
for i in (1, 2, 3):
    cdpump_points(f"CDP-1-{i}", "Condenser-water plant")
for i in (1, 2):
    cdpump_points(f"CDP-2-{i}", "Condenser-water plant")
freecool_points("EX-1", "Rooftop (winter free-cooling)")
emerg_hx_points("EMHX-1", "Low-floor emergency cooling (to DHC)")
emerg_hx_points("EMHX-2", "Low-floor emergency cooling (to DHC)")

# --- Low-rise distribution (低層系統) ---
oa_station_points("ST-1", "Low-rise OA station", "低層 OA sensor")
for i in (1, 2, 3):
    PANEL_HINT[f"CP-1-{i}"] = "LCP-HSL"
    vfd_pump_points(f"CP-1-{i}", "Primary CHW pump (low-rise)", "B2F machine room",
                    "Heat source", notes_dp="CP-1; sequenced/群起動; VFD; DP-reset")
for i in (1, 2, 3):
    PANEL_HINT[f"HP-1-{i}"] = "LCP-HSL"
    vfd_pump_points(f"HP-1-{i}", "Hot-water pump (low-rise)", "B2F machine room",
                    "Heat source", notes_dp="HP-1; sequenced; 4-pipe heating loop")
for i in (1, 2):
    PANEL_HINT[f"CP-4-{i}"] = "LCP-HSL"
    vfd_pump_points(f"CP-4-{i}", "B1F kitchen CHW pump", "B2F machine room", "Heat source",
                    notes_dp="CP-4; open/close + duty rotation")
for i in (1, 2):
    PANEL_HINT[f"CP-5-{i}"] = "LCP-HSL"
    vfd_pump_points(f"CP-5-{i}", "B1F kitchen pump", "B2F machine room", "Heat source",
                    notes_dp="CP-5")
ex_hx_points("EX-2-1", "B2F (laundry)", "LCP-HSL", "laundry")
ex_hx_points("EX-2-2", "B2F (laundry)", "LCP-HSL", "laundry")
ext_points("EXT-1", "B2F machine room (low-rise)", "LCP-HSL")

# --- High-rise distribution (高層系統) ---
oa_station_points("ST-2", "High-rise OA station", "高層 OA sensor")
for i in (1, 2, 3):
    PANEL_HINT[f"CP-2-{i}"] = "LCP-HSH"
    vfd_pump_points(f"CP-2-{i}", "Primary CHW pump (high-rise)", "PH machine room",
                    "Heat source", notes_dp="CP-2; sequenced/群起動; VFD; DP-reset")
for i in (1, 2, 3):
    PANEL_HINT[f"HP-2-{i}"] = "LCP-HSH"
    vfd_pump_points(f"HP-2-{i}", "Hot-water pump (high-rise)", "PH machine room",
                    "Heat source", notes_dp="HP-2; sequenced; 4-pipe heating loop")
for i in (1, 2):
    PANEL_HINT[f"CP-3-{i}"] = "LCP-HSH"
    vfd_pump_points(f"CP-3-{i}", "High-rise CHW pump (INV)", "PH machine room", "Heat source",
                    notes_dp="CP-3; INV; control-flow l/m")
for i in (1, 2):
    PANEL_HINT[f"CP-6-{i}"] = "LCP-HSH"
    vfd_pump_points(f"CP-6-{i}", "High-rise pump (seq changeover)", "PH machine room",
                    "Heat source", notes_dp="CP-6; 順序切替 ON:№1優先")
ex_hx_points("EX-3-1", "PH (high-rise kitchen)", "LCP-HSH", "kitchen")
ex_hx_points("EX-3-2", "PH (high-rise kitchen)", "LCP-HSH", "kitchen")
ext_points("EXT-2", "PH machine room (high-rise)", "LCP-HSH")

# --- 36/37F local heat source (R-1 + DHC changeover) ---
chiller_points("R-1", "37F PH machine room")
for i in (1, 2):
    cdpump_points(f"CDP-3-{i}", "37F PH (R-1 condenser)")
    PANEL_HINT[f"CDP-3-{i}"] = "LCP-3637"
tower_cell_points("CT-3-1", "37F PH (R-1)", inv=True)
tower_common_points("CT-3", "37F PH (R-1)", filtration=False)
for i in (1, 2):
    PANEL_HINT[f"CP-7-{i}"] = "LCP-3637"
    vfd_pump_points(f"CP-7-{i}", "R-1 CHW primary pump", "37F PH machine room", "Heat source",
                    notes_dp="CP-7; duty/standby; 発停/故障")
for i in (1, 2):
    PANEL_HINT[f"HP-4-{i}"] = "LCP-3637"
    vfd_pump_points(f"HP-4-{i}", "36/37F secondary loop pump", "37F PH machine room",
                    "Heat source", notes_dp="HP-4; EX4 secondary -> FCU/AC/EVU")
hex_points("HEX-1", "37F PH machine room")
for i in (1, 2, 3):
    PANEL_HINT[f"CP-8-{i}"] = "LCP-3637"
    vfd_pump_points(f"CP-8-{i}", "HEX-1 (DHC) primary pump", "37F PH machine room",
                    "Heat source", notes_dp="CP-8; DHC side of HEX-1; metered")
ex_hx_points("EX4", "37F PH (source/secondary buffer)", "LCP-3637", "36/37F loop")
ext_points("EXT-3", "37F PH machine room", "LCP-3637")
changeover_valve_points("CHGV-1", "37F PH (R-1 branch)", "R-1 evaporator branch valve (open = chiller mode)")
changeover_valve_points("CHGV-2", "37F PH (through-main)", "Through-main valve (open = DHC bypass mode)")

# Air side
for n in range(1, 28):
    ahu_points(n)
for n in range(1, 16):
    evu_points(n)
for tag, area, panel in FCU_GROUPS:
    fcu_group_points(tag, area, panel)

# Ventilation fans
for tag, parent, inv in FANS:
    fan_points(tag, parent, inv)

# Packaged units + kitchen refrigeration
for tag, area, always in PKG_UNITS:
    pkg_points(tag, area, always)
refrigeration_points()

# Domestic hot water (給湯) + sanitary/plumbing (衛生) + filtration (ろ過) + smoke (排煙)
for dev, area, basis in DHW_ZONES:
    dhw_points(dev, area, basis)
sanitary_points()
filtration_points()
smoke_points()

# Lighting (common-area / facade) + metering + common
lighting_points()
meter_points()
common_points()

# --------------------------------------------------------------------------
# Assign controllers (group devices per panel into DDCs of ~4 devices)
# --------------------------------------------------------------------------
# order devices by panel then first appearance
from collections import OrderedDict, defaultdict
panel_devices = OrderedDict()
for r in rows:
    r["Panel"] = r["Panel"] or PANEL_HINT.get(r["Device ID"], "LCP-PH")
    panel_devices.setdefault(r["Panel"], [])
    if r["Device ID"] not in panel_devices[r["Panel"]]:
        panel_devices[r["Panel"]].append(r["Device ID"])

def dev_category(d: str) -> str:
    """Class used to keep unlike equipment off the same DDC controller."""
    if d.startswith("DHC-CHW"): return "dhc_chw"
    if d.startswith("DHC-STEAM"): return "steam"
    if d.startswith("HWT-"): return "hotwell"
    if d.startswith(("HP-3", "HP-5")): return "hwpump"
    if d.startswith("CT-"): return "tower"
    if d.startswith("CDP-"): return "cwp"
    if d in ("EX-1",) or d.startswith("EMHX"): return "freecool"
    if d.startswith("R-1"): return "chiller"
    if d.startswith("HEX-"): return "hex"
    if d.startswith("CP-7") or d.startswith("CP-8"): return "chwp3637"
    if d.startswith("HP-4"): return "hwpump"
    if d in ("EX4",) or d.startswith("EXT"): return "vessel"
    if d.startswith("CHGV"): return "valve"
    if d.startswith("CP-"): return "chwp"
    if d.startswith("HP-"): return "hwpump"
    if d.startswith("EX-"): return "hex"
    if d.startswith("ST-"): return "oa"
    if d.startswith("MTR-"): return "meter"
    if d.startswith("DHW-"): return "dhw"
    if d.startswith(("MP-", "P-", "DP-", "BL-", "RCVT", "ELVT", "REUSET", "FIRET")): return "sanitary"
    if d.startswith("FILT-"): return "filtration"
    if d.startswith(("SMF-", "SMK-")): return "smoke"
    if d.startswith("POOL"): return "pool"
    if d.startswith("AC-"): return "ahu"
    if d.startswith("EVU-"): return "oau"
    if d.startswith("FCU-"): return "fcu"
    if d.startswith(("PAC-", "PCU-", "PMAC")): return "pkg"
    if d.startswith("REFR"): return "refr"
    if d.startswith("LTG-"): return "ltg"
    return "fan"   # EF/SF/RF

# Order categories so controllers list reads DHC -> plant -> air -> aux
CAT_ORDER = ["dhc_chw", "steam", "hotwell", "tower", "cwp", "freecool",
             "chiller", "hex", "chwp3637", "chwp", "hwpump", "vessel", "valve",
             "oa", "ahu", "oau", "fcu", "fan", "pkg", "dhw", "sanitary",
             "filtration", "smoke", "refr", "ltg", "pool", "meter"]

dev_controller = {}
controllers = []   # (ctrl_id, panel, devices)
for panel, devs in panel_devices.items():
    by_cat = OrderedDict()
    for d in devs:
        by_cat.setdefault(dev_category(d), []).append(d)
    pshort = panel.replace("LCP-", "").replace("BMS-", "")
    idx = 1
    for cat in CAT_ORDER:
        clist = by_cat.get(cat, [])
        # chunk same-category devices by 3 into one controller
        for i in range(0, len(clist), 3):
            chunk = clist[i:i + 3]
            cid = f"DDC-{pshort}-{idx:02d}"
            for d in chunk:
                dev_controller[d] = cid
            controllers.append([cid, panel, ", ".join(chunk)])
            idx += 1

for r in rows:
    r["Controller"] = dev_controller.get(r["Device ID"], "")

# --------------------------------------------------------------------------
# Physical-panel cross-reference — bridge to the 판넬별 포인트 정리 physical
# schedule (panels_schedule.json). Each physical RS/RCP/CP panel now lists the
# equipment it serves (AC-* AHUs, EVU-* OAUs, PCU packaged units, SF/EF fans);
# we map those tags back to the functional Device IDs to stamp each point with
# the physical panel + floor it is wired into. Best-effort: air-side AHUs/OAUs
# match cleanly; SF/EF fans are not individually deviced in the functional model
# so they stay blank. This never fabricates points.
# --------------------------------------------------------------------------
import json as _json
import re as _re


def _load_phys_schedule():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "panels_schedule.json")
    try:
        return _json.load(open(path, encoding="utf-8"))
    except Exception:
        return None


PHYS_SCHED = _load_phys_schedule()


def _equip_tokens(equip):
    if not equip:
        return []
    s = _re.sub(r"\(.*?\)", "", equip)
    out = []
    for part in _re.split(r"[,\.;/]", s):
        t = part.strip().upper().replace(" ", "")
        if _re.match(r"^(AC|EVU|PCU|PAC|PMAC|SF|EF)-?\d", t):
            out.append(t)
    return out


DEV2PANEL = {}
if PHYS_SCHED:
    for _p in PHYS_SCHED["panels"]:
        for _t in _equip_tokens(_p.get("equip", "")):
            DEV2PANEL.setdefault(_t, (_p["name"], _p.get("floor", "")))


def _phys_lookup(device_id):
    d = _re.sub(r"\(.*?\)", "", str(device_id or "")).strip().upper().replace(" ", "")
    for cand in (d, _re.sub(r"-\d+$", "", d)):
        if cand in DEV2PANEL:
            return DEV2PANEL[cand]
    return ("", "")


# --------------------------------------------------------------------------
# Build the workbook.  The device/point model above (rows, controllers,
# PANELS, PANEL descriptions) is importable without side effects; only the
# actual .xlsx write happens under the __main__ guard at the bottom.
# --------------------------------------------------------------------------
def build_workbook():
  wb = Workbook()
  HDR_FILL = PatternFill("solid", fgColor="1F3864")
  HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
  TITLE_FONT = Font(bold=True, size=14, color="1F3864")
  SUB_FONT = Font(italic=True, size=10, color="555555")
  thin = Side(style="thin", color="D9D9D9")
  BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
  SYS_FILL = {
      "DHC interface": "FFF2CC", "Heat source": "FCE4D6", "Condenser water": "DDEBF7",
      "Air side": "E2EFDA", "Ventilation": "DEEBF7", "Packaged units": "FCE8F3",
      "Domestic hot water": "FBE2D5", "Plumbing/Sanitary": "DAEEF3",
      "Water filtration": "E4F0D0", "Smoke control": "F4CCCC",
      "Lighting": "FFF7DA", "Metering": "EDEDED", "Common": "F2F2F2",
  }

  def style_header(ws, ncols, row=1):
      for c in range(1, ncols + 1):
          cell = ws.cell(row=row, column=c)
          cell.fill = HDR_FILL; cell.font = HDR_FONT
          cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
          cell.border = BORDER

  def set_widths(ws, widths):
      for i, w in enumerate(widths, 1):
          ws.column_dimensions[get_column_letter(i)].width = w

  # ---- Cover ----
  ws = wb.active
  ws.title = "Cover"
  ws["A1"] = "Red5-DHCP — BMS I/O List & Panel / Controller Schedule"
  ws["A1"].font = TITLE_FONT
  ws["A2"] = "ANA InterContinental Tokyo — District Heating & Cooling (DHC) connected hotel BMS"
  ws["A2"].font = SUB_FONT
  cover = [
      ("", ""),
      ("Building", "SRC, B3–37F, ~98,331 m²; Azbil savic-net FX / FX2 BMS"),
      ("Energy source", "District Heating & Cooling (PRIMARY): DHC chilled-water intake (冷水受入 CS/CR, GJ metered) + DHC steam (蒸気受入 SS 0.8→0.2 MPa, condensate HR to hot-well). Local plant is backup/peaking."),
      ("Distribution", "Low-rise (低層): CP-1 ×3 CHW + HP-1 ×3 HW pumps, EX-2 laundry HX. High-rise (高層): CP-2 ×3 CHW + HP-2 ×3 HW, CP-3(INV)/CP-6, EX-3 HX. Expansion EXT-1/2; OA stations ST-1/ST-2."),
      ("36/37F local plant", "R-1 single water-cooled SCREW chiller (Ebara RHS DW202M2, 370 kW / 82.2 kW in / COP 4.5, twin 45 kW screw comp, R-407C) in changeover with DHC via HEX-1 + CP-8; CDP-3 + CT-3 condenser; CP-7 CHW; HP-4 + EX4 + EXT-3 secondary; 2 source-changeover valves"),
      ("Condenser water", "CT-1 (2 INV cells) + CT-2 (2 cells + filtration) + CDP-1 ×3 / CDP-2 ×2 pumps rejecting heat from packaged units (PCU/PAC/PMAC) & kitchen refrigeration; EX-1 winter free-cooling; 2 emergency cooling HX to DHC"),
      ("Air side", "AC-1..27 AHUs (public/kitchen), EVU-1..15 outdoor-air units, guest-room FCU zone-groups (5-20F & 21-35F × N/NE/SE/S/SW, 4-pipe) + 36/37F FCU, EF/SF/RF fans"),
      ("Also on savic-net", "Packaged units PCU/PAC/PMAC (status/alarm), kitchen refrigeration alarms, common-area/facade lighting (照明一覧)"),
    ("Building services", "Domestic hot water (給湯: steam-heated storage DHW-1..4 by vertical zone), plumbing/sanitary (上水/中水 booster MP/P, receiving/elevated/fire tanks, drainage sump pumps DP-1..14 + aeration blowers BL), recreational-water filtration (baths/sauna/pool/waterfalls/fountain/rainwater), smoke control (排煙 fans SMF-1..22) — surfaced from the live savic-net FX summary graphs (graphs 4100/4200/4300/4400/5000)"),
    ("Comms / network", "savic-netFX2: BMS/SMS/DSS servers + 2 operator PCs (B2F 防災室 / 2F 防災センター) on Ethernet (ESW1/2/3); 6 SCS System Core Servers → ≤24 NC-bus lines → RS/DDC (= 67 physical panels), IPEV-S 0.9mm² trunk — see Comms-Network sheet"),
    ("ESCO control scope", "Pump optimization, outdoor-air-unit optimization, thermal-demand control (per 共-01 spec)"),
    ("BMS platform", "Azbil savic-net FX2, S/W 機能仕様 dated 2015/04/09 — confirmed from the 144-page Azbil 納入仕様書 (software function spec, not an equipment list)"),
    ("", ""),
    ("Sheets", "Panels · Controllers · IO_List (+ Phys. panel/floor cross-ref) · IO_Summary · Reconciliation · Points-by-Floor · LCP-to-Physical · Delta-Controllers · Comms-Network · Legend"),
    ("Basis", "Azbil 完成図 savic-net graphics (熱源設備 全体/低層/高層/36,37F/冷却塔/ホットウェル/DHC受入, graphs 1000-1100; PMAC-PAC & 照明一覧), air-side summary graphs (空調関連/客室等), M-* CAD equipment schedules, 共-01 ESCO spec, 空調機スケジュール_20251127.xlsx"),
    ("Note", "Device tags/counts read directly from the savic-net graphics. Per-unit packaged (PCU/PAC/PMAC) and per-floor lighting lists are modelled as representative supervised groups — expand from 照明一覧1/2 + PMAC/PAC一覧 for the full point count."),
  ]
  r = 4
  for k, v in cover:
      ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
      ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
      r += 1
  set_widths(ws, [22, 110])
  for rr in range(5, r):
      ws.row_dimensions[rr].height = 30

  # ---- Panels ----
  ws = wb.create_sheet("Panels")
  pcols = ["Panel ID", "Description", "Location", "Areas / Systems served",
           "Controllers", "I/O points"]
  ws.append(pcols)
  pt_by_panel = defaultdict(int)
  for rr in rows:
      pt_by_panel[rr["Panel"]] += 1
  ctrl_by_panel = defaultdict(int)
  for cid, panel, devs in controllers:
      ctrl_by_panel[panel] += 1
  for pid, desc, loc, served in PANELS:
      ws.append([pid, desc, loc, served, ctrl_by_panel.get(pid, 0), pt_by_panel.get(pid, 0)])
  style_header(ws, len(pcols))
  set_widths(ws, [14, 34, 26, 60, 12, 11])
  ws.freeze_panes = "A2"
  ws.auto_filter.ref = f"A1:{get_column_letter(len(pcols))}{ws.max_row}"
  for rr in range(2, ws.max_row + 1):
      for c in range(1, len(pcols) + 1):
          ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
          ws.cell(row=rr, column=c).border = BORDER

  # ---- Controllers ----
  ws = wb.create_sheet("Controllers")
  ccols = ["Controller ID", "Panel", "Type", "Devices served", "Point count"]
  ws.append(ccols)
  pts_by_ctrl = defaultdict(int)
  for rr in rows:
      pts_by_ctrl[rr["Controller"]] += 1
  for cid, panel, devs in controllers:
      ws.append([cid, panel, "DDC field controller (Azbil Infilex/savic-net class)",
                 devs, pts_by_ctrl.get(cid, 0)])
  style_header(ws, len(ccols))
  set_widths(ws, [16, 12, 42, 60, 12])
  ws.freeze_panes = "A2"
  ws.auto_filter.ref = f"A1:{get_column_letter(len(ccols))}{ws.max_row}"
  for rr in range(2, ws.max_row + 1):
      for c in range(1, len(ccols) + 1):
          ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
          ws.cell(row=rr, column=c).border = BORDER

  # ---- IO_List ----
  ws = wb.create_sheet("IO_List")
  cols = ["Point Tag", "System", "Device ID", "Device Type", "Area / Served",
          "Panel", "Controller", "Phys. panel (26.07.30)", "Phys. floor",
          "Point Description", "I/O Type",
          "Signal / Range", "Units", "SP/Sched", "Alarm", "Trend", "Notes / Basis"]
  ws.append(cols)
  _src_cols = ["Point Tag", "System", "Device ID", "Device Type", "Area / Served",
               "Panel", "Controller", "Point Description", "I/O Type",
               "Signal / Range", "Units", "SP/Sched", "Alarm", "Trend", "Notes / Basis"]
  for rr in rows:
      pp, pf = _phys_lookup(rr["Device ID"])
      vals = {c: rr[c] for c in _src_cols}
      vals["Phys. panel (26.07.30)"] = pp
      vals["Phys. floor"] = pf
      ws.append([vals[c] for c in cols])
  style_header(ws, len(cols))
  set_widths(ws, [16, 13, 12, 24, 34, 10, 13, 15, 9, 34, 8, 15, 8, 8, 7, 7, 46])
  ws.freeze_panes = "C2"
  ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
  # system color banding + borders
  for rr in range(2, ws.max_row + 1):
      sysname = ws.cell(row=rr, column=2).value
      fill = SYS_FILL.get(sysname)
      for c in range(1, len(cols) + 1):
          cell = ws.cell(row=rr, column=c)
          cell.border = BORDER
          cell.alignment = Alignment(wrap_text=True, vertical="top", size=9) if False else Alignment(wrap_text=True, vertical="top")
          cell.font = Font(size=9)
          if fill:
              cell.fill = PatternFill("solid", fgColor=fill)

  # ---- IO_Summary ----
  ws = wb.create_sheet("IO_Summary")
  systems = ["DHC interface", "Heat source", "Condenser water", "Air side",
             "Ventilation", "Packaged units", "Domestic hot water", "Plumbing/Sanitary",
             "Water filtration", "Smoke control", "Lighting", "Metering", "Common"]
  counts = {s: {"AI": 0, "AO": 0, "BI": 0, "BO": 0} for s in systems}
  for rr in rows:
      counts.setdefault(rr["System"], {"AI": 0, "AO": 0, "BI": 0, "BO": 0})
      counts[rr["System"]][rr["I/O Type"]] += 1
  scols = ["System", "AI", "AO", "BI", "BO", "Total"]
  ws.append(scols)
  tot = {"AI": 0, "AO": 0, "BI": 0, "BO": 0}
  for s in systems:
      c = counts[s]
      row_total = sum(c.values())
      ws.append([s, c["AI"], c["AO"], c["BI"], c["BO"], row_total])
      for k in tot:
          tot[k] += c[k]
  ws.append(["GRAND TOTAL", tot["AI"], tot["AO"], tot["BI"], tot["BO"], sum(tot.values())])
  style_header(ws, len(scols))
  set_widths(ws, [18, 8, 8, 8, 8, 10])
  last = ws.max_row
  for c in range(1, len(scols) + 1):
      ws.cell(row=last, column=c).font = Font(bold=True)
      ws.cell(row=last, column=c).fill = PatternFill("solid", fgColor="D6DCE4")
  for rr in range(2, ws.max_row + 1):
      for c in range(1, len(scols) + 1):
          ws.cell(row=rr, column=c).border = BORDER
  # device + point tallies
  ws.cell(row=last + 2, column=1, value="Devices").font = Font(bold=True)
  ws.cell(row=last + 2, column=2, value=len({r['Device ID'] for r in rows}))
  ws.cell(row=last + 3, column=1, value="Panels").font = Font(bold=True)
  ws.cell(row=last + 3, column=2, value=len(PANELS))
  ws.cell(row=last + 4, column=1, value="Controllers").font = Font(bold=True)
  ws.cell(row=last + 4, column=2, value=len(controllers))
  ws.cell(row=last + 5, column=1, value="Total I/O points").font = Font(bold=True)
  ws.cell(row=last + 5, column=2, value=len(rows))

  # ---- Reconciliation (functional model vs 26.07.30 physical schedule) ----
  ws = wb.create_sheet("Reconciliation")
  func_pts = len(rows)
  func_type = {"AI": 0, "AO": 0, "BI": 0, "BO": 0}
  for rr in rows:
      func_type[rr["I/O Type"]] = func_type.get(rr["I/O Type"], 0) + 1
  linked = sum(1 for rr in rows if _phys_lookup(rr["Device ID"])[0])
  if PHYS_SCHED:
      pt = PHYS_SCHED["typeTotals"]
      phys_used = PHYS_SCHED["usedTotal"]
      phys_cap = PHYS_SCHED["capTotal"]
      phys_pan = len(PHYS_SCHED["panels"])
      phys_mod = sum(PHYS_SCHED["modTotals"])
      phys_equip = sum(1 for _p in PHYS_SCHED["panels"] if _p.get("equip"))
      src = PHYS_SCHED.get("source", "")
  else:
      pt = {"DO": 0, "DI": 0, "BTOT": 0, "AI": 0, "AO": 0}
      phys_used = phys_cap = phys_pan = phys_mod = phys_equip = 0
      src = "(panels_schedule.json missing)"
  ws["A1"] = "Reconciliation — functional model vs physical panel schedule"
  ws["A1"].font = TITLE_FONT
  ws["A2"] = (f"This workbook is the drawing-derived FUNCTIONAL model (tagged points + SOO). "
              f"판넬별 포인트 정리 {src} is the PHYSICAL panel schedule (per-panel type counts).")
  ws["A2"].font = SUB_FONT
  ws.append([])
  block = [
      ["Metric", "Functional model", f"Physical schedule ({src})"],
      ["Scope", "tagged points + description + SOO", "per-panel type counts (no per-point tags)"],
      ["Total I/O points", func_pts, phys_used],
      ["Panels", len(PANELS), phys_pan],
      ["Controllers / modules", len(controllers), phys_mod],
      ["Installed capacity (points)", "—", phys_cap],
      ["Panels naming equipment served", "—", phys_equip],
      ["", "", ""],
      ["Point types", "functional (AI/AO/BI/BO)", "physical (DO/DI/BTOT/AI/AO)"],
      ["Analog input (AI)", func_type.get("AI", 0), pt["AI"]],
      ["Analog output (AO)", func_type.get("AO", 0), pt["AO"]],
      ["Binary input (BI ≈ DI)", func_type.get("BI", 0), pt["DI"]],
      ["Binary output (BO ≈ DO)", func_type.get("BO", 0), pt["DO"]],
      ["Pulse / totalizer (BTOT)", "—", pt["BTOT"]],
      ["", "", ""],
      ["Overlap (characterized, purpose known)", func_pts, func_pts],
      ["Newly surfaced, count-only (tags TBD)", "—", max(phys_used - func_pts, 0)],
      ["Functional points cross-linked to a physical panel",
       linked, f"{(100 * linked // func_pts) if func_pts else 0}% of {func_pts}"],
  ]
  for row in block:
      ws.append(row)
  style_header(ws, 3, row=4)
  set_widths(ws, [46, 30, 40])
  for rr in range(5, ws.max_row + 1):
      for c in (1, 2, 3):
          ws.cell(row=rr, column=c).border = BORDER
          ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
      ws.cell(row=rr, column=1).font = Font(bold=True, size=10)
  ws.append([])
  note_r = ws.max_row + 1
  ws.cell(row=note_r, column=1, value="Note")
  ws.cell(row=note_r, column=1).font = Font(bold=True, size=10)
  ws.cell(row=note_r, column=2, value=(
      "The physical schedule supplies only per-panel type counts, not individual point tags, so the "
      f"{max(phys_used - func_pts, 0)} count-only points cannot be expanded into fully-described tagged "
      "points from that source. The full physical inventory (per panel × type, floor, equipment served) "
      "lives in Red5-DHCP-Panels-IO; every physical point is enumerated as a positional tag in the "
      "red5-dhcp_commissioning sheets. The 'Phys. panel' / 'Phys. floor' columns on IO_List link the "
      "functional AHU/OAU/packaged devices to the physical panel that carries them."))
  ws.cell(row=note_r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
  ws.row_dimensions[note_r].height = 90

  # ---- Points-by-Floor (physical 4,467 accounting across the 67 panels) ----
  ws = wb.create_sheet("Points-by-Floor")
  ws["A1"] = "Physical panels accounting for the full 4,467 points (판넬별 포인트 정리 physical schedule)"
  ws["A1"].font = TITLE_FONT
  ws["A2"] = ("The 4,467 installed I/O points live in 67 physical panels (RS/RCP/CP cabinets), grouped "
              "here by floor. The 1,160 functional points in this workbook are a drawing-derived subset "
              "bucketed into 12 logical LCP panels — see the LCP-to-Physical sheet for how they relate.")
  ws["A2"].font = SUB_FONT
  ws.append([])
  AREA_ORDER = ["Penthouse (PH)", "Floor 31", "Floor 21", "Floor 10", "Floor 4", "Floor 3",
                "Floor 2", "Floor 1", "Basement B1", "Basement B2", "Basement B3"]
  if PHYS_SCHED:
      by_area = defaultdict(list)
      for _p in PHYS_SCHED["panels"]:
          by_area[_p["area"]].append(_p)
      ws.append(["Floor / area", "Panels", "Used points", "Capacity"])
      sum_hrow = ws.max_row
      style_header(ws, 4, row=sum_hrow)
      gp = gu = gc = 0
      for a in AREA_ORDER:
          ps = by_area.get(a, [])
          u = sum(x["usedTot"] for x in ps)
          c = sum(x["capTot"] for x in ps)
          ws.append([a, len(ps), u, c])
          gp += len(ps); gu += u; gc += c
      ws.append(["TOTAL", gp, gu, gc])
      for c in range(1, 5):
          ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
          ws.cell(row=ws.max_row, column=c).fill = PatternFill("solid", fgColor="D6DCE4")
      for rr in range(sum_hrow, ws.max_row + 1):
          for c in range(1, 5):
              ws.cell(row=rr, column=c).border = BORDER
      ws.append([])
      ws.append(["Per-panel detail (67 physical panels)"])
      ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
      det_hrow = ws.max_row + 1
      ws.append(["#", "Panel", "Floor", "Equipment served", "Used", "Cap"])
      style_header(ws, 6, row=det_hrow)
      i = 0
      for a in AREA_ORDER:
          for _p in by_area.get(a, []):
              i += 1
              ws.append([i, _p["name"], _p.get("floor", ""), _p.get("equip", ""),
                         _p["usedTot"], _p["capTot"]])
      for rr in range(det_hrow, ws.max_row + 1):
          for c in range(1, 7):
              ws.cell(row=rr, column=c).border = BORDER
              ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
      set_widths(ws, [5, 16, 10, 42, 8, 8])
  else:
      ws.append(["panels_schedule.json not found — physical accounting unavailable"])

  # ---- LCP-to-Physical (logical grouping → physical panels) ----
  ws = wb.create_sheet("LCP-to-Physical")
  ws["A1"] = "Logical (LCP) panels → physical panels"
  ws["A1"].font = TITLE_FONT
  ws["A2"] = ("Each functional LCP grouping mapped to the physical RS/RCP/CP panel(s) that carry its "
              "equipment, via the equipment-served bridge. Plant / DHC / metering / lighting devices are "
              "not equipment-tagged in the physical schedule, so those rows show the drawing location + an "
              "engineering-judgement hint instead of a hard match.")
  ws["A2"].font = SUB_FONT
  ws.append([])
  CURATED = {
      "BMS-CENTRAL": "= 시스템 제어반 (B2F) — the physical central/system panel",
      "LCP-DHC": "DHC intake devices in B2F machine-room CP-* panels (not equip-tagged)",
      "LCP-CT": "Tower/condenser devices on rooftop / PH plant panels (not equip-tagged)",
      "LCP-HSL": "Low-rise CHW/HW pumps + HX in basement machine-room CP-* panels",
      "LCP-HSH": "High-rise + 36/37F pumps/HX incl. R-1 → CP-R-1 (PH1) + basement CP-*",
      "LCP-LTG": "Common-area/facade lighting distributed across RS-* room panels",
      "LCP-PKG": "Packaged units → CP-B1-2 / CP-B1-4 / RCP-B1-5 (PCU tags)",
  }
  pts_by_lcp = defaultdict(int)
  linked_by_lcp = defaultdict(int)
  phys_by_lcp = defaultdict(set)
  for rr in rows:
      lcp = rr["Panel"]
      pts_by_lcp[lcp] += 1
      pp, pf = _phys_lookup(rr["Device ID"])
      if pp:
          linked_by_lcp[lcp] += 1
          phys_by_lcp[lcp].add(f"{pp} ({pf})")
  lcp_cols = ["LCP panel", "Drawing location", "Points", "Matched physical panels (floor)",
              "Linked pts", "Coverage", "Note (engineering judgement)"]
  ws.append(lcp_cols)
  lcp_hrow = ws.max_row
  style_header(ws, len(lcp_cols), row=lcp_hrow)
  for pid, desc, loc, served in PANELS:
      pl = pts_by_lcp.get(pid, 0)
      matched = ", ".join(sorted(phys_by_lcp.get(pid, []))) or "—"
      lk = linked_by_lcp.get(pid, 0)
      cov = f"{(100 * lk // pl) if pl else 0}%"
      ws.append([pid, loc, pl, matched, lk, cov, CURATED.get(pid, "")])
  for rr in range(lcp_hrow, ws.max_row + 1):
      for c in range(1, len(lcp_cols) + 1):
          ws.cell(row=rr, column=c).border = BORDER
          ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
  set_widths(ws, [14, 26, 8, 46, 10, 10, 46])
  ws.append([])
  cav_r = ws.max_row + 1
  ws.cell(row=cav_r, column=1, value="How to read this").font = Font(bold=True, size=10)
  ws.cell(row=cav_r, column=2, value=(
      "The 'Matched physical panels' column is a BEST-EFFORT bridge: a functional device (e.g. AC-17) is "
      "linked to a physical panel only when that panel's 'equipment served' column names the same tag. "
      "So coverage is high for the air-side floor LCPs (AHUs/OAUs) and partial elsewhere. Rows at 0% "
      "coverage (BMS-CENTRAL, LCP-DHC, LCP-CT, LCP-LTG) hold plant / DHC / metering / lighting points that "
      "are NOT equipment-tagged in the physical schedule — use the drawing location + judgement note for "
      "those. Caveat: matching is by tag string only, so a repeated fan/unit number can produce a "
      "cross-floor coincidence (e.g. a 2F SF tag also appearing on a basement fan panel) — verify against "
      "the drawings before relying on a single-panel assignment."))
  ws.cell(row=cav_r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
  ws.row_dimensions[cav_r].height = 96

  # ---- Delta-Controllers (physical controller/module tally + 4,467 points by panel) ----
  ws = wb.create_sheet("Delta-Controllers")
  ws["A1"] = "Delta controllers & I/O modules — type tally + all 4,467 points grouped by panel"
  ws["A1"].font = TITLE_FONT
  ws["A2"] = ("The physical schedule lists, per panel, the Delta 'controller & I/O module' fleet that "
              "carries its points. Below: the type tally (unit counts) and every panel's controller "
              "composition with its used-point total, so the full 4,467 points roll up to the 904 units.")
  ws["A2"].font = SUB_FONT
  ws.append([])
  MOD_LABELS = ["Red5-PLUS-ROOM", "CON-T1L", "Red5-EXPAND-04",
                "Red5-MODULE-4F4xP", "Red5-MODULE-8xP", "Red5-MODULE-8PxP"]
  MOD_ROLE = ["Room controller (BACnet, onboard I/O)", "Network/BACnet controller",
              "Bus expander (hosts I/O modules)", "I/O module", "I/O module", "I/O module"]
  if PHYS_SCHED:
      mt = PHYS_SCHED["modTotals"]
      total_mod = sum(mt)
      used_tot = PHYS_SCHED["usedTotal"]
      cap_tot = PHYS_SCHED["capTotal"]
      # --- type tally ---
      ws.append(["Controller / module type", "Role", "Units", "Share of fleet"])
      t_hrow = ws.max_row
      style_header(ws, 4, row=t_hrow)
      for i, lab in enumerate(MOD_LABELS):
          share = f"{(100 * mt[i] / total_mod):.1f}%" if total_mod else "0%"
          ws.append([lab, MOD_ROLE[i], mt[i], share])
      ws.append(["TOTAL", "", total_mod, "100%"])
      for c in range(1, 5):
          ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
          ws.cell(row=ws.max_row, column=c).fill = PatternFill("solid", fgColor="D6DCE4")
      for rr in range(t_hrow, ws.max_row + 1):
          for c in range(1, 5):
              ws.cell(row=rr, column=c).border = BORDER
              ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
      # --- role rollup ---
      controllers_n = mt[0] + mt[1]
      expanders_n = mt[2]
      iomods_n = mt[3] + mt[4] + mt[5]
      ws.append([])
      ws.append(["Rollup", "", "", ""])
      roll_h = ws.max_row
      ws.cell(row=roll_h, column=1).font = Font(bold=True, size=11)
      for lab, val in [
          ("Controllers (ROOM + CON-T1L)", controllers_n),
          ("Bus expanders (EXPAND-04)", expanders_n),
          ("I/O modules (4F4xP + 8xP + 8PxP)", iomods_n),
          ("Total controllers + modules", total_mod),
          ("Used I/O points", used_tot),
          ("Installed point capacity", cap_tot),
          ("Fleet point utilisation", f"{(100 * used_tot / cap_tot):.0f}%" if cap_tot else "—"),
      ]:
          ws.append([lab, val])
          ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10)
      # --- per-panel controller composition + points ---
      ws.append([])
      ws.append(["Per-panel controller composition (all 4,467 points grouped to each panel's fleet)"])
      ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
      p_hrow = ws.max_row + 1
      pcols = ["#", "Panel", "Floor", "ROOM", "CON-T1L", "EXPAND-04",
               "4F4xP", "8xP", "8PxP", "Modules", "Used pts", "Cap"]
      ws.append(pcols)
      style_header(ws, len(pcols), row=p_hrow)
      AREA_ORD = ["Penthouse (PH)", "Floor 31", "Floor 21", "Floor 10", "Floor 4", "Floor 3",
                  "Floor 2", "Floor 1", "Basement B1", "Basement B2", "Basement B3"]
      by_area2 = defaultdict(list)
      for _p in PHYS_SCHED["panels"]:
          by_area2[_p["area"]].append(_p)
      i = 0
      col_mod = [0, 0, 0, 0, 0, 0]
      for a in AREA_ORD:
          for _p in by_area2.get(a, []):
              i += 1
              m = _p["mods"]
              for k in range(6):
                  col_mod[k] += m[k]
              ws.append([i, _p["name"], _p.get("floor", ""), m[0], m[1], m[2], m[3], m[4], m[5],
                         sum(m), _p["usedTot"], _p["capTot"]])
      ws.append(["", "TOTAL", "", col_mod[0], col_mod[1], col_mod[2], col_mod[3], col_mod[4], col_mod[5],
                 total_mod, used_tot, cap_tot])
      for c in range(1, len(pcols) + 1):
          ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
      for rr in range(p_hrow, ws.max_row + 1):
          for c in range(1, len(pcols) + 1):
              ws.cell(row=rr, column=c).border = BORDER
              ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
      set_widths(ws, [5, 16, 10, 8, 9, 11, 8, 7, 7, 9, 9, 7])
      ws.append([])
      note_r2 = ws.max_row + 1
      ws.cell(row=note_r2, column=1, value="Note").font = Font(bold=True, size=10)
      ws.cell(row=note_r2, column=2, value=(
          "The physical schedule tallies points at the panel level and controllers/modules by type; it "
          "does not assign each point to a specific module. Points are therefore grouped to the panel's "
          "whole controller fleet (Used pts vs installed capacity), and the type tally above gives the "
          "controller-type counts. Naming: Red5-PLUS-ROOM & CON-T1L are controllers; Red5-EXPAND-04 is a "
          "bus expander; Red5-MODULE-4F4xP/8xP/8PxP are I/O point modules."))
      ws.cell(row=note_r2, column=2).alignment = Alignment(wrap_text=True, vertical="top")
      ws.row_dimensions[note_r2].height = 70
  else:
      ws.append(["panels_schedule.json not found — controller tally unavailable"])

  # ---- Comms-Network (savic-netFX2 central + fieldbus architecture) ----
  ws = wb.create_sheet("Comms-Network")
  ws["A1"] = "Central monitoring & fieldbus architecture (Azbil savic-netFX2)"
  ws["A1"].font = TITLE_FONT
  ws["A2"] = ("Read from the ESCO hardware delivery spec 工番 1-LHP1-11 (2015/04/09): drawings "
              "LHP1-11-200-01 system config, -200-02 trunk, -100-B0xx equipment lists, -202-xx SCS "
              "connections. Trunk floor layout from the 2002 伝送幹線系統図 (dwg A-20).")
  ws["A2"].font = SUB_FONT
  ws.append([])
  ws.append(["Central station (Ethernet backbone — IPv4/IPv6, 100BASE-TX, fibre risers)"])
  ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
  h1 = ws.max_row + 1
  ws.append(["Component", "Role", "Model / detail", "Location"])
  style_header(ws, 4, row=h1)
  central = [
      ("BMS", "Building-management server", "NEC FC-E21A (ビルマネジメントサーバ)", "B2F 防災室 (panel LHP1-11-101)"),
      ("SMS", "System-management server", "Azbil SI-net BCY45300W0020", "B2F 防災室 (101)"),
      ("DSS", "Data-storage server", "Azbil SI-net BCY46300W0020", "B2F 防災室 (101)"),
      ("監視用PC1", "Operator workstation #1", "Fujitsu ESPRIMO D551/G · EIZO 22\" LCD", "B2F 防災室 (101)"),
      ("監視用PC2", "Operator workstation #2", "Fujitsu ESPRIMO D551/G · EIZO 22\" LCD", "2F 防災センター (panel LHP1-11-103)"),
      ("CLP", "Colour laser printer", "Brother HL-4570CDW", "B2F 防災室 (101)"),
      ("ESW1 / ESW3", "Ethernet switches (8-port)", "Hitachi APLGB108SS", "101 / 103"),
      ("ESW2", "Ethernet switch (16-port)", "Allied Telesis FS816S", "System-control panel 102"),
      ("IP plan", "Addressing", "192.168.30.x / 192.168.31.x (+ IPv6 FE80::)", "per -200-02 address table"),
  ]
  for row in central:
      ws.append(list(row))
  for rr in range(h1, ws.max_row + 1):
      for c in range(1, 5):
          ws.cell(row=rr, column=c).border = BORDER
          ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
  ws.append([])
  ws.append(["Fieldbus hierarchy (system-control panel 102 → field)"])
  ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
  h2 = ws.max_row + 1
  ws.append(["Tier", "Device", "Qty", "Notes"])
  style_header(ws, 4, row=h2)
  fieldbus = [
      ("Sub-server", "SCS — システム・コア・サーバ (System Core Server)", "6", "BCY44200W0000, on ESW2; each heads up to 4 NC-bus lines"),
      ("Fieldbus", "NC-bus lines", "≤24", "≤4 lines per SCS; transmission trunk IPEV-S 0.9 mm² (TX-IN→TX-OUT daisy-chain)"),
      ("Remote station", "RS / DDC (= the 67 physical panels)", "67", "RS-* / RCP-* / CP-* cabinets — see Points-by-Floor & Delta-Controllers sheets (4,467 pts)"),
      ("Fire/ext I/O", "IFGD1 Inf-GD module", "1", "DI 16 / DO 8 — fire-alarm & external signals (火災入力・外部信号)"),
  ]
  for row in fieldbus:
      ws.append(list(row))
  for rr in range(h2, ws.max_row + 1):
      for c in range(1, 5):
          ws.cell(row=rr, column=c).border = BORDER
          ws.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
  set_widths(ws, [16, 44, 8, 60])
  ws.append([])
  nr = ws.max_row + 1
  ws.cell(row=nr, column=1, value="Note").font = Font(bold=True, size=10)
  ws.cell(row=nr, column=2, value=(
      "The functional model in this workbook groups points into 16 logical LCP panels for control-logic "
      "authoring; the drawings show the true physical segmentation is 6 SCS → ≤24 NC-bus lines → 67 RS/DDC "
      "remote stations. The 2002 伝送幹線系統図 organises those panels by floor onto named UIC1–4 LINE "
      "segments (high-rise risers to 10F/21F/31F; basement UIC4 group) — the trunk this FX2 retrofit rides on. "
      "Full write-up: docs/reconciliation_2026-08.md."))
  ws.cell(row=nr, column=2).alignment = Alignment(wrap_text=True, vertical="top")
  ws.row_dimensions[nr].height = 84

  # ---- Legend ----
  ws = wb.create_sheet("Legend")
  legend = [
      ("Abbreviation", "Meaning"),
      ("AI", "Analog Input (sensor: temperature, humidity, pressure, flow, power, position fb)"),
      ("AO", "Analog Output (modulating command: valve %, damper %, VFD speed %)"),
      ("BI", "Binary Input (status / fault / proof dry contact)"),
      ("BO", "Binary Output (start/stop / open-close relay)"),
      ("SP/Sched", "Point carries a setpoint or time/seasonal/event schedule"),
      ("CHW / CW", "Chilled water / Condenser water"),
      ("SAT / RAT", "Supply-air temp / Return-air temp"),
      ("CCV / HCV", "Cooling-coil valve / Heating-coil (steam) valve"),
      ("INV", "Inverter / VFD fan speed control"),
      ("Zero-energy band", "Deadband between heating & cooling to prevent simultaneous operation"),
      ("DHC", "District Heating & Cooling (external steam / chilled supply) — PRIMARY source"),
      ("HEX / EX", "Plate heat exchanger (HEX-1 DHC tie-in; EX-2 laundry, EX-3 kitchen, EX4 36/37F buffer)"),
      ("EX-1", "Winter free-cooling heat exchanger (冬期用熱交換器)"),
      ("CS / CR", "DHC chilled supply / return header"),
      ("SS / HR", "DHC steam supply / hot (condensate) return; 8k=0.8 MPa, 2k=0.2 MPa"),
      ("R-1", "Local 36/37F water-cooled SCREW chiller — Ebara RHS DW202M2 (370 kW, COP 4.5, twin screw, R-407C), DHC-backup"),
      ("CDP", "Condenser-water pump (cooling-tower loop)"),
      ("CP / HP", "Chilled-water pump / hot-water (heating-loop) pump"),
      ("CT", "Cooling tower (INV fan cell + basin common points)"),
      ("HWT-1", "Hot-well / condensate tank"),
      ("PCU/PAC/PMAC", "Packaged (water-cooled) air-conditioners — status/alarm on BMS"),
      ("OAU (EVU)", "Outdoor-air / make-up-air unit"),
      ("順序 / 群起動", "Duty sequence / group-start staging of like pumps"),
  ]
  for row in legend:
      ws.append(row)
  style_header(ws, 2)
  set_widths(ws, [22, 95])
  for rr in range(2, ws.max_row + 1):
      ws.cell(row=rr, column=1).font = Font(bold=True, size=10)
      ws.cell(row=rr, column=2).alignment = Alignment(wrap_text=True, vertical="top")
      for c in (1, 2):
          ws.cell(row=rr, column=c).border = BORDER

  wb.save(OUT)
  print("wrote", OUT)
  print("total I/O points:", len(rows))
  print("devices:", len({r['Device ID'] for r in rows}),
        "| panels:", len(PANELS), "| controllers:", len(controllers))


if __name__ == "__main__":
    build_workbook()
